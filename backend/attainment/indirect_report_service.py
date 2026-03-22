import io
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
from django.db.models import Avg
from django.db import models
from surveys.models import SurveyMaster, SurveyQuestion, SurveyAnswer
from academics.models import Program, PO, PSO, Batch
import re

class IndirectReportService:
    # Colors and Styles
    HEADER_BLUE = "1a237e"
    HEADER_TEXT = "ffffff"
    BORDER_COLOR = "000000"
    
    @staticmethod
    def _get_styles():
        border = Border(
            left=Side(style='thin', color="000000"),
            right=Side(style='thin', color="000000"),
            top=Side(style='thin', color="000000"),
            bottom=Side(style='thin', color="000000")
        )
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        header_font = Font(bold=True, size=12, color="ffffff")
        header_fill = PatternFill(start_color="1a237e", end_color="1a237e", fill_type="solid")
        title_font = Font(bold=True, size=14)
        bold_font = Font(bold=True)
        return border, center_align, header_font, header_fill, title_font, bold_font

    @staticmethod
    def _add_company_header(ws, title, program_name, batch_name, col_count):
        _, center_align, _, _, title_font, _ = IndirectReportService._get_styles()
        
        # Row 1-4 Meta Info
        rows = [
            "Sandip Foundation's",
            "Sandip Polytechnic, Nashik",
            f"Department of {program_name}",
            f"{title} for Batch {batch_name}"
        ]
        
        for i, text in enumerate(rows, 1):
            ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=col_count)
            cell = ws.cell(row=i, column=1, value=text)
            cell.alignment = center_align
            cell.font = title_font
        
        return 5 # Next available row

    @staticmethod
    def _sanitize_title(title):
        title = (title[:25] if title else "Sheet")
        for char in r"\/[]:*?":
            title = title.replace(char, "_")
        return title

    @staticmethod
    def _get_batch_years(batch_raw):
        """Helper to get list of years corresponding to a batch. The frontend sends the graduation year (TY)."""
        years = []
        try:
            if isinstance(batch_raw, Batch):
                grad_start = batch_raw.batch_year
            else:
                grad_start = int(str(batch_raw).split('-')[0].strip())
            
            for i in range(3): # FY, SY, TY
                y = grad_start - 2 + i
                next_yr_short = (y + 1) % 100
                years.append(f"{y} - {next_yr_short:02d}")
                years.append(f"{y}-{next_yr_short:02d}")
            return years
        except Exception:
            return []

    @staticmethod
    def _get_survey_category(survey):
        """Determines a human-readable category for a survey based on activity_type or name."""
        atype = (survey.activity_type or "").strip()
        name = (survey.survey_name or "").strip()
        
        # Resource Person check
        if "RESOURCE" in atype.upper() or "RESOURCE" in name.upper() or "RP" in atype.upper():
            return "Resource Person Feedback"
        # Alumni check
        if "ALUMNI" in atype.upper() or "ALUMNI" in name.upper():
            return "Alumni Feedback"
        # Exit check
        if "EXIT" in atype.upper() or "EXIT" in name.upper():
            return "Programme Exit Feedback"
        # Industry Visit check
        if any(x in atype.upper() for x in ["IV", "VISIT"]) or "VISIT" in name.upper():
            return "Industry Visit Feedback"
        # Expert Lecture
        if any(x in atype.upper() for x in ["EL", "LECTURE"]) or "LECTURE" in name.upper():
            return "Expert Lecture Feedback"
        # VAP
        if "VAP" in atype.upper() or "VAP" in name.upper() or "VALUE ADDED" in name.upper():
            return "Value Added Program Feedback"
            
        if name and (not atype or atype.lower() in ['other', 'general', 'misc', 'na']):
            return name
        return atype.capitalize() if atype else (name if name else "Other Feedback")

    @staticmethod
    def generate_indirect_attainment_report(program_id, batch_id):
        wb = Workbook()
        # Remove default sheet
        default_sheet = wb.active
        wb.remove(default_sheet)
        
        program = Program.objects.get(program_id=program_id)
        if isinstance(batch_id, Batch):
            batch = batch_id
        else:
            from attainment.views import resolve_batch
            batch = resolve_batch(batch_id)
            if not batch:
                 raise ValueError(f"Batch {batch_id} not found")
        
        pos = list(PO.objects.filter(program_id=program_id, is_active=True).order_by('po_number'))
        psos = list(PSO.objects.filter(program_id=program_id, is_active=True).order_by('pso_number'))
        outcomes = pos + psos
        num_cols = len(outcomes)
        
        batch_years = IndirectReportService._get_batch_years(batch)
        
        # 1. Fetch ALL relevant surveys (be broad)
        all_surveys = SurveyMaster.objects.filter(
            academic_year__in=batch_years
        ).filter(
            models.Q(program_id=program) | models.Q(program_id__isnull=True)
        ).filter(
            survey_category__in=['indirect']
        ).exclude(status='DRAFT').order_by('-academic_year', '-survey_id')

        # Group surveys by category
        category_map = {} # label -> [surveys]
        for s in all_surveys:
            cat_label = IndirectReportService._get_survey_category(s)
            if cat_label not in category_map:
                category_map[cat_label] = []
            category_map[cat_label].append(s)

        border, center_align, header_font, header_fill, title_font, bold_font = IndirectReportService._get_styles()

        # 1. Overall Summary Sheet (Reverted to 4 categories)
        ws_sum = wb.create_sheet(title="Overall Summary")
        next_row = IndirectReportService._add_company_header(ws_sum, "Overall Indirect Attainment", program.program_name, batch, num_cols + 2)
        
        sum_headers = ["Sr. no.", "Feedback Category"] + [outcome.po_number if isinstance(outcome, PO) else outcome.pso_number for outcome in outcomes]
        for c, h in enumerate(sum_headers, 1):
            cell = ws_sum.cell(row=next_row, column=c, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = border
        next_row += 1
        
        # Restore original 4 rows for summary
        fixed_categories = [
            ("Overall Curricular and Extra Curricular Activities Feedback", ["Expert Lecture Feedback", "Industry Visit Feedback", "Value Added Program Feedback"]),
            ("Programme Exit Feedback", ["Programme Exit Feedback"]),
            ("Feedback from Alumni", ["Alumni Feedback"]),
            ("Overall Activities Resource person feedback", ["Resource Person Feedback"])
        ]
        
        row_avg_vals = []
        for idx, (label, dynamic_cats) in enumerate(fixed_categories, 1):
            ws_sum.cell(row=next_row, column=1, value=idx).border = border
            ws_sum.cell(row=next_row, column=2, value=label).border = border
            
            # Combine all surveys from matched dynamic categories
            surveys_for_row = []
            for dcat in dynamic_cats:
                surveys_for_row.extend(category_map.get(dcat, []))
            
            outcome_avgs = []
            for col_idx, outcome in enumerate(outcomes, 3):
                survey_avgs = []
                for s in surveys_for_row:
                    ans_filter = {'question_id__survey_id': s.survey_id}
                    if isinstance(outcome, PO): ans_filter['question_id__po_id'] = outcome
                    else: ans_filter['question_id__pso_id'] = outcome
                    
                    s_avg = SurveyAnswer.objects.filter(**ans_filter).aggregate(Avg('answer_value'))['answer_value__avg']
                    if s_avg is not None: survey_avgs.append(s_avg)
                
                row_val = round(sum(survey_avgs) / len(survey_avgs), 2) if survey_avgs else 0
                cell = ws_sum.cell(row=next_row, column=col_idx, value=row_val if row_val > 0 else "")
                cell.border = border
                cell.alignment = center_align
                outcome_avgs.append(row_val)
            
            row_avg_vals.append(outcome_avgs)
            next_row += 1
            
        # Average row
        ws_sum.cell(row=next_row, column=1).border = border
        ws_sum.cell(row=next_row, column=2, value="AVERAGE").border = border
        ws_sum.cell(row=next_row, column=2).font = bold_font
        for col_idx in range(3, num_cols + 3):
            col_vals = [r[col_idx-3] for r in row_avg_vals if r[col_idx-3] > 0]
            avg_val = round(sum(col_vals) / len(col_vals), 2) if col_vals else 0
            cell = ws_sum.cell(row=next_row, column=col_idx, value=avg_val if avg_val > 0 else "")
            cell.font = bold_font
            cell.border = border
            cell.alignment = center_align
        
        ws_sum.column_dimensions['B'].width = 60

        # 2. Categorywise Averages Sheet (New)
        ws_cat_sum = wb.create_sheet(title="Categorywise Averages")
        next_row = IndirectReportService._add_company_header(ws_cat_sum, "Categorywise Indirect Averages", program.program_name, batch, num_cols + 2)
        
        for c, h in enumerate(sum_headers, 1):
            cell = ws_cat_sum.cell(row=next_row, column=c, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = border
        next_row += 1
        
        cat_row_avgs = []
        sorted_labels = sorted(category_map.keys())
        for idx, cat_label in enumerate(sorted_labels, 1):
            ws_cat_sum.cell(row=next_row, column=1, value=idx).border = border
            ws_cat_sum.cell(row=next_row, column=2, value=cat_label).border = border
            
            outcome_avgs = []
            for col_idx, outcome in enumerate(outcomes, 3):
                survey_avgs = []
                for s in category_map[cat_label]:
                    ans_filter = {'question_id__survey_id': s.survey_id}
                    if isinstance(outcome, PO): ans_filter['question_id__po_id'] = outcome
                    else: ans_filter['question_id__pso_id'] = outcome
                    avg = SurveyAnswer.objects.filter(**ans_filter).aggregate(Avg('answer_value'))['answer_value__avg']
                    if avg is not None: survey_avgs.append(avg)
                
                val = round(sum(survey_avgs)/len(survey_avgs), 2) if survey_avgs else 0
                cell = ws_cat_sum.cell(row=next_row, column=col_idx, value=val if val > 0 else "")
                cell.border = border
                cell.alignment = center_align
                outcome_avgs.append(val)
            cat_row_avgs.append(outcome_avgs)
            next_row += 1
        
        # Add Overall Average Row for Categorywise Sheet
        next_row += 1
        ws_cat_sum.cell(row=next_row, column=1).border = border
        ws_cat_sum.cell(row=next_row, column=2, value="AVERAGE").border = border
        ws_cat_sum.cell(row=next_row, column=2).font = bold_font
        for col_idx in range(3, num_cols + 3):
            col_vals = [r[col_idx-3] for r in cat_row_avgs if r[col_idx-3] > 0]
            val = round(sum(col_vals)/len(col_vals), 2) if col_vals else 0
            c = ws_cat_sum.cell(row=next_row, column=col_idx, value=val if val > 0 else "")
            c.border = border
            c.alignment = center_align
            c.font = bold_font
            
        ws_cat_sum.column_dimensions['B'].width = 60
        ws_cat_sum.column_dimensions['B'].width = 50

        # 3. Detailed Sheets (One for each category)
        for cat_label in sorted_labels:
            sheet_title = IndirectReportService._sanitize_title(cat_label)
            ws = wb.create_sheet(title=sheet_title)
            
            surveys = category_map[cat_label]
            curr_row = IndirectReportService._add_company_header(ws, cat_label, program.program_name, batch, num_cols + 2)
            
            for survey in surveys:
                ws.merge_cells(start_row=curr_row, start_column=1, end_row=curr_row, end_column=num_cols+2)
                date_str = f" | Date: {survey.conducted_date.strftime('%d-%m-%Y')}" if survey.conducted_date else ""
                cell = ws.cell(row=curr_row, column=1, value=f"{survey.survey_name} ({survey.academic_year}){date_str}")
                cell.font = bold_font
                curr_row += 1
                
                table_headers = ["Sr. no.", "Respondent Name"] + [outcome.po_number if isinstance(outcome, PO) else outcome.pso_number for outcome in outcomes]
                for c, h in enumerate(table_headers, 1):
                    cell = ws.cell(row=curr_row, column=c, value=h)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.border = border
                    cell.alignment = center_align
                curr_row += 1
                
                responses = survey.responses.all()
                if not responses.exists():
                    ws.merge_cells(start_row=curr_row, start_column=1, end_row=curr_row, end_column=num_cols+2)
                    ws.cell(row=curr_row, column=1, value="No responses submitted yet.").border = border
                    curr_row += 2
                    continue

                for idx, res in enumerate(responses, 1):
                    ws.cell(row=curr_row, column=1, value=idx).border = border
                    ws.cell(row=curr_row, column=2, value=res.respondent_name or f"Respondent {res.enrollment_no or idx}").border = border
                    
                    for col_idx, outcome in enumerate(outcomes, 3):
                        ans_query = {'response_id': res}
                        if isinstance(outcome, PO): ans_query['question_id__po_id'] = outcome
                        else: ans_query['question_id__pso_id'] = outcome
                        ans = SurveyAnswer.objects.filter(**ans_query).first()
                        val = ans.answer_value if ans else ""
                        cell = ws.cell(row=curr_row, column=col_idx, value=val)
                        cell.border = border
                        cell.alignment = center_align
                    curr_row += 1
                
                # Average per Survey
                ws.cell(row=curr_row, column=1).border = border
                ws.cell(row=curr_row, column=2, value="Average").font = bold_font
                ws.cell(row=curr_row, column=2).border = border
                for col_idx, outcome in enumerate(outcomes, 3):
                    ans_filter = {'question_id__survey_id': survey.survey_id}
                    if isinstance(outcome, PO): ans_filter['question_id__po_id'] = outcome
                    else: ans_filter['question_id__pso_id'] = outcome
                    avg = SurveyAnswer.objects.filter(**ans_filter).aggregate(Avg('answer_value'))['answer_value__avg']
                    val = round(avg, 2) if avg is not None else ""
                    cell = ws.cell(row=curr_row, column=col_idx, value=val)
                    cell.font = bold_font
                    cell.border = border
                curr_row += 2
            ws.column_dimensions['B'].width = 40

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    @staticmethod
    def get_indirect_attainment_summary_data(program_id, batch_id):
        program = Program.objects.get(program_id=program_id)
        pos = list(PO.objects.filter(program_id=program_id, is_active=True).order_by('po_number'))
        psos = list(PSO.objects.filter(program_id=program_id, is_active=True).order_by('pso_number'))
        outcomes = pos + psos
        
        if isinstance(batch_id, Batch): batch = batch_id
        else:
            from attainment.views import resolve_batch
            batch = resolve_batch(batch_id)
            if not batch: return []
            
        batch_years = IndirectReportService._get_batch_years(batch)
        all_surveys = SurveyMaster.objects.filter(
            academic_year__in=batch_years
        ).filter(
            models.Q(program_id=program) | models.Q(program_id__isnull=True)
        ).filter(survey_category__in=['indirect', 'feedback', 'course_exit']).order_by('-academic_year')

        category_map = {}
        for s in all_surveys:
            cat_label = IndirectReportService._get_survey_category(s)
            if cat_label not in category_map: category_map[cat_label] = []
            category_map[cat_label].append(s)
            
        summary = []
        for outcome in outcomes:
            category_vals = []
            for cat_label, surveys in category_map.items():
                survey_avgs = []
                for s in surveys:
                    ans_filter = {'question_id__survey_id': s.survey_id}
                    if isinstance(outcome, PO): ans_filter['question_id__po_id'] = outcome
                    else: ans_filter['question_id__pso_id'] = outcome
                    s_avg = SurveyAnswer.objects.filter(**ans_filter).aggregate(Avg('answer_value'))['answer_value__avg']
                    if s_avg is not None: survey_avgs.append(s_avg)
                if survey_avgs:
                    category_vals.append(sum(survey_avgs) / len(survey_avgs))
            
            outcome_final_avg = sum(category_vals) / len(category_vals) if category_vals else 0
            summary.append({
                'id': outcome.po_id if isinstance(outcome, PO) else outcome.pso_id,
                'type': 'PO' if isinstance(outcome, PO) else 'PSO',
                'number': outcome.po_number if isinstance(outcome, PO) else outcome.pso_number,
                'label': str(outcome.po_number) if isinstance(outcome, PO) else str(outcome.pso_number),
                'achieved': round(outcome_final_avg, 2)
            })
        return summary
