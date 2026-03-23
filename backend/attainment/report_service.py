import pandas as pd
import io
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from academics.models import Course, CO, PO, PSO, COPOMapping, COPSOMapping, Batch, Program
from attainment.models import COAttainment
from surveys.models import SurveyMaster, SurveyQuestion, SurveyAnswer
from django.db.models import Avg

class ReportService:
    @staticmethod
    def generate_batch_evaluation_report(program_id, batch_id):
        program = Program.objects.get(pk=program_id)
        if isinstance(batch_id, Batch):
            batch = batch_id
        else:
            # Fallback if ID/string passed
            from attainment.views import resolve_batch
            batch = resolve_batch(batch_id)
            if not batch:
                raise ValueError(f"Batch {batch_id} not found")
        
        # 1. Fetch all courses for this program and scheme
        courses = Course.objects.filter(
            program_id=program,
            scheme_id=batch.scheme_id,
            is_active=True
        ).order_by('semester', 'course_code')

        # 2. Prepare PO/PSO headers
        pos = PO.objects.filter(program_id=program, is_active=True).order_by('po_number')
        psos = PSO.objects.filter(program_id=program, is_active=True).order_by('pso_number')
        
        po_pso_headers = [p.po_number for p in pos] + [p.pso_number for p in psos]
        po_map_dict = {p.po_id: p.po_number for p in pos}
        pso_map_dict = {p.pso_id: p.pso_number for p in psos}
        po_ids = list(po_map_dict.keys())
        pso_ids = list(pso_map_dict.keys())

        # 3. Aggregate Data for Left Table (Mappings) and Right Table (Attainment)
        mapping_avg_data = []
        evaluation_data = []
        
        for course in courses:
            # Course info
            course_display_name = f"{course.course_code} {course.course_name}"
            
            # --- Left Table: Mapping Averages ---
            # Group by PO/PSO and average the weights
            row_mappings = {h: 0.0 for h in po_pso_headers}
            
            # PO Mappings
            po_maps = COPOMapping.objects.filter(co_id__course_id=course, weightage__gt=0)
            for po_id in po_ids:
                matches = [m.weightage for m in po_maps if m.po_id_id == po_id]
                if matches:
                    avg_w = sum(matches) / len(matches)
                    row_mappings[po_map_dict[po_id]] = round(avg_w, 2)
            
            # PSO Mappings
            pso_maps = COPSOMapping.objects.filter(co_id__course_id=course, weightage__gt=0)
            for pso_id in pso_ids:
                matches = [m.weightage for m in pso_maps if m.pso_id_id == pso_id]
                if matches:
                    avg_w = sum(matches) / len(matches)
                    row_mappings[pso_map_dict[pso_id]] = round(avg_w, 2)
            
            mapping_avg_data.append({
                'Course Name': course_display_name,
                **row_mappings
            })

            # --- Right Table: Result of Evaluation ---
            # Fetch direct course attainment (average of CO attainments)
            # Note: We need to decide which Academic Year to use for each course.
            # In a batch report, each semester happened in a different AY.
            # E.g. 2021-22 Sem1, 2022-23 Sem3...
            # For simplicity, we'll fetch the latest overall_attainment for each course.
            # Realistically, we should filter by the AYs covered by this batch.
            
            co_atts = COAttainment.objects.filter(course_id=course)
            if co_atts.exists():
                direct_att = co_atts.aggregate(Avg('overall_attainment'))['overall_attainment__avg'] or 0.0
            else:
                direct_att = 0.0
            
            # Attained scores = Mapping * Attainment / 3
            row_evaluation = {h: 0.0 for h in po_pso_headers}
            for h in po_pso_headers:
                row_evaluation[h] = round((row_mappings[h] * direct_att) / 3, 2)
            
            evaluation_data.append({
                'Attainment': round(direct_att, 2),
                'Course Name': course_display_name,
                **row_evaluation
            })

        # --- Footer Logic ---
        # Average of evaluation rows
        eval_df = pd.DataFrame(evaluation_data)
        avg_row = {}
        for h in po_pso_headers:
            avg_row[h] = round(eval_df[h].mean(), 2) if not eval_df.empty else 0.0
        
        # Indirect Attainment (from surveys)
        # Category='indirect' surveys for this program and cohort years
        indirect_atts = {h: 3.0 for h in po_pso_headers} # Default 3.0
        
        from attainment.indirect_report_service import IndirectReportService
        batch_years = IndirectReportService._get_batch_years(batch)

        for po in pos:
            ans_avg = SurveyAnswer.objects.filter(
                question_id__survey_id__academic_year__in=batch_years,
                question_id__survey_id__program_id=program,
                question_id__survey_id__survey_category='indirect',
                question_id__po_id=po
            ).aggregate(Avg('answer_value'))['answer_value__avg']
            if ans_avg:
                indirect_atts[po.po_number] = round(ans_avg, 2)
        
        for pso in psos:
            ans_avg = SurveyAnswer.objects.filter(
                question_id__survey_id__academic_year__in=batch_years,
                question_id__survey_id__program_id=program,
                question_id__survey_id__survey_category='indirect',
                question_id__pso_id=pso
            ).aggregate(Avg('answer_value'))['answer_value__avg']
            if ans_avg:
                indirect_atts[pso.pso_number] = round(ans_avg, 2)

        # Final calc
        # Final PO attainment = 0.8*direct attainment + 0.2*indirect attainment
        # In the context of the bottom row, "direct attainment" is the average level across courses
        final_attainment = {}
        for h in po_pso_headers:
            final_attainment[h] = round(0.8 * avg_row[h] + 0.2 * indirect_atts[h], 2)

        # 4. Excel Generation with openpyxl for formatting
        wb = Workbook()
        ws = wb.active
        ws.title = "Result of Evaluation"
        
        # Styles
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        header_fill_tan = PatternFill(start_color="F9E4B7", end_color="F9E4B7", fill_type="solid") # Tan for left
        header_fill_blue = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid") # Changed to match image's green-ish/blue
        header_fill_dark_blue = PatternFill(start_color="CFE2F3", end_color="CFE2F3", fill_type="solid")
        
        bold_font = Font(bold=True)
        center_align = Alignment(horizontal='center', vertical='center')

        # HEADERS
        # Row 1: Titles
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(po_pso_headers)+1)
        ws.cell(row=1, column=1, value="CO-PO-PSO Mapping Average").font = bold_font
        ws.cell(row=1, column=1).alignment = center_align
        ws.cell(row=1, column=1).fill = header_fill_tan

        ws.merge_cells(start_row=1, start_column=len(po_pso_headers)+3, end_row=1, end_column=len(po_pso_headers)*2 + 4)
        ws.cell(row=1, column=len(po_pso_headers)+3, value="Result of Evaluation").font = bold_font
        ws.cell(row=1, column=len(po_pso_headers)+3).alignment = center_align
        ws.cell(row=1, column=len(po_pso_headers)+3).fill = header_fill_dark_blue

        # Row 2: Subtitle for Right table
        ws.merge_cells(start_row=2, start_column=len(po_pso_headers)+3, end_row=2, end_column=len(po_pso_headers)*2 + 4)
        # Display as graduation academic year (e.g. 2025-26)
        batch_label = f"{batch.end_year-1}-{str(batch.end_year)[-2:]}"
        ws.cell(row=2, column=len(po_pso_headers)+3, value=f"Batch {batch_label} (PO and PSO attainment)").font = bold_font
        ws.cell(row=2, column=len(po_pso_headers)+3).alignment = center_align

        # Row 3: Column Headers
        ws.cell(row=3, column=1, value="Course Name").fill = header_fill_tan
        for i, h in enumerate(po_pso_headers):
            ws.cell(row=3, column=i+2, value=h).fill = header_fill_tan
        
        ws.cell(row=3, column=len(po_pso_headers)+3, value="Attainment").fill = header_fill_dark_blue
        ws.cell(row=3, column=len(po_pso_headers)+4, value="Course Name").fill = header_fill_dark_blue
        for i, h in enumerate(po_pso_headers):
            ws.cell(row=3, column=len(po_pso_headers)+5+i, value=h).fill = header_fill_dark_blue

        # DATA ROWS
        current_row = 4
        for i in range(len(mapping_avg_data)):
            # Left Table
            ws.cell(row=current_row, column=1, value=mapping_avg_data[i]['Course Name'])
            for j, h in enumerate(po_pso_headers):
                ws.cell(row=current_row, column=j+2, value=mapping_avg_data[i][h])
            
            # Right Table
            ws.cell(row=current_row, column=len(po_pso_headers)+3, value=evaluation_data[i]['Attainment'])
            ws.cell(row=current_row, column=len(po_pso_headers)+4, value=evaluation_data[i]['Course Name'])
            for j, h in enumerate(po_pso_headers):
                ws.cell(row=current_row, column=len(po_pso_headers)+5+j, value=evaluation_data[i][h])
            
            current_row += 1

        # FOOTERS (Right Table)
        # Average
        ws.cell(row=current_row, column=len(po_pso_headers)+4, value="Average").font = bold_font
        for i, h in enumerate(po_pso_headers):
            ws.cell(row=current_row, column=len(po_pso_headers)+5+i, value=avg_row[h]).font = Font(color="7030A0", bold=True)
        current_row += 1
        
        # Direct Attainment
        direct_cell = ws.cell(row=current_row, column=len(po_pso_headers)+4, value="Direct Attainment")
        direct_cell.font = Font(color="FF0000", bold=True)
        for i, h in enumerate(po_pso_headers):
            ws.cell(row=current_row, column=len(po_pso_headers)+5+i, value=avg_row[h]).font = bold_font
        current_row += 1

        # 80% of Direct Attainment
        da80_cell = ws.cell(row=current_row, column=len(po_pso_headers)+4, value="80% of Direct Attainment")
        da80_cell.font = Font(color="FF0000", bold=True)
        for i, h in enumerate(po_pso_headers):
            ws.cell(row=current_row, column=len(po_pso_headers)+5+i, value=round(avg_row[h] * 0.8, 2)).font = bold_font
        current_row += 1

        # Spacer row
        current_row += 1

        # Indirect Attainment
        ia_cell = ws.cell(row=current_row, column=len(po_pso_headers)+4, value="Indirect Attainment")
        ia_cell.font = Font(color="FF0000", bold=True)
        for i, h in enumerate(po_pso_headers):
            ws.cell(row=current_row, column=len(po_pso_headers)+5+i, value=indirect_atts[h]).font = bold_font
        current_row += 1

        # 20% of Indirect attainment
        ia20_cell = ws.cell(row=current_row, column=len(po_pso_headers)+4, value="20% of Indirect attainment")
        ia20_cell.font = Font(color="FF0000", bold=True)
        for i, h in enumerate(po_pso_headers):
            ws.cell(row=current_row, column=len(po_pso_headers)+5+i, value=round(indirect_atts[h] * 0.2, 2)).font = bold_font
        current_row += 1

        # Spacer
        current_row += 1

        # PO Attainment
        po_cell = ws.cell(row=current_row, column=len(po_pso_headers)+4, value="PO Attainment")
        po_cell.font = Font(color="0000FF", bold=True)
        for i, h in enumerate(po_pso_headers):
            ws.cell(row=current_row, column=len(po_pso_headers)+5+i, value=final_attainment[h]).font = bold_font
        
        # Borders for all cells
        for row in ws.rows:
            for cell in row:
                cell.border = thin_border

        # Save to buffer
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    @staticmethod
    def generate_course_attainment_report(course_id, academic_year):
        from academics.models import Course, CO
        from attainment.attainment_service import AttainmentService
        from .models import COAttainment, CourseATR, COTarget
        from django.db.models import Q

        course = Course.objects.get(pk=course_id)

        ay_clean = academic_year.replace(' ', '') if academic_year else ""
        ay_spaced = ay_clean.replace('-', ' - ')
        ay_query = Q(academic_year__icontains=academic_year) | Q(academic_year__icontains=ay_clean) | Q(academic_year__icontains=ay_spaced)

        # Fetch tool-wise attainment per CO
        tool_data = AttainmentService._calculate_detailed_tool_attainment(course_id, academic_year)
        all_cos = CO.objects.filter(course_id=course_id, is_active=True).order_by('co_number')

        # Build per-CO rows
        rows = []
        for co in all_cos:
            tools = tool_data.get(co.co_id, {})

            def get_level(key, prefix=None):
                if prefix:
                    k = f"{prefix}_{key}"
                    if k in tools:
                        v = tools[k]
                        return v.get('level', '-') if isinstance(v, dict) else (v if v is not None else '-')
                    return '-'
                
                if key in tools:
                    v = tools[key]
                    return v.get('level', '-') if isinstance(v, dict) else (v if v is not None else '-')
                
                for full_key, v in tools.items():
                    if full_key.endswith(f"_{key}"):
                        if isinstance(v, dict):
                            return v.get('level', '-')
                        return v if v is not None else '-'
                return '-'

            # Use dynamic category-aware lookups
            ct1      = get_level('FA_TH_1', 'INTERNAL')
            ct2      = get_level('FA_TH_2', 'INTERNAL')
            sla      = get_level('SLA')
            fa_pr    = get_level('FA_PR', 'INTERNAL')
            sa_pr_i  = get_level('SA_PR', 'INTERNAL')
            sa_th    = get_level('SA_TH', 'EXTERNAL')
            sa_pr_e  = get_level('SA_PR', 'EXTERNAL')

            # Redefine logic for Avg(I) and Avg(E) based on tool dictionary keys
            internal_vals = [v['level'] for k, v in tools.items() if k.startswith('INTERNAL_') and isinstance(v, dict)]
            avg_i = round(sum(internal_vals) / len(internal_vals), 2) if internal_vals else '-'

            external_vals = [v['level'] for k, v in tools.items() if k.startswith('EXTERNAL_') and isinstance(v, dict)]
            avg_e = round(sum(external_vals) / len(external_vals), 2) if external_vals else '-'

            if isinstance(avg_i, float) and isinstance(avg_e, float):
                avg_b = round(0.4 * avg_i + 0.6 * avg_e, 2)
            elif isinstance(avg_i, float):
                avg_b = avg_i
            elif isinstance(avg_e, float):
                avg_b = avg_e
            else:
                avg_b = '-'

            # Target & gap
            co_att = COAttainment.objects.filter(ay_query, co_id=co).first()
            target_obj = COTarget.objects.filter(ay_query, co_id=co).first()
            target = round(target_obj.target_value, 2) if target_obj else 3.0
            overall = round(co_att.overall_attainment, 2) if co_att else (avg_b if isinstance(avg_b, float) else 0)
            gap = round(target - overall, 2) if isinstance(overall, float) else '-'

            rows.append({
                'co_number': co.co_number,
                'ct1': ct1, 'ct2': ct2, 'sla': sla, 'fa_pr': fa_pr,
                'sa_pr_i': sa_pr_i, 'avg_i': avg_i,
                'sa_th': sa_th, 'sa_pr_e': sa_pr_e,
                'avg_b': avg_b,
                'target': target, 'gap': gap
            })

        # --- Excel Layout ---
        wb = Workbook()
        ws = wb.active
        ws.title = "CO Attainment"

        thin = Border(left=Side(style='thin'), right=Side(style='thin'),
                      top=Side(style='thin'), bottom=Side(style='thin'))
        bold = Font(bold=True)
        center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

        yellow_fill  = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # Internal 40%
        blue_fill    = PatternFill(start_color="DDEEFF", end_color="DDEEFF", fill_type="solid")  # External 60%
        header_fill  = PatternFill(start_color="C9DAF8", end_color="C9DAF8", fill_type="solid")
        title_fill   = PatternFill(start_color="A9C4F5", end_color="A9C4F5", fill_type="solid")
        gap_red_fill = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")
        ok_green     = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")

        def bc(row, col, val='', fill=None, font=None, align=center, border=thin):
            cell = ws.cell(row=row, column=col, value=val)
            if fill:  cell.fill  = fill
            if font:  cell.font  = font
            cell.alignment = align
            cell.border = border
            return cell

        # Row 1: Meta
        meta = [
            (1, 'Academic Year', academic_year),
            (2, 'Semester', ''),
            (3, 'Scheme', ''),
            (4, 'Name of Faculty', ''),
            (5, 'Name of Course & Code', f"{course.course_name} {course.course_code}"),
            (6, 'Class & Division', ''),
        ]
        for r_off, label, val in meta:
            bc(r_off, 1, label, font=bold, align=left_align)
            ws.merge_cells(start_row=r_off, start_column=2, end_row=r_off, end_column=4)
            bc(r_off, 2, val, align=left_align)

        # Row 8: Title
        TITLE_ROW = 8
        ws.merge_cells(start_row=TITLE_ROW, start_column=1, end_row=TITLE_ROW, end_column=11)
        bc(TITLE_ROW, 1, "CO Attainment", fill=title_fill, font=Font(bold=True, size=13), align=center)

        # Row 9: Section headers
        SECT_ROW = 9
        ws.merge_cells(start_row=SECT_ROW, start_column=2, end_row=SECT_ROW, end_column=7)
        bc(SECT_ROW, 2, "Internal Assessment 40%", fill=yellow_fill, font=bold, align=center)
        ws.merge_cells(start_row=SECT_ROW, start_column=8, end_row=SECT_ROW, end_column=11)
        bc(SECT_ROW, 8, "External Assessment 60%", fill=blue_fill, font=bold, align=center)

        # Row 10: Column headers
        HDR_ROW = 10
        headers = ['CO', 'CT1', 'CT2', 'Assignment (SLA)', 'FA-PR', 'SA-PR\n(Internal)', 'Avg(I)',
                   'SA-TH', 'SA-PR\n(External)', 'Avg(B)', 'Target / Gap']
        fills   = [header_fill, yellow_fill, yellow_fill, yellow_fill, yellow_fill, yellow_fill, yellow_fill,
                   blue_fill, blue_fill, blue_fill, header_fill]
        for i, (h, f) in enumerate(zip(headers, fills)):
            bc(HDR_ROW, i+1, h, fill=f, font=bold, align=center)

        # Data rows
        DATA_START = 11
        for r_i, row in enumerate(rows):
            r = DATA_START + r_i
            vals = [row['co_number'], row['ct1'], row['ct2'], row['sla'], row['fa_pr'],
                    row['sa_pr_i'], row['avg_i'], row['sa_th'], row['sa_pr_e'], row['avg_b'],
                    f"{row['avg_b']} / {row['gap']}" if isinstance(row['avg_b'], float) else '-']
            row_fills = [None, yellow_fill, yellow_fill, yellow_fill, yellow_fill, yellow_fill, yellow_fill,
                         blue_fill, blue_fill, blue_fill, None]
            for c_i, (v, f) in enumerate(zip(vals, row_fills)):
                cell = bc(r, c_i+1, v, fill=f)
                # Highlight gap column
                if c_i == 10:
                    if isinstance(row['gap'], float) and row['gap'] > 0:
                        cell.fill = gap_red_fill
                    elif isinstance(row['gap'], float):
                        cell.fill = ok_green

        # Footer: ATR
        c_atr = CourseATR.objects.filter(course_id=course_id, academic_year=academic_year).first()
        if c_atr:
            atr_row = DATA_START + len(rows) + 2
            ws.merge_cells(start_row=atr_row, start_column=1, end_row=atr_row, end_column=11)
            bc(atr_row, 1, "ACTION TAKEN REPORT (ATR)", fill=title_fill, font=bold)
            ws.merge_cells(start_row=atr_row+1, start_column=1, end_row=atr_row+4, end_column=11)
            bc(atr_row+1, 1, c_atr.action_proposed, align=Alignment(wrap_text=True, vertical='top'))

        # Footer note
        note_row = DATA_START + len(rows) + (7 if c_atr else 2)
        ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=11)
        bc(note_row, 1, "CT1 & CT2 : Class Test 1 & 2   SLA : Self Learning Assessments   FA PR : Formative Assessment Practical",
           font=Font(italic=True, size=9), align=left_align)

        # Column widths
        col_widths = [14, 7, 7, 18, 9, 12, 9, 9, 12, 9, 14]
        for idx, w in enumerate(col_widths):
            ws.column_dimensions[ws.cell(row=1, column=idx+1).column_letter].width = w

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output
