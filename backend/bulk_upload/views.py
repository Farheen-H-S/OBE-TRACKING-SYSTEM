from django.shortcuts import get_object_or_404
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from rest_framework.parsers import MultiPartParser, FormParser
import csv
import io
from users.models import Student, User, UserRole
from academics.models import Program, Batch, Course, CO, Scheme
from assessments.models import Assessment, MarksEntry, AssessmentCOMapping
from attainment.attainment_service import AttainmentService
from teaching_plan.models import TeachingPlan, TeachingPlanLecture
from django.db import transaction
import re
import openpyxl
from openpyxl.utils import get_column_letter

from django.db.models import Q
import pandas as pd
from django.db import IntegrityError
from django.http import HttpResponse
import os
from django.conf import settings
from notifications.utils import send_obe_notification
from .bulk_cis_service import generate_cis_multi_sheet_template, process_bulk_cis_apply

def natural_sort_key(s):
    return [int(text) if text.isdigit() else text.lower()
            for text in re.split('([0-9]+)', str(s))]

class DownloadStudentTemplateView(APIView):
    """
    Serves the pre-generated Excel template for student bulk upload.
    """
    def get(self, request):
        file_path = settings.BASE_DIR / 'bulk_upload' / 'template' / 'Student_Bulk_Upload_Template.xlsx'
        if not os.path.exists(file_path):
            return Response({"error": "Template file not found. Please contact administrator."}, status=404)
        
        with open(file_path, 'rb') as f:
            response = HttpResponse(f.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename=Student_Bulk_Upload_Template.xlsx'
            return response

class DownloadCourseTemplateView(APIView):
    """
    Generates and serves an Excel template for course bulk upload.
    """
    def get(self, request):
        cols = [
            'Program Name', 'Scheme Name', 'Course Code', 'Course Name', 
            'Course Title', 'Course Abbr', 'Semester', 'Class Year'
        ]
        df = pd.DataFrame(columns=cols)
        
        # Add a sample row
        df.loc[0] = ['Computer Engineering', 'Scheme 2023', '22101', 'English', 'Communication Skills', 'ENG', '1', 'FY']
        
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Courses')
        output.seek(0)
        
        response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=Marks_Bulk_Upload_Template.xlsx'
        return response

class DownloadCISTemplateView(APIView):
    """
    Generates a pre-filled Excel template for CIS marks based on the course and tool.
    Matches the Marks Entry UI exactly: Enrollment, Roll No, Name, Label (Q/Wt/CO), Questions.
    """
    def get(self, request):
        course_id = request.query_params.get('course_id')
        tool_name = request.query_params.get('tool_name', 'CT-1')
        
        if not course_id:
            return Response({"error": "Course ID is required"}, status=400)
            
        course = get_object_or_404(Course, pk=course_id)
        students = list(Student.objects.filter(program_id=course.program_id, is_active=True))
        students.sort(key=lambda x: natural_sort_key(x.roll_no or ""))
        
        assessment_config = course.assessment_tools or {}
        tool_data = assessment_config.get(tool_name, {})
        col_count = tool_data.get('columnCount', tool_data.get('questionCount', 6))
        
        output = io.BytesIO()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "CIS Marks Entry"
        
        # Column Headers - Labels are row-specific for the 4th column
        ws['A1'] = "ENROLLMENT NO"
        ws['B1'] = "Roll No."
        ws['C1'] = "Name of Student"
        
        # Merge A-C for the 3 header rows
        for col_char in ['A', 'B', 'C']:
            ws.merge_cells(f'{col_char}1:{col_char}3')
            cell = ws[f'{col_char}1']
            apply_header_style(cell, fill_color="2F5597", font_color="FFFFFF")

        # Row 1-3 Labels in Column D
        ws['D1'] = "Q"
        ws['D2'] = "Wt"
        ws['D3'] = "CO"
        for r in [1, 2, 3]:
            apply_header_style(ws.cell(row=r, column=4), fill_color="2F5597", font_color="FFFFFF")

        # Questions (Row 1), Weights (Row 2), COs (Row 3) Starting from Col E (5)
        for i in range(col_count):
            col_idx = 5 + i
            ws.cell(row=1, column=col_idx, value=f"Q{i+1}")
            ws.cell(row=2, column=col_idx, value=tool_data.get('customWeights', [5]*20)[i] if 'customWeights' in tool_data else 5)
            ws.cell(row=3, column=col_idx, value=tool_data.get('userCos', [f"CO{(i%6)+1}"]*20)[i] if 'userCos' in tool_data else f"CO{(i%6)+1}")
            
            apply_header_style(ws.cell(row=1, column=col_idx), fill_color="DEEBF7", font_color="000000")
            apply_header_style(ws.cell(row=2, column=col_idx), fill_color="DAE3F3", font_color="000000")
            apply_header_style(ws.cell(row=3, column=col_idx), fill_color="DEEBF7", font_color="000000")

        # Add Students starting row 4
        for idx, student in enumerate(students, start=4):
            ws.cell(row=idx, column=1, value=student.enrollment_no)
            ws.cell(row=idx, column=2, value=student.roll_no)
            ws.cell(row=idx, column=3, value=student.name)
            # Empty Label in Col D for students
            ws.cell(row=idx, column=4, value="") 
            
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=Marks_Bulk_Upload_Template.xlsx'
        return response

class DownloadTeachingPlanTemplateView(APIView):
    """
    Generates and serves an Excel template for Teaching Plan.
    Headers: Date, Unit, Topic, Description
    """
    def get(self, request):
        cols = ['Date', 'Unit', 'Topic', 'Description']
        df = pd.DataFrame(columns=cols)
        
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='TeachingPlan')
            
            # Auto-adjust column width for better readability
            worksheet = writer.sheets['TeachingPlan']
            for i, col in enumerate(cols):
                column_letter = get_column_letter(i + 1)
                worksheet.column_dimensions[column_letter].width = 20
        
        output.seek(0)
        
        response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=Teaching_Plan_Template.xlsx'
        return response

class UploadTeachingPlanExcelView(APIView):
    """
    Processes an Excel file to populate a Teaching Plan.
    Headers: Date, Unit, Topic, Description
    """
    parser_classes = (MultiPartParser, FormParser)

    def post(self, request):
        file_obj = request.FILES.get('file')
        course_id = request.data.get('course_id')
        academic_year = request.data.get('academic_year')
        semester = request.data.get('semester')
        scheme_id = request.data.get('scheme_id')
        batch_id = request.data.get('batch_id')

        if not all([file_obj, course_id, academic_year, semester, scheme_id]):
            return Response({"error": "Missing required fields (file, course_id, academic_year, semester, scheme_id)"}, status=400)

        try:
            course = get_object_or_404(Course, pk=course_id)
            academic_year = str(academic_year).replace(" ", "")

            with transaction.atomic():
                # Resolve Batch
                res_batch = None
                if batch_id and str(batch_id).isdigit():
                    res_batch = Batch.objects.filter(pk=batch_id).first()
                
                if not res_batch:
                    res_batch = Batch.objects.filter(scheme_id_id=scheme_id).first()
                
                if not res_batch:
                    return Response({"error": "No valid batch found for this scheme. Please select a batch."}, status=400)

                plan, created = TeachingPlan.objects.get_or_create(
                    course_id=course,
                    academic_year=academic_year,
                    semester=int(semester),
                    scheme_id_id=scheme_id,
                    defaults={
                        'user_id': request.user if request.user.is_authenticated else User.objects.first(),
                        'batch_id': res_batch,
                        'is_active': True
                    }
                )

                df = pd.read_excel(file_obj)
                df.columns = [str(c).strip().lower() for c in df.columns]
                
                required_cols = ['date', 'unit', 'topic', 'description']
                missing_cols = [c for c in required_cols if c not in df.columns]
                if missing_cols:
                    return Response({"error": f"Missing columns in Excel: {', '.join(missing_cols)}"}, status=400)

                plan.lectures.all().delete()

                lectures_to_create = []
                for index, row in df.iterrows():
                    if pd.isnull(row['topic']) and pd.isnull(row['date']):
                        continue
                    
                    l_date = row['date']
                    if isinstance(l_date, str):
                        try:
                            l_date = pd.to_datetime(l_date).date()
                        except:
                            l_date = None
                    elif hasattr(l_date, 'date'):
                        l_date = l_date.date()
                    
                    if pd.isnull(l_date):
                        continue

                    lectures_to_create.append(TeachingPlanLecture(
                        teaching_plan_id=plan,
                        lecture_no=index + 1,
                        lecture_date=l_date,
                        unit_no=int(row['unit']) if pd.notnull(row['unit']) and str(row['unit']).isdigit() else 1,
                        topic_planned=str(row['topic']) if pd.notnull(row['topic']) else "",
                        actual_topic=str(row['description']) if pd.notnull(row['description']) else "",
                        remark=""
                    ))
                
                TeachingPlanLecture.objects.bulk_create(lectures_to_create)

            return Response({"message": f"Successfully uploaded {len(lectures_to_create)} lectures."}, status=201)

        except Exception as e:
            return Response({"error": f"Upload Failed: {str(e)}"}, status=500)

def apply_header_style(cell, fill_color="2F5597", font_color="FFFFFF"):
    """Helper for template styling"""
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    cell.font = Font(bold=True, color=font_color)
    cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

class BulkStudentUploadView(APIView):
    parser_classes = (MultiPartParser, FormParser)

    def post(self, request):
        file_obj = request.FILES.get('file')
        if not file_obj:
            return Response({"error": "No file uploaded"}, status=status.HTTP_400_BAD_REQUEST)

        # Get defaults from request data
        default_batch_id = request.data.get('batch_id')
        default_academic_year = request.data.get('academic_year')
        default_semester = request.data.get('semester')
        default_class_year = request.data.get('class_year')
        default_division = request.data.get('division')
        default_program_id = request.data.get('program_id')

        # --- PRE-VALIDATION ---
        if not str(default_batch_id).strip(): default_batch_id = None
        if not str(default_program_id).strip(): default_program_id = None

        if not (file_obj.name.endswith('.xlsx') or file_obj.name.endswith('.xls')):
            return Response({"error": "Only Excel files (.xlsx, .xls) are allowed"}, status=status.HTTP_400_BAD_REQUEST)

        results = {
            "total": 0,
            "success": 0, "updated": 0, "skipped": 0, "errors": []
        }

        try:
            excel_file = pd.ExcelFile(file_obj)
            sheet_names = excel_file.sheet_names
            
            student_role = UserRole.objects.get_or_create(role_name='Student')[0]
            
            # Resolve Default Batch
            def_batch = None
            scheme_id = request.data.get('scheme_id')
            if default_batch_id:
                b_year = None
                if str(default_batch_id).isdigit():
                    def_batch = Batch.objects.filter(pk=default_batch_id).first()
                elif '-' in str(default_batch_id):
                    try: b_year = int(str(default_batch_id).split('-')[0])
                    except: pass
                
                if not def_batch and b_year:
                    def_batch = Batch.objects.filter(batch_year=b_year).first()
                    if not def_batch and scheme_id:
                        scheme = Scheme.objects.filter(pk=scheme_id).first()
                        if scheme:
                            def_batch = Batch.objects.create(
                                batch_year=b_year,
                                scheme_id=scheme,
                                start_year=b_year,
                                end_year=b_year + 3
                            )
            
            def_prog = Program.objects.filter(pk=default_program_id).first() if default_program_id and str(default_program_id).isdigit() else None

            alias_map = {
                'program_name': ['programname', 'program', 'department', 'dept', 'branch', 'branchname', 'stream', 'course'],
                'batch_year': ['batchyear', 'batch', 'admyear', 'admissionyear', 'batchperiod', 'acadbatch'],
                'academic_year': ['academicyear', 'ay', 'academic', 'year', 'session', 'academicsession'],
                'semester': ['semester', 'sem', 'term', 'semesterid'],
                'class_year': ['class', 'classyear', 'year', 'yr', 'grade', 'classid'],
                'division': ['division', 'div', 'section', 'group', 'divid'],
                'enrollment_no': [
                    'enrollmentno', 'enrollment', 'enrollno', 'enrolment', 'regno', 'regnumber', 'prn', 'studentid', 'sid', 'id'
                ],
                'roll_no': [
                    'rollno', 'rollnumber', 'roll', 'rno', 'rollcallno', 'rn', 'seatno', 'seatnumber', 'srno'
                ],
                'student_name': [
                    'studentname', 'name', 'fullname', 'nameofstudent', 'stname', 'student', 'fname'
                ],
                'email': ['email', 'mail', 'emailaddress', 'studentemail', 'mailid', 'emailid'],
                'is_active': ['isactive', 'active', 'status', 'enabled']
            }

            def clean_key(s):
                return "".join(c.lower() for c in str(s) if c.isalnum())

            clean_alias_map = {k: [clean_key(a) for a in v] for k, v in alias_map.items()}

            for sheet_name in sheet_names:
                try:
                    preview_df = pd.read_excel(file_obj, sheet_name=sheet_name, header=None, nrows=20)
                except:
                    continue

                header_row_idx = -1
                detection_keywords = ['enroll', 'roll', 'student', 'name', 'prn', 'id']
                for i, row in preview_df.iterrows():
                    row_str = " ".join([str(val).lower() for val in row if pd.notnull(val)])
                    hits = sum(1 for k in detection_keywords if k in row_str)
                    if hits >= 2:
                        header_row_idx = i
                        break
                
                if header_row_idx == -1:
                    continue # Skip sheets that don't look like student lists

                df = pd.read_excel(file_obj, sheet_name=sheet_name, skiprows=header_row_idx, dtype=str)
                
                # Dynamic Rename Map
                rename_map = {}
                for col in df.columns:
                    ck = clean_key(col)
                    if not ck: continue
                    for internal_key, aliases in clean_alias_map.items():
                        if ck in aliases and internal_key not in rename_map.values():
                            rename_map[col] = internal_key
                            break
                
                df = df.rename(columns=rename_map)
                
                # Validate critical columns for this sheet
                if 'enrollment_no' not in df.columns or 'roll_no' not in df.columns or 'student_name' not in df.columns:
                    continue # This sheet is probably something else

                results["total"] += len(df)

                for index, row in df.iterrows():
                    row_num = index + header_row_idx + 2
                    try:
                        with transaction.atomic():
                            def get_val(key):
                                v = row.get(key)
                                if pd.isnull(v) or str(v).strip().lower() in ['nan', 'none', '']: return None
                                s = str(v).strip()
                                if 'e' in s.lower() or '+' in s:
                                    try: s = format(float(s), 'f').split('.')[0]
                                    except: pass
                                if s.endswith('.0'): s = s[:-2]
                                return s

                            enroll_no = get_val('enrollment_no')
                            roll_no = get_val('roll_no')
                            name = get_val('student_name')
                            
                            if not enroll_no or not roll_no or not name:
                                if row.isnull().all():
                                    results["total"] -= 1
                                    continue
                                results["errors"].append(f"Sheet '{sheet_name}', Row {row_num}: Missing critical data (Enroll: '{enroll_no}', Roll: '{roll_no}', Name: '{name}')")
                                results["skipped"] += 1
                                continue

                            # Value Parsing
                            ay_val = get_val('academic_year') or default_academic_year
                            if ay_val:
                                if '-' not in str(ay_val):
                                    try:
                                        y = int(float(str(ay_val)))
                                        ay_val = f"{y}-{(y+1)%100:02d}"
                                    except: ay_val = None
                            
                            if not ay_val:
                                results["errors"].append(f"Sheet '{sheet_name}', Row {row_num}: Invalid Academic Year.")
                                results["skipped"] += 1
                                continue

                            sem_val = get_val('semester') or default_semester
                            try:
                                sem = int(float(str(sem_val))) if sem_val else None
                            except: sem = None
                            
                            if not sem:
                                results["errors"].append(f"Sheet '{sheet_name}', Row {row_num}: Invalid Semester '{sem_val}'.")
                                results["skipped"] += 1
                                continue

                            c_year = get_val('class_year') or default_class_year
                            div = get_val('division') or default_division
                            is_active = str(row.get('is_active', '')).strip().lower() != 'false'
                            email = get_val('email')

                            # Mapping
                            prog_name = get_val('program_name')
                            program = Program.objects.filter(program_name__iexact=prog_name).first() if prog_name else def_prog
                            
                            batch = None
                            batch_year_input = get_val('batch_year')
                            scheme_id = request.data.get('scheme_id')

                            if batch_year_input:
                                try:
                                    y = int(batch_year_input.split('-')[0]) if '-' in batch_year_input else int(float(batch_year_input))
                                    batch = Batch.objects.filter(batch_year=y).first()
                                    
                                    # Auto-create batch if year and scheme are known
                                    if not batch and scheme_id:
                                        scheme = Scheme.objects.filter(pk=scheme_id).first()
                                        if scheme:
                                            batch = Batch.objects.create(
                                                batch_year=y,
                                                scheme_id=scheme,
                                                start_year=y,
                                                end_year=y + (scheme.end_year - scheme.start_year if scheme.end_year and scheme.start_year else 3)
                                            )
                                except Exception as e:
                                    print(f"DEBUG: Batch creation failed: {e}")
                            
                            if not batch: batch = def_batch

                            if not program or not batch:
                                results["errors"].append(f"Sheet '{sheet_name}', Row {row_num}: Missing Program or Batch mapping (Prog: {program}, Batch: {batch}).")
                                results["skipped"] += 1
                                continue

                            # Resolution
                            student = Student.objects.filter(roll_no=roll_no, batch_id=batch).first()
                            if not student:
                                student = Student.objects.filter(enrollment_no=enroll_no).first()

                            if student:
                                # Overwrite protection for enrollment_no conflicts
                                other = Student.objects.filter(enrollment_no=enroll_no).exclude(pk=student.pk).first()
                                if other:
                                    other.enrollment_no = f"X_{other.enrollment_no}_{other.pk}"
                                    other.is_active = False
                                    other.save()

                                student.roll_no = roll_no
                                student.enrollment_no = enroll_no
                                student.name = name
                                student.class_year = c_year
                                student.division = div
                                student.semester = sem
                                student.academic_year = ay_val
                                student.batch_id = batch
                                student.program_id = program
                                student.is_active = is_active
                                student.save()
                                
                                user = getattr(student, 'user_id', None)
                                if user:
                                    user.name = name
                                    user.username = enroll_no
                                    if email and '@' in email: user.email = email
                                    user.save()
                                results["updated"] += 1
                            else:
                                # Create new
                                user_obj, created = User.objects.get_or_create(
                                    username=enroll_no, 
                                    defaults={
                                        'name': name,
                                        'email': email if email and '@' in email else f"{enroll_no}@example.com",
                                        'role_id': student_role
                                    }
                                )
                                if created:
                                    user_obj.set_password("Student@123")
                                    user_obj.save()
                                
                                Student.objects.create(
                                    user_id=user_obj, enrollment_no=enroll_no, roll_no=roll_no,
                                    name=name, program_id=program, batch_id=batch,
                                    class_year=c_year, semester=sem, division=div,
                                    academic_year=ay_val, is_active=is_active
                                )
                                results["success"] += 1

                    except IntegrityError as ie:
                        results["errors"].append(f"Sheet '{sheet_name}', Row {row_num}: Database Conflict - {str(ie)}")
                        results["skipped"] += 1
                    except Exception as e:
                        results["errors"].append(f"Sheet '{sheet_name}', Row {row_num}: {str(e)}")
                        results["skipped"] += 1

            return Response(results, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({"error": f"General Error: {str(e)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class BulkMarksUploadView(APIView):
    parser_classes = (MultiPartParser, FormParser)

    def post(self, request):
        file_obj = request.FILES.get('file')
        assessment_id = request.data.get('assessment_id')

        if not file_obj or not assessment_id:
            return Response({"error": "file and assessment_id are required"}, status=status.HTTP_400_BAD_REQUEST)

        assessment = get_object_or_404(Assessment, pk=assessment_id)
        set_by = request.user if request.user and not request.user.is_anonymous else User.objects.first()

        try:
            decoded_file = file_obj.read().decode('utf-8')
            io_string = io.StringIO(decoded_file)
            reader = csv.DictReader(io_string)
            
            marks_created = 0

            with transaction.atomic():
                for row in reader:
                    roll_no = row.get('roll_no')
                    marks_obtained = row.get('marks')
                    
                    student = Student.objects.filter(roll_no=roll_no).first()
                    if student:
                        MarksEntry.objects.update_or_create(
                            assessment_id=assessment,
                            student_id=student,
                            defaults={
                                'marks_obtained': marks_obtained,
                                'entered_by': set_by
                            }
                        )
                        marks_created += 1

            return Response({"message": f"Successfully uploaded marks for {marks_created} students"}, status=status.HTTP_201_CREATED)
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class PromoteStudentsView(APIView):
    """
    Carries forward students from a previous semester to the current selection
    within the same batch and division.
    """
    def post(self, request):
        batch_id = request.data.get('batch_id')
        division = request.data.get('division')
        target_sem = request.data.get('semester')
        target_class = request.data.get('class_year')
        target_ay = request.data.get('academic_year')
        
        if not batch_id or not target_sem:
            return Response({"error": "Batch and Semester are required"}, status=400)

        try:
            source_sem = int(target_sem) - 1
            if source_sem < 1:
                return Response({"error": "Cannot promote from before Semester 1"}, status=400)
        except:
            return Response({"error": "Invalid semester format"}, status=400)

        # Resolve Batch
        batch = None
        if batch_id:
            if str(batch_id).isdigit():
                batch = Batch.objects.filter(pk=batch_id).first()
            elif '-' in str(batch_id):
                try:
                    b_year = int(str(batch_id).split('-')[0])
                    batch = Batch.objects.filter(batch_year=b_year).first()
                except: pass

        if not batch:
            return Response({"error": "Valid Batch is required"}, status=400)

        # Find students in the previous semester for this batch and division
        students = Student.objects.filter(
            batch_id=batch,
            semester=source_sem,
            division=division,
            is_active=True
        )

        if not students.exists():
            return Response({
                "error": f"No students found in the previous semester (Sem {source_sem}).",
                "details": f"To carry forward, students must first be registered in Semester {source_sem}."
            }, status=404)

        created_count = 0
        skipped_count = 0
        errors = []

        with transaction.atomic():
            for student in students:
                try:
                    # Check if this enrollment_no is already taken in the target semester/year
                    exists_enroll = Student.objects.filter(
                        enrollment_no=student.enrollment_no,
                        semester=target_sem,
                        academic_year=target_ay
                    ).exists()

                    if exists_enroll:
                        skipped_count += 1
                        continue

                    # Check if this roll_no is already taken in the target semester for this batch
                    exists_roll = Student.objects.filter(
                        batch_id=batch,
                        roll_no=student.roll_no,
                        semester=target_sem
                    ).exists()

                    if exists_roll:
                        errors.append(f"Student '{student.name}' (Roll {student.roll_no}) already has a different entry in Semester {target_sem}.")
                        continue

                    # Create a NEW explicit record (Copy and Paste)
                    Student.objects.create(
                        name=student.name,
                        roll_no=student.roll_no,
                        enrollment_no=student.enrollment_no,
                        program_id=student.program_id,
                        batch_id=student.batch_id,
                        class_year=target_class,
                        division=student.division,
                        semester=target_sem,
                        academic_year=target_ay,
                        user_id=student.user_id,
                        is_active=True
                    )
                    created_count += 1
                except Exception as e:
                    errors.append(f"Could not carry forward '{student.name}': {str(e)}")

        if errors and created_count == 0:
            # Combine the main error and specific details so they show up in the frontend alert()
            full_error = "Carry forward failed for all students. Reasons: " + " | ".join(errors[:3])
            return Response({
                "error": full_error,
                "details": errors
            }, status=400)

        return Response({
            "message": f"Successfully carried forward {created_count} students to Semester {target_sem}.",
            "count": created_count,
            "skipped": skipped_count,
            "errors": errors if errors else None
        }, status=status.HTTP_200_OK)

class BulkCISUploadView(APIView):
    """
    Highly robust bulk upload for CIS marks.
    Detects formats:
    1. System Format (3 rows: Q, Wt, CO)
    2. Sandip/Green Format (2 rows: CO, Q/Desc)
    """
    parser_classes = (MultiPartParser, FormParser)

    def post(self, request):
        file_obj = request.FILES.get('file')
        course_id = request.data.get('course_id')
        academic_year = request.data.get('academic_year')
        tool_name = request.data.get('tool_name')
        semester = request.data.get('semester')
        
        if not all([file_obj, course_id, academic_year, tool_name, semester]):
            return Response({"error": "Missing required fields (course_id, tool_name, academic_year, semester)"}, status=400)

        try:
            df = pd.read_excel(file_obj, header=None)
            
            # 1. Orientation: Find Enrollment Row
            enroll_row = -1
            enroll_col = -1
            for r in range(min(10, len(df))): # Look in first 10 rows
                for c in range(min(5, len(df.columns))):
                    val = str(df.iloc[r, c]).upper()
                    if "ENROLLMENT" in val or "ENROLMENT" in val:
                        enroll_row, enroll_col = r, c
                        break
                if enroll_row != -1: break

            if enroll_row == -1:
                return Response({"error": "Could not find 'ENROLLMENT NO' column header."}, status=400)

            # 2. Header Block Identification
            # Scan 3 rows around enroll_row (from idx 0 to enroll_row)
            header_range = df.iloc[0 : enroll_row + 1]
            
            q_row_idx, wt_row_idx, co_row_idx = -1, -1, -1
            
            for r_idx in range(len(header_range)):
                row_str = " ".join([str(v).upper() for v in header_range.iloc[r_idx] if pd.notnull(v)])
                if co_row_idx == -1 and ("CO" in row_str or "COURSE OUTCOME" in row_str):
                    co_row_idx = r_idx
                elif wt_row_idx == -1 and ("WT" in row_str or "WEIGHT" in row_str or "MAX MARKS" in row_str):
                    wt_row_idx = r_idx
                elif q_row_idx == -1 and ("QUESTION" in row_str or " Q " in row_str or " Q1 " in row_str or "DESCRIPTION" in row_str or "DEVELOP" in row_str):
                    q_row_idx = r_idx

            # Fallbacks based on position
            if q_row_idx == -1: q_row_idx = enroll_row
            
            # Start column for questions
            q_start_col = -1
            q_names_row = df.iloc[q_row_idx]
            # Excluded column names that might act as "Labels"
            excluded_headers = [
                "NAN", "", "NAME", "NAME OF STUDENT", "SR. NO.", "SR.NO.", "SR", 
                "ROLL NO.", "ROLL NO", "ROLL", "ENROLLMENT", "ENROLMENT", "ENROLLM ENT NO",
                "Q", "WT", "CO", "WEIGHT", "COURSE OUTCOME"
            ]
            for c in range(enroll_col + 1, len(q_names_row)):
                val = str(q_names_row[c]).strip().upper()
                if val not in excluded_headers:
                    q_start_col = c
                    break
            
            if q_start_col == -1: q_start_col = enroll_col + 3 
            
            # Data usually starts immediately after the header row (enroll_row)
            # BUT if it's a multi-row header (Wt/CO present below name), it starts later.
            # For the Green format in our image, data starts at enroll_row + 1
            data_start_row_final = enroll_row + 1
            
            # 3. Extract Metadata
            # Forward-fill merged cells in metadata rows
            q_names_row = df.iloc[q_row_idx].fillna(method='ffill')
            wt_row = df.iloc[wt_row_idx].fillna(method='ffill') if wt_row_idx != -1 else None
            co_row = df.iloc[co_row_idx].fillna(method='ffill') if co_row_idx != -1 else None

            custom_questions, user_cos, custom_weights = [], [], []
            
            # Pull defaults from course config for holes
            course = get_object_or_404(Course, pk=course_id)
            tool_defaults = (course.assessment_tools or {}).get(tool_name, {})
            
            for c in range(q_start_col, len(df.columns)):
                q_text = str(q_names_row[c]).strip()
                if q_text.upper() in ["TOTAL", "PERCENTAGE", "AVERAGE", "NAN", ""]:
                    if len(custom_questions) > 0 and c > q_start_col + 1: break # End of questions if we hit a total column
                    q_text = f"Q{len(custom_questions)+1}"
                
                custom_questions.append(q_text)
                
                # Weight
                w_val = 5.0 # Default fallback
                if wt_row is not None:
                    try: w_val = float(wt_row[c])
                    except: pass
                elif 'customWeights' in tool_defaults and len(custom_weights) < len(tool_defaults['customWeights']):
                     w_val = tool_defaults['customWeights'][len(custom_weights)]
                elif "CES" in tool_name.upper(): w_val = 4.0
                custom_weights.append(w_val)
                
                # CO
                co_val = ""
                if co_row is not None:
                    co_val = str(co_row[c]).strip().upper()
                    if co_val and not co_val.startswith("CO") and any(char.isdigit() for char in co_val):
                        # Convert things like "1" to "CO1"
                        digits = "".join([char for char in co_val if char.isdigit()])
                        if digits: co_val = f"CO{digits}"
                elif 'userCos' in tool_defaults and len(user_cos) < len(tool_defaults['userCos']):
                    co_val = tool_defaults['userCos'][len(user_cos)]
                user_cos.append(co_val if co_val != "NAN" else "")

            # 4. Extract Student Marks
            marks_map, marks_list = {}, []
            for r in range(data_start_row_final, len(df)):
                enroll_val = str(df.iloc[r, enroll_col]).strip().replace('.0', '')
                if enroll_val.upper() in ["AVERAGE", "TOTAL", "CO ATTAINMENT", "NAN", "", "NUMBER OF"]: continue
                if not any(char.isdigit() for char in enroll_val): continue
                
                total_m, q_marks = 0, {}
                for i in range(len(custom_questions)):
                    try: 
                        m = float(df.iloc[r, q_start_col + i])
                        if pd.isna(m): m = 0
                    except: m = 0
                    q_marks[str(i)] = m
                    total_m += m
                
                tool_base = tool_name.split('-')[0]
                is_avg_tool = tool_base == "FA" and "PR" in tool_name or tool_base == "SLA"
                
                if is_avg_tool and len(custom_questions) > 0:
                    total_m = round(total_m / len(custom_questions), 2)
                    
                q_marks['total'] = total_m
                
                marks_map[enroll_val] = q_marks
                marks_list.append({"enrollment_no": enroll_val, "marks": total_m})

            # 5. Atomic Upsert
            user = request.user if request.user.is_authenticated else User.objects.first()
            max_m = sum(custom_weights)
            if "CES" in tool_name.upper() and max_m > 50: max_m = 4.0 # Sanity check for surveys

            with transaction.atomic():
                tool_base = tool_name.split('-')[0]
                t_suffix = "_TH" if "TH" in tool_name else "_PR" if "PR" in tool_name else ""
                t_type = f"{tool_base}{t_suffix}"
                
                # Special cases
                if "SLA" in tool_name: t_type = "SLA"
                elif "CES" in tool_name: t_type = "CES"

                assessment, _ = Assessment.objects.update_or_create(
                    course_id=course, assessment_name=tool_name, 
                    academic_year=academic_year, semester=semester,
                    defaults={
                        'assessment_type': t_type, 'max_marks': max_m,
                        'weightage': 1.0, # Default internal weightage
                        'configuration': {
                            'columnCount': len(custom_questions), 'customQuestions': custom_questions,
                            'customWeights': custom_weights, 'userCos': user_cos,
                            'marksData': marks_map, 'toolKey': tool_name
                        },
                        'user_id': user, 'is_active': True
                    }
                )

                # Mapping & Marks
                AssessmentCOMapping.objects.filter(assessment_id=assessment).delete()
                co_sums = {}
                for idx, co_n in enumerate(user_cos):
                    if co_n: co_sums[co_n] = co_sums.get(co_n, 0) + custom_weights[idx]
                
                for co_n, weight in co_sums.items():
                    co_obj = CO.objects.filter(course_id=course, co_number__iexact=co_n).first()
                    if co_obj: AssessmentCOMapping.objects.create(assessment_id=assessment, co_id=co_obj, co_weightage=weight)

                for item in marks_list:
                    student = Student.objects.filter(enrollment_no=item['enrollment_no'], program_id=course.program_id).first()
                    if student:
                        MarksEntry.objects.update_or_create(
                            assessment_id=assessment, student_id=student,
                            defaults={'marks_obtained': item['marks'], 'user_id': user}
                        )

            AttainmentService.calculate_attainment(course_id, academic_year)
            return Response({"message": f"Successfully uploaded {len(marks_list)} records."}, status=201)

        except Exception as e:
            return Response({"error": f"Upload Failed: {str(e)}"}, status=500)

class DownloadCISMultiSheetTemplateView(APIView):
    """
    Generates a multi-sheet Excel template for all CIS tools.
    """
    def get(self, request):
        course_id = request.query_params.get('course_id')
        academic_year = request.query_params.get('academic_year')
        division = request.query_params.get('division') or None
        if not course_id:
            return Response({"error": "Course ID is required"}, status=400)
            
        try:
            excel_data = generate_cis_multi_sheet_template(course_id, academic_year, division=division)
            course = get_object_or_404(Course, pk=course_id)
            
            response = HttpResponse(excel_data, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename=Marks_Bulk_Upload_Template.xlsx'
            return response
        except Exception as e:
            return Response({"error": str(e)}, status=500)


class BulkCISApplyView(APIView):
    """
    Processes a multi-sheet CIS Excel file and applies marks to all corresponding tools.
    """
    parser_classes = (MultiPartParser, FormParser)

    def post(self, request):
        file_obj = request.FILES.get('file')
        course_id = request.data.get('course_id')
        academic_year = request.data.get('academic_year')
        semester = request.data.get('semester')
        division = request.data.get('division') or None
        
        if not all([file_obj, course_id, academic_year, semester]):
            return Response({"error": "Missing required fields (file, course_id, academic_year, semester)"}, status=400)

        try:
            report = process_bulk_cis_apply(
                file=file_obj,
                course_id=course_id,
                academic_year=academic_year,
                semester=semester,
                user=request.user if request.user.is_authenticated else User.objects.first(),
                division=division
            )
            return Response({"message": "Bulk Apply completed", "report": report}, status=200)
        except Exception as e:
            import traceback
            traceback.print_exc()
            return Response({"error": f"Bulk Apply Failed: {str(e)}"}, status=500)

class BulkCourseUploadView(APIView):
    """
    Handles bulk upload of courses from an Excel file.
    Maps headers like 'Course Code', 'Course Name', 'Type', 'Class', 'Sem' etc.
    """
    parser_classes = (MultiPartParser, FormParser)

    def post(self, request):
        file_obj = request.FILES.get('file')
        if not file_obj:
            return Response({"error": "No file uploaded"}, status=status.HTTP_400_BAD_REQUEST)

        # Defaults from frontend filters
        default_program_id = request.data.get('program_id')
        default_scheme_id = request.data.get('scheme_id')
        default_academic_year = request.data.get('academic_year')

        if not (file_obj.name.endswith('.xlsx') or file_obj.name.endswith('.xls')):
            return Response({"error": "Only Excel files (.xlsx, .xls) are allowed"}, status=status.HTTP_400_BAD_REQUEST)

        try:
            # 1. Header Detection (More robust)
            preview_df = pd.read_excel(file_obj, header=None, nrows=15)
            header_row_idx = 0
            # Higher threshold and specific combination check
            keywords = ['course', 'code', 'name', 'title', 'abbr', 'sem', 'class', 'scheme', 'dept', 'program']
            for i, row in preview_df.iterrows():
                row_vals = [str(val).lower() for val in row if pd.notnull(val)]
                row_str = " ".join(row_vals)
                
                # Count hits
                hits = sum(1 for k in keywords if k in row_str)
                
                # Check for critical combination: code AND name
                has_critical = any('code' in v for v in row_vals) and any('name' in v for v in row_vals)
                
                if hits >= 3 or (hits >= 2 and has_critical):
                    header_row_idx = i
                    break
            
            file_obj.seek(0)
            df = pd.read_excel(file_obj, skiprows=header_row_idx, dtype=str)
            
            # Remove any completely empty columns
            df = df.dropna(axis=1, how='all')
            
            # 2. Alias Mapping
            def clean(s): return "".join(c.lower() for c in str(s) if c.isalnum())
            
            alias_map = {
                'course_code': ['coursecode', 'code', 'subcode', 'subjectcode', 'ccode', 'id', 'courseid'],
                'course_name': ['coursename', 'name', 'subjectname', 'subject', 'cname', 'subname'],
                'course_title': ['coursetitle', 'title', 'subjecttitle', 'ctitle'],
                'course_abbr': ['courseabbr', 'abbr', 'abbreviatedname', 'shortname', 'subjabbr', 'subabbr'],
                'semester': ['semester', 'sem', 'sme', 'term'],
                'class_year': ['class', 'classyear', 'year', 'yr', 'grade'],
                'program_name': ['program', 'department', 'dept', 'branch', 'pname', 'programname', 'deptname'],
                'scheme_name': ['scheme', 'schemename', 'revision', 'sname', 'reg', 'regulation']
            }

            rename_map = {}
            for col in df.columns:
                if 'unnamed' in str(col).lower(): continue
                c_col = clean(col)
                if not c_col: continue
                
                found = False
                # Try exact clean match first
                for internal, aliases in alias_map.items():
                    if c_col in [clean(a) for a in aliases]:
                        if internal not in rename_map.values():
                            rename_map[col] = internal
                            found = True
                        break
                
                # Fallback: substring match
                if not found:
                    for internal, aliases in alias_map.items():
                        if any(clean(a) in c_col for a in aliases if len(a) > 3):
                            if internal not in rename_map.values():
                                rename_map[col] = internal
                            break
            
            df = df.rename(columns=rename_map)
            
            # Critical Validation
            if 'course_code' not in df.columns or 'course_name' not in df.columns:
                return Response({
                    "error": "Required columns 'Course Code' and 'Course Name' not found.",
                    "detected_headers": list(df.columns),
                    "header_row_identified": header_row_idx + 1
                }, status=status.HTTP_400_BAD_REQUEST)

            courses_created = 0
            courses_updated = 0
            errors = []

            with transaction.atomic():
                for idx, row in df.iterrows():
                    try:
                        code = str(row.get('course_code')).strip()
                        name = str(row.get('course_name')).strip()
                        if not code or not name or code == "nan": continue

                        # Resolve Program
                        p_name = str(row.get('program_name', '')).strip()
                        if p_name and p_name != "nan":
                            program = Program.objects.filter(Q(program_name__iexact=p_name) | Q(program_name__icontains=p_name)).first()
                            if not program:
                                program = Program.objects.create(program_name=p_name)
                        elif default_program_id:
                            program = Program.objects.filter(pk=default_program_id).first()
                        else:
                            program = Program.objects.first() # Total fallback

                        if not program:
                            errors.append(f"Row {idx+2}: Could not resolve program.")
                            continue

                        # Resolve Scheme
                        s_name = str(row.get('scheme_name', '')).strip()
                        if s_name and s_name != "nan":
                            scheme = Scheme.objects.filter(Q(scheme_name__iexact=s_name) | Q(scheme_name__icontains=s_name)).first()
                            if not scheme:
                                # Default years if creating on the fly
                                scheme = Scheme.objects.create(scheme_name=s_name, start_year=2023)
                        elif default_scheme_id:
                            scheme = Scheme.objects.filter(pk=default_scheme_id).first()
                        else:
                            scheme = None

                        # Resolve Sem and Class
                        sem = str(row.get('semester', '1')).replace('Semester', '').strip()
                        try: sem_int = int(float(sem))
                        except: sem_int = 1

                        class_y = str(row.get('class_year', '')).strip()
                        if not class_y or class_y == "nan":
                            if sem_int <= 2: class_y = "FY"
                            elif sem_int <= 4: class_y = "SY"
                            else: class_y = "TY"

                        # Upsert Course
                        course, created = Course.objects.update_or_create(
                            course_code=code,
                            program_id=program,
                            defaults={
                                'course_name': name,
                                'course_title': str(row.get('course_title', name)).strip(),
                                'course_abbr': str(row.get('course_abbr', code[:5])).strip(),
                                'semester': sem_int,
                                'class_year': class_y,
                                'scheme_id': scheme,
                                'is_active': True
                            }
                        )
                        
                        if created: courses_created += 1
                        else: courses_updated += 1
                        
                    except Exception as row_err:
                        errors.append(f"Row {idx+2}: {str(row_err)}")

            return Response({
                "message": f"Successfully processed courses: {courses_created} created, {courses_updated} updated.",
                "errors": errors[:10] # Limit reported errors
            }, status=status.HTTP_201_CREATED)

        except Exception as e:
            return Response({"error": f"Failed to parse file: {str(e)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class DownloadUserTemplateView(APIView):
    """
    Generates and serves an Excel template for user bulk upload (HOD, Coordinator, Faculty, Auditor).
    """
    def get(self, request):
        cols = [
            'Name', 'Email', 'Contact No', 'Role', 'Department Name', 'Date of Joining'
        ]
        df = pd.DataFrame(columns=cols)
        
        # Add sample rows
        df.loc[0] = ['John Doe', 'john@example.com', '9876543210', 'Faculty', 'Computer Engineering', '2023-01-15']
        df.loc[1] = ['Jane Smith', 'jane@example.com', '9123456789', 'HOD', 'Information Technology', '2022-06-01']
        
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Users')
        output.seek(0)
        
        response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=User_Bulk_Upload_Template.xlsx'
        return response

class BulkUserUploadView(APIView):
    parser_classes = (MultiPartParser, FormParser)

    def post(self, request):
        file_obj = request.FILES.get('file')
        if not file_obj:
            return Response({"error": "No file uploaded"}, status=status.HTTP_400_BAD_REQUEST)

        if not (file_obj.name.endswith('.xlsx') or file_obj.name.endswith('.xls')):
            return Response({"error": "Only Excel files (.xlsx, .xls) are allowed"}, status=status.HTTP_400_BAD_REQUEST)

        results = {"total": 0, "success": 0, "updated": 0, "skipped": 0, "errors": []}

        try:
            df = pd.read_excel(file_obj, dtype=str)
            df = df.fillna('')
            
            # Map column names
            def clean_k(s): return "".join(c.lower() for c in str(s) if c.isalnum())
            
            alias_map = {
                'name': ['name', 'fullname', 'username', 'user'],
                'email': ['email', 'emailid', 'mail', 'emailaddress'],
                'contact_no': ['contactno', 'contact', 'mobile', 'phone', 'mobileno'],
                'role': ['role', 'designation', 'usertype'],
                'dept_name': ['departmentname', 'department', 'dept', 'program', 'programname'],
                'doj': ['dateofjoining', 'doj', 'joiningdate']
            }

            rename_map = {}
            for col in df.columns:
                ck = clean_k(col)
                for internal_key, aliases in alias_map.items():
                    if ck in aliases:
                        rename_map[col] = internal_key
                        break
            
            df = df.rename(columns=rename_map)
            results["total"] = len(df)

            for index, row in df.iterrows():
                row_num = index + 2
                try:
                    name = str(row.get('name', '')).strip()
                    email = str(row.get('email', '')).strip().lower()
                    role_name = str(row.get('role', '')).strip()
                    dept_name = str(row.get('dept_name', '')).strip()
                    
                    if not email or not role_name:
                        results["errors"].append(f"Row {row_num}: Missing email or role.")
                        results["skipped"] += 1
                        continue

                    if role_name.upper() == 'STUDENT':
                        results["errors"].append(f"Row {row_num}: Students should be uploaded via Student Bulk Upload.")
                        results["skipped"] += 1
                        continue

                    # Resolve Role
                    role_obj = UserRole.objects.filter(role_name__iexact=role_name).first()
                    if not role_obj:
                        results["errors"].append(f"Row {row_num}: Role '{role_name}' not found.")
                        results["skipped"] += 1
                        continue

                    # Resolve Department
                    dept_obj = None
                    if dept_name:
                        from academics.models import Program
                        dept_obj = Program.objects.filter(Q(program_name__icontains=dept_name) | Q(program_abbr__iexact=dept_name)).first()

                    with transaction.atomic():
                        user, created = User.objects.get_or_create(
                            email=email,
                            defaults={
                                'username': email,
                                'name': name or email.split('@')[0],
                                'role_id': role_obj,
                                'department': dept_obj,
                                'contact_no': str(row.get('contact_no', '')).strip(),
                                'date_of_joining': str(row.get('doj', '')).strip()
                            }
                        )

                        if created:
                            user.set_password("Admin@123")
                            user.save()
                            results["success"] += 1
                        else:
                            # Update existing
                            user.name = name or user.name
                            user.role_id = role_obj
                            user.department = dept_obj or user.department
                            user.contact_no = str(row.get('contact_no', '')).strip() or user.contact_no
                            user.date_of_joining = str(row.get('doj', '')).strip() or user.date_of_joining
                            user.save()
                            results["updated"] += 1

                except Exception as e:
                    results["errors"].append(f"Row {row_num}: {str(e)}")
                    results["skipped"] += 1

            return Response(results, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

# =====================================================================
# DATA SEEDING ONLY - REMOVE LATER
# =====================================================================

class DownloadSurveyTemplateView(APIView):
    """
    Generates a pre-filled Excel template for Survey responses (Course Exit / Indirect).
    Matches the CIS Marks Entry UI exactly: Enrollment, Roll No, Name, and dynamic Survey item columns.
    """
    def get(self, request):
        survey_id = request.query_params.get('survey_id')
        
        if not survey_id:
            return Response({"error": "Survey ID is required"}, status=400)
            
        from surveys.models import SurveyMaster, SurveyQuestion
        survey = get_object_or_404(SurveyMaster, pk=survey_id)
        
        output = io.BytesIO()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Survey {survey.survey_category}"
        
        from openpyxl.styles import Border, Side
        # Column Headers - Labels are row-specific for the 4th column
        ws['A1'] = "ENROLLMENT NO"
        ws['B1'] = "Roll no."
        ws['C1'] = "Name of Student"
        
        # Merge A-C for the 3 header rows
        for col_char in ['A', 'B', 'C']:
            ws.merge_cells(f'{col_char}1:{col_char}3')
            cell = ws[f'{col_char}1']
            apply_header_style(cell, fill_color="2F5597", font_color="FFFFFF")

        # Row 1-3 Labels in Column D (Just purely cosmetic spacers to match CIS)
        ws['D1'] = ""
        ws['D2'] = ""
        ws['D3'] = "Survey Item"
        for r in [1, 2, 3]:
            apply_header_style(ws.cell(row=r, column=4), fill_color="2F5597", font_color="FFFFFF")

        students = []
        col_headers = []
        
        if survey.survey_category == 'course_exit' and survey.course_id:
            students = list(Student.objects.filter(program_id=survey.course_id.program_id, is_active=True))
            cos = CO.objects.filter(course_id=survey.course_id, is_active=True).order_by('co_number')
            col_headers = [co.co_number for co in cos]
        elif survey.survey_category == 'indirect' and survey.program_id:
            students = list(Student.objects.filter(program_id=survey.program_id, is_active=True))
            from academics.models import PO, PSO
            pos = PO.objects.filter(program_id=survey.program_id, is_active=True).order_by('po_number')
            psos = PSO.objects.filter(program_id=survey.program_id, is_active=True).order_by('pso_number')
            col_headers = [po.po_number for po in pos] + [pso.pso_number for pso in psos]
        else:
            # Fallback to existing questions
            questions = SurveyQuestion.objects.filter(survey_id=survey)
            col_headers = [f"Q{q.question_id}" for q in questions]
            
        students.sort(key=lambda x: natural_sort_key(x.roll_no or ""))

        # Target Items (Row 3) Starting from Col E (5)
        for i, header in enumerate(col_headers):
            col_idx = 5 + i
            ws.cell(row=1, column=col_idx, value="")
            ws.cell(row=2, column=col_idx, value="")
            ws.cell(row=3, column=col_idx, value=header)
            
            apply_header_style(ws.cell(row=1, column=col_idx), fill_color="FCE4D6", font_color="000000")
            apply_header_style(ws.cell(row=2, column=col_idx), fill_color="FCE4D6", font_color="000000")
            apply_header_style(ws.cell(row=3, column=col_idx), fill_color="FCE4D6", font_color="000000")

        # Add Students starting row 4
        for idx, student in enumerate(students, start=4):
            ws.cell(row=idx, column=1, value=student.enrollment_no).border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            ws.cell(row=idx, column=2, value=student.roll_no).border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            ws.cell(row=idx, column=3, value=student.name).border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            # Empty Label in Col D for students
            ws.cell(row=idx, column=4, value="").border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            
            for i in range(len(col_headers)):
                 ws.cell(row=idx, column=5+i).border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=Survey_Bulk_Upload_Template.xlsx'
        return response

class BulkSurveyUploadView(APIView):
    """
    Bulk uploads survey responses (strictly for Data Seeding).
    Connects to the specified survey_id and creates SurveyResponse/SurveyAnswer records.
    """
    parser_classes = (MultiPartParser, FormParser)

    def post(self, request):
        file_obj = request.FILES.get('file')
        survey_id = request.data.get('survey_id')
        
        if not all([file_obj, survey_id]):
            return Response({"error": "Missing required fields (file, survey_id)"}, status=400)

        from surveys.models import SurveyMaster, SurveyQuestion, SurveyResponse, SurveyAnswer
        try:
            survey = get_object_or_404(SurveyMaster, pk=survey_id)
            df = pd.read_excel(file_obj, header=None)
            
            # 1. Orientation: Find Enrollment Row
            enroll_row = -1
            enroll_col = -1
            for r in range(min(10, len(df))): # Look in first 10 rows
                for c in range(min(5, len(df.columns))):
                    val = str(df.iloc[r, c]).upper()
                    if "ENROLLMENT" in val or "ENROLMENT" in val:
                        enroll_row, enroll_col = r, c
                        break
                if enroll_row != -1: break

            if enroll_row == -1:
                return Response({"error": "Could not find 'ENROLLMENT NO' column header."}, status=400)

            # Define headers directly from row 3 (idx 2) if standard CIS format is used
            # Default to header being 2 rows below the enrollment NO row
            header_row_idx = enroll_row + 2
            if len(df) <= header_row_idx:
                return Response({"error": "Empty data sheet"}, status=400)
                
            q_start_col = enroll_col + 4 # Skip Enrollment, Roll, Name, Label Col
            header_row = df.iloc[header_row_idx].ffill()
            
            # Map columns to question IDs
            survey_questions = {} # idx (int) -> question_id
            
            for c in range(q_start_col, len(df.columns)):
                col_name = str(header_row[c]).strip().upper()
                if col_name in ["NAN", "", "NONE"]: continue
                
                # Resolve the column to an existing question, or create it
                question = None
                
                if survey.survey_category == 'course_exit':
                    # Expecting COs
                    digits = "".join([char for char in col_name if char.isdigit()])
                    co_number = f"CO{digits}" if digits else col_name
                    target_co = None
                    if survey.course_id:
                        cos = CO.objects.filter(course_id=survey.course_id, is_active=True)
                        for co in cos:
                            if str(co.co_number).upper().endswith(digits) or str(co.co_number).upper() == co_number:
                                target_co = co
                                break
                    
                    if target_co:
                        question, _ = SurveyQuestion.objects.get_or_create(
                            survey_id=survey,
                            co_id_id=target_co.co_id,
                            defaults={'question_text': f"Evaluation for {target_co.co_number}"}
                        )
                elif survey.survey_category == 'indirect':
                    # Expecting POs and PSOs
                    from academics.models import PO, PSO
                    target_obj = None
                    digits = "".join([char for char in col_name if char.isdigit()])
                    val_number = f"PO{digits}" if "PO" in col_name and "PSO" not in col_name else (f"PSO{digits}" if "PSO" in col_name else col_name)

                    if survey.program_id:
                        if "PSO" in col_name:
                             for p in PSO.objects.filter(program_id=survey.program_id, is_active=True):
                                 if str(p.pso_number).upper().endswith(digits) or str(p.pso_number).upper() == val_number:
                                     target_obj = p
                                     break
                             if target_obj:
                                 question, _ = SurveyQuestion.objects.get_or_create(
                                     survey_id=survey, pso_id=target_obj,
                                     defaults={'question_text': f"Evaluation for {target_obj.pso_number}"}
                                 )
                        else:
                             for p in PO.objects.filter(program_id=survey.program_id, is_active=True):
                                 if str(p.po_number).upper().endswith(digits) or str(p.po_number).upper() == val_number:
                                     target_obj = p
                                     break
                             if target_obj:
                                 question, _ = SurveyQuestion.objects.get_or_create(
                                     survey_id=survey, po_id=target_obj,
                                     defaults={'question_text': f"Evaluation for {target_obj.po_number}"}
                                 )
                
                # Generic fallback if no mapping found
                if not question:
                    question, _ = SurveyQuestion.objects.get_or_create(
                        survey_id=survey,
                        question_text__iexact=f"Evaluation for {col_name}",
                        defaults={'question_text': f"Evaluation for {col_name}"}
                    )
                
                survey_questions[c] = question.question_id
                
            data_start_row_final = header_row_idx + 1
            results = {"total_processed": 0, "responses_created": 0}
            
            with transaction.atomic():
                for r in range(data_start_row_final, len(df)):
                    enroll_val = str(df.iloc[r, enroll_col]).strip().replace('.0', '')
                    name_val = str(df.iloc[r, enroll_col + 2]).strip()
                    
                    if enroll_val.upper() in ["AVERAGE", "TOTAL", "CO ATTAINMENT", "NAN", "", "NUMBER OF"]: continue
                    if not any(char.isdigit() for char in enroll_val): continue
                    
                    student = Student.objects.filter(enrollment_no=enroll_val).first()
                    
                    response_obj, created = SurveyResponse.objects.update_or_create(
                        survey_id=survey,
                        student_id=student,
                        defaults={
                            'respondent_name': student.name if student else (name_val if name_val != 'nan' else "Guest"),
                            'enrollment_no': enroll_val
                        }
                    )
                    
                    if created: results["responses_created"] += 1
                    
                    # Wipe existing answers for this response
                    SurveyAnswer.objects.filter(response_id=response_obj).delete()
                    
                    answers_to_create = []
                    for c_idx, q_id in survey_questions.items():
                        try:
                            m = float(df.iloc[r, c_idx])
                            if pd.notna(m):
                                answers_to_create.append(SurveyAnswer(
                                    response_id=response_obj,
                                    question_id_id=q_id,
                                    answer_value=int(m)
                                ))
                        except Exception as e:
                            pass # Blank or invalid answer
                            
                    SurveyAnswer.objects.bulk_create(answers_to_create)
                    results["total_processed"] += 1
                    
            return Response({"message": f"Successfully uploaded {results['responses_created']} responses ({results['total_processed']} rows processed)."}, status=status.HTTP_201_CREATED)

        except Exception as e:
            return Response({"error": f"General Error: {str(e)}"}, status=500)

