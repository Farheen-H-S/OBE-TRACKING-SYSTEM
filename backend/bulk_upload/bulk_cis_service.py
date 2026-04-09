import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from django.shortcuts import get_object_or_404
from academics.models import Course, CO
from users.models import Student
from assessments.models import Assessment, MarksEntry, AssessmentCOMapping
from attainment.attainment_service import AttainmentService
from django.db import transaction
import pandas as pd
import io
import re

def apply_header_style(cell, fill_color="2F5597", font_color="FFFFFF"):
    cell.font = Font(bold=True, color=font_color)
    cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

def natural_sort_key(s):
    return [int(text) if text.isdigit() else text.lower()
            for text in re.split('([0-9]+)', str(s))]

def normalize_roll(r):
    if pd.isna(r): return ""
    s = str(r).strip()
    if s.lower() == 'nan': return ""
    if s.endswith('.0'): s = s[:-2]
    if s.endswith('.'): s = s[:-1]
    
    # Remove leading zeros if it's purely numeric, but keep it as string
    if s.isdigit():
        return str(int(s)).lower()
    
    return s.lower()

def generate_cis_multi_sheet_template(course_id, academic_year=None, division=None):
    course = get_object_or_404(Course, pk=course_id)
    
    def get_students(filt_base, ay):
        from django.db.models import Q
        qs = Student.objects.filter(**filt_base)
        if ay:
            ay_c = str(ay).replace(" ", "")
            ay_s = ay_c.replace("-", " - ")
            qs = qs.filter(Q(academic_year=ay_c) | Q(academic_year=ay_s) | Q(academic_year=ay))
        return qs.distinct()

    # Stage 1: Strict (Prog + Sem + Class + Batch)
    base_filters = {
        'program_id': course.program_id,
        'semester': course.semester,
        'class_year': course.class_year,
        'is_active': True
    }
    if course.batches.exists():
        base_filters['batch_id__in'] = course.batches.all()
    
    students = get_students(base_filters, academic_year)

    # Stage 2: Cohort-focused (Prog + Batch) - Ignore Sem/Class
    if not students.exists() and course.batches.exists():
        cohort_filters = {'program_id': course.program_id, 'batch_id__in': course.batches.all(), 'is_active': True}
        students = get_students(cohort_filters, academic_year)

    # Stage 3: Program-focused (Prog + AY)
    if not students.exists():
        prog_ay_filters = {'program_id': course.program_id, 'is_active': True}
        students = get_students(prog_ay_filters, academic_year)
    
    # Stage 4: Broadest (Prog only)
    if not students.exists():
        students = Student.objects.filter(program_id=course.program_id, is_active=True).distinct()
    
    # Apply natural sort to students by roll_no
    students_list = sorted(list(students), key=lambda x: natural_sort_key(x.roll_no))
    
    output = io.BytesIO()
    wb = openpyxl.Workbook()
    
    # Sheets to create
    sheets_config = [
        {"name": "Class Test 1 (FA-TH)", "tool": "FA-TH-CT1", "type": "CT"},
        {"name": "Class Test 2 (FA-TH)", "tool": "FA-TH-CT2", "type": "CT"},
        {"name": "FA-PR (K3)", "tool": "FA-PR", "type": "PR"},
        {"name": "SLA", "tool": "SLA", "type": "SLA"},
        {"name": "SA-TH", "tool": "SA-TH", "type": "SA"},
        {"name": "SA-PR", "tool": "SA-PR", "type": "SA"},
    ]
    
    for idx, config in enumerate(sheets_config):
        if idx == 0:
            ws = wb.active
            ws.title = config["name"]
        else:
            ws = wb.create_sheet(title=config["name"])
            
        # ... logic ...
        tool_type = config["type"]
        
        # Generic Headers - A1:A3 merged for CT/PR/SLA (3 header rows), single for SA
        if tool_type in ("CT", "PR", "SLA"):
            ws.merge_cells('A1:A3')
            ws.merge_cells('B1:B3')
        
        ws['A1'] = "Roll No"
        apply_header_style(ws['A1'], fill_color="2F5597", font_color="FFFFFF")
        ws.column_dimensions['A'].width = 12

        ws['B1'] = "Name of Student"
        apply_header_style(ws['B1'], fill_color="D9E1F2", font_color="1F3864")
        ws.column_dimensions['B'].width = 28
        
        if tool_type == "CT":
            # Col C: row labels (Q / Wt / CO)
            ws['C1'] = "Q"
            ws['C2'] = "Wt"
            ws['C3'] = "CO"
            apply_header_style(ws['C1'])
            apply_header_style(ws['C2'])
            apply_header_style(ws['C3'])
            ws.column_dimensions['C'].width = 6

            # 1(a)-1(g), 2(a)-2(g) starting from Col D (col_idx=4, 1-based)
            col_idx = 4
            for q_num in range(1, 3):
                for sub in 'abcdefg':
                    cell = ws.cell(row=1, column=col_idx, value=f"{q_num}({sub})")
                    apply_header_style(cell, fill_color="FFFF00", font_color="000000")
                    ws.cell(row=2, column=col_idx, value=2 if q_num == 1 else 4)
                    ws.cell(row=3, column=col_idx, value="CO1")
                    col_idx += 1
                
        elif tool_type in ("PR", "SLA"):
            # Col C: row labels
            label_row1 = "Practical No" if tool_type == "PR" else "Assignment"
            label_row3 = "Course Outcome" if tool_type == "PR" else "CO"
            ws['C1'] = label_row1
            ws['C2'] = "Max Marks"
            ws['C3'] = label_row3
            apply_header_style(ws['C1'])
            apply_header_style(ws['C2'])
            apply_header_style(ws['C3'])
            ws.column_dimensions['C'].width = 14

            # Questions start from Col D (col_idx=4, 1-based)
            count = 10 if tool_type == "PR" else 4
            for i in range(1, count + 1):
                col_idx = 3 + i  # 4, 5, 6 ...
                cell = ws.cell(row=1, column=col_idx, value=i)
                apply_header_style(cell, fill_color="FFFF00", font_color="000000")
                ws.cell(row=2, column=col_idx, value=25 if tool_type == "PR" else 20)
                ws.cell(row=3, column=col_idx, value=f"CO{i}")

        elif tool_type == "SA":
            # SA has only 1 header row: Roll No | Name of Student | Total Marks
            ws['C1'] = "Total Marks"
            apply_header_style(ws['A1'])
            apply_header_style(ws['C1'])
            ws.column_dimensions['C'].width = 15
                
        # Add Students starting row 4 (row 2 for SA since it has only 1 header row)
        data_start_row = 4 if tool_type != "SA" else 2
        for s_idx, student in enumerate(students_list, start=data_start_row):
            ws.cell(row=s_idx, column=1, value=student.roll_no)
            ws.cell(row=s_idx, column=2, value=student.name)

            
    wb.save(output)
    output.seek(0)
    return output.read()

def process_bulk_cis_apply(file, course_id, academic_year, semester, user, division=None):
    course = get_object_or_404(Course, pk=course_id)
    all_sheets = pd.read_excel(file, sheet_name=None, header=None)
    
    report = {}
    
    # Mapping sheet names/tools to internal types
    tool_map = {
        "Class Test 1 (FA-TH)": {"tool_name": "FA-TH-CT1", "tool_type": "FA_TH", "parser": "CT"},
        "Class Test 2 (FA-TH)": {"tool_name": "FA-TH-CT2", "tool_type": "FA_TH", "parser": "CT"},
        "FA-PR (K3)": {"tool_name": "FA-PR", "tool_type": "FA_PR", "parser": "PR"},
        "SLA": {"tool_name": "SLA", "tool_type": "SLA", "parser": "SLA"},
        "SA-TH": {"tool_name": "SA-TH", "tool_type": "SA_TH", "parser": "SA"},
        "SA-PR": {"tool_name": "SA-PR", "tool_type": "SA_PR", "parser": "SA"},
    }
    
    def get_student_qs(filt_base, ay):
        from django.db.models import Q
        qs = Student.objects.filter(**filt_base)
        if ay:
            ay_c = str(ay).replace(" ", "")
            ay_s = ay_c.replace("-", " - ")
            qs = qs.filter(Q(academic_year=ay_c) | Q(academic_year=ay_s) | Q(academic_year=ay))
        return qs

    # Same fallback logic for matching
    base_f = {'program_id': course.program_id, 'semester': course.semester, 'class_year': course.class_year, 'is_active': True}
    if course.batches.exists(): base_f['batch_id__in'] = course.batches.all()
    
    student_qs = get_student_qs(base_f, academic_year)
    
    if not student_qs.exists() and course.batches.exists():
        student_qs = get_student_qs({'program_id': course.program_id, 'batch_id__in': course.batches.all(), 'is_active': True}, academic_year)
        
    if not student_qs.exists():
        student_qs = get_student_qs({'program_id': course.program_id, 'is_active': True}, academic_year)

    if not student_qs.exists():
        student_qs = Student.objects.filter(program_id=course.program_id, is_active=True)

    students_in_context = {normalize_roll(s.roll_no): s for s in student_qs}
    
    for sheet_name, df in all_sheets.items():
        if sheet_name not in tool_map:
            continue
            
        config = tool_map[sheet_name]
        try:
            with transaction.atomic():
                result = _parse_and_save_sheet(df, config, course, academic_year, course.semester, students_in_context, user)
                report[sheet_name] = result
        except Exception as e:
            report[sheet_name] = f"Error: {str(e)}"
            
    # Trigger recalculation
    AttainmentService.calculate_attainment(course_id, academic_year)
    
    return report

def _parse_and_save_sheet(df, config, course, ay, sem, students_map, user):
    tool_name = config["tool_name"]
    tool_type = config["tool_type"]
    parser = config["parser"]
    
    # Identifiers
    questions = []
    weights = []
    cos = []
    marks_data = [] # List of (student, marks_json, total_sum)
    errors = []
    unmatched_rolls = []  # Track roll numbers not found in roster
    
    if parser == "CT":
        # Row 0: Qs, Row 1: Wts, Row 2: COs
        # Col C (index 2) = row labels (Q/Wt/CO) — skip it, start reading from Col D (index 3)
        for col in range(3, len(df.columns)):
            q = str(df.iloc[0, col]).strip()
            if not q or q == "nan": break
            questions.append(q)
            weights.append(float(df.iloc[1, col]) if pd.notnull(df.iloc[1, col]) else 0.0)
            cos.append(str(df.iloc[2, col]).strip())
            
        for row in range(3, len(df)):
            roll = str(df.iloc[row, 0]).strip()
            if not roll or roll == "nan": continue
            norm_roll = normalize_roll(roll)
            if norm_roll in students_map:
                student = students_map[norm_roll]
                s_marks = {}
                row_total = 0
                for col_idx, q in enumerate(questions):
                    m = df.iloc[row, 3 + col_idx]  # col D (index 3) onwards
                    try: 
                        m_val = float(m) if pd.notnull(m) and str(m).strip() not in ('', '-', 'ab', 'AB', 'Abs', 'abs', 'absent', 'Absent') else None
                    except: m_val = None
                    
                    if m_val is not None:
                        max_w = weights[col_idx]
                        if max_w > 0 and m_val > max_w:
                            errors.append(f"Row {row+1}, Q '{q}': {m_val} exceeds max {max_w} — capped")
                            m_val = max_w  # Cap and warn, don't abort
                        
                    s_marks[col_idx] = m_val
                    # Logic for CT usually best of 5, but for storage we just need marksData
                
                # Smart Choice detection: If sum of weights <= max marks, it's MCQ/Mandatory
                # We need max_m which is calculated later, but we can calculate it here too
                weight_sum_total = sum(weights)
                # Attempt to find max_marks from config or default to something sensible
                tool_config = course.assessment_tools.get(tool_name, {}) if course.assessment_tools else {}
                max_marks_limit = float(tool_config.get('maxMarks', 30.0))
                
                has_choice = weight_sum_total > (max_marks_limit + 0.1)

                total = 0
                if has_choice:
                    def get_best_5(marks_dict, start_idx, end_idx):
                        vals = []
                        for i in range(start_idx, end_idx + 1):
                            v = marks_dict.get(i)
                            if v is not None:
                                vals.append(v)
                        vals.sort(reverse=True)
                        return sum(vals[:5])
                    
                    total = get_best_5(s_marks, 0, 6) + get_best_5(s_marks, 7, 13)
                else:
                    total = sum(v for v in s_marks.values() if v is not None)
                    
                s_marks['total'] = total
                marks_data.append((student, s_marks, total))
            else:
                unmatched_rolls.append(roll)
                
    elif parser in ["PR", "SLA"]:
        # Col C (index 2) = row labels (Practical No/Assignment/Max Marks/CO) — skip, read from Col D (index 3)
        for col in range(3, len(df.columns)):
            q = str(df.iloc[0, col]).strip()
            if not q or q == "nan": break
            questions.append(q)
            weights.append(float(df.iloc[1, col]) if pd.notnull(df.iloc[1, col]) else 0.0)
            cos.append(str(df.iloc[2, col]).strip())
            
        for row in range(3, len(df)):
            roll = str(df.iloc[row, 0]).strip()
            if not roll or roll == "nan": continue
            norm_roll = normalize_roll(roll)
            if norm_roll in students_map:
                student = students_map[norm_roll]
                s_marks = {}
                row_sum = 0
                for col_idx, q in enumerate(questions):
                    m = df.iloc[row, 3 + col_idx]  # col D (index 3) onwards
                    try: m_val = float(m) if pd.notnull(m) and str(m).strip() not in ('', '-', 'ab', 'AB', 'Abs', 'abs', 'absent', 'Absent') else None
                    except: m_val = None
                    
                    if m_val is not None:
                        max_w = weights[col_idx]
                        if max_w > 0 and m_val > max_w:
                            errors.append(f"Row {row+1}, Q '{q}': {m_val} exceeds max {max_w} — capped")
                            m_val = max_w  # Cap instead of abort
                    
                    s_marks[col_idx] = m_val
                    if m_val is not None:
                        row_sum += m_val
                
                valid_marks = [v for v in s_marks.values() if v is not None]
                total = sum(valid_marks) / len(valid_marks) if valid_marks else 0
                s_marks['total'] = round(total, 2) if valid_marks else None
                marks_data.append((student, s_marks, s_marks['total']))
            else:
                unmatched_rolls.append(roll)

    elif parser == "SA":
        # Template: Roll No (col A=0), Name (col B=1), Total Marks (col C=2)
        # Find the header row to skip it, then read data rows
        tool_config = course.assessment_tools.get(tool_name, {}) if course.assessment_tools else {}
        max_w = float(tool_config.get('maxMarks', 100))

        # Detect header row — find row where col 0 says "Roll No" or similar
        data_start_row = 1  # default: skip row 0 (header)
        first_row_val = str(df.iloc[0, 0]).strip().lower()
        if any(kw in first_row_val for kw in ['roll', 'enroll', 'sr', 'no']):
            data_start_row = 1  # real header row is row 0, data starts at 1
        
        # Detect which column holds the Total Marks
        # In our template: col C (index 2). Fallback: scan header row for "total"/"marks"
        total_col = 2  # default to column C
        if len(df.columns) > 2:
            for c in range(len(df.columns)):
                hdr = str(df.iloc[0, c]).strip().lower()
                if 'total' in hdr or 'marks' in hdr:
                    total_col = c
                    break

        for row in range(data_start_row, len(df)):
            roll = str(df.iloc[row, 0]).strip()
            if not roll or roll == "nan": continue
            # Skip footer rows like "Average", "Total" etc.
            if any(kw in roll.lower() for kw in ['total', 'avg', 'average', 'co attainment']): continue
            norm_roll = normalize_roll(roll)
            if norm_roll in students_map:
                student = students_map[norm_roll]
                try:
                    raw = df.iloc[row, total_col]
                    if pd.isna(raw) or str(raw).strip() in ('', '-', 'ab', 'AB', 'Abs', 'abs', 'absent', 'Absent'):
                        total = None
                    else:
                        total = float(raw)
                except:
                    total = None
                
                if total is not None and total > max_w:
                    errors.append(f"Row {row+1}: Total marks ({total}) exceed max configuration ({max_w})")
                    
                marks_data.append((student, {'0': total, 'total': total}, total))
            else:
                unmatched_rolls.append(roll)
        
        questions = ["Total"]
        weights = [max_w]
        cos = ["CO1"]  # Default for SA if not specified
    
    # If no student rows found — return descriptive message, no DB change
    if not marks_data and not questions:
        return "Skipped (empty sheet — no headers found)"
    
    if not marks_data:
        return f"No matching students found (check roll numbers match the system)"
    
    # Build warning prefix if any caps were applied
    warning_prefix = ""
    if errors:
        warning_prefix = "⚠️ Warning (marks capped): " + "; ".join(errors[:3])
        if len(errors) > 3:
            warning_prefix += f" ...and {len(errors)-3} more."
        warning_prefix += "\n"
    
    # Save Assessment
    if parser == "SA":
        max_m = float(weights[0]) if weights else 100.0
    elif parser in ["PR", "SLA"]:
        # For average-based tools, max_m should be the average of weights
        max_m = sum(float(w) for w in weights) / len(weights) if weights else 0.0
    else:
        max_m = sum(float(w) for w in weights)
    
    assessment, _ = Assessment.objects.update_or_create(
        course_id=course, assessment_name=tool_name, 
        academic_year=ay, semester=sem,
        defaults={
            'assessment_type': tool_type, 'max_marks': max_m,
            'weightage': 1.0, 
            'configuration': {
                'columnCount': len(questions), 'customQuestions': questions,
                'customWeights': weights, 'userCos': cos,
                'marksData': {m[0].enrollment_no: m[1] for m in marks_data}, # Map by Enrollment for consistency
                'toolKey': tool_name
            },
            'user_id': user, 'is_active': True
        }
    )
    
    # Save CO Mappings
    AssessmentCOMapping.objects.filter(assessment_id=assessment).delete()
    co_sums = {}
    for idx, co_n in enumerate(cos):
        if co_n: 
            # Normalize CO name
            co_norm = f"CO{co_n}" if not str(co_n).upper().startswith("CO") else str(co_n).upper()
            co_sums[co_norm] = co_sums.get(co_norm, 0) + float(weights[idx])
            
    for co_num, weight in co_sums.items():
        co_obj = CO.objects.filter(course_id=course, co_number__iexact=co_num).first()
        if co_obj:
            AssessmentCOMapping.objects.create(assessment_id=assessment, co_id=co_obj, co_weightage=weight)

    # Bulk Save Marks — skip absent (None total) students entirely
    MarksEntry.objects.filter(assessment_id=assessment).delete()
    entries = [
        MarksEntry(assessment_id=assessment, student_id=m[0], marks_obtained=m[2], user_id=user)
        for m in marks_data if m[2] is not None
    ]
    MarksEntry.objects.bulk_create(entries)
    
    success_msg = f"Success: {len(marks_data)} student(s) updated"
    if unmatched_rolls:
        # Limit to 5 roll numbers displayed to avoid super long messages
        sample = unmatched_rolls[:5]
        extra = f" ...+{len(unmatched_rolls)-5} more" if len(unmatched_rolls) > 5 else ""
        success_msg += f" | ⚠️ {len(unmatched_rolls)} roll(s) not matched: {', '.join(str(r) for r in sample)}{extra}"
    return warning_prefix + success_msg if warning_prefix else success_msg

