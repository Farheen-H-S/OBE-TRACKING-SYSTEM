import io
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference, Series, PieChart, BarChart3D
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

class StressExcelReportGenerator:
    @staticmethod
    def generate(data):
        wb = Workbook()
        
        # 1. Report_Info
        StressExcelReportGenerator._create_info_sheet(wb, data)
        
        # 2. Overall_Stress_Summary
        StressExcelReportGenerator._create_summary_sheet(wb, data)
        
        # 3. Domain_Wise_Stress
        StressExcelReportGenerator._create_domain_sheet(wb, data)
        
        # 4. Question_Wise_Analysis
        StressExcelReportGenerator._create_question_sheet(wb, data)
        
        # 5. Stress_Level_Distribution
        StressExcelReportGenerator._create_distribution_sheet(wb, data)
        
        # 6. Raw_Aggregated_Data
        StressExcelReportGenerator._create_raw_data_sheet(wb, data)
        
        # 7. HOD_Action_Plan
        StressExcelReportGenerator._create_action_plan_sheet(wb, data)
        
        # Remove default sheet if exists
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]
            
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    @staticmethod
    def _apply_header_style(cell):
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        StressExcelReportGenerator._apply_border(cell)

    @staticmethod
    def _apply_border(cell):
        thin = Side(border_style="thin", color="000000")
        cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

    @staticmethod
    def _auto_adjust_columns(ws):
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

    @staticmethod
    def _create_info_sheet(wb, data):
        ws = wb.active
        ws.title = "Report_Info"
        info = data["survey_info"]
        
        rows = [
            ["Institution Name", "Sandip Polytechnic"],
            ["Department", "Computer Engineering"],
            ["Program", "Diploma in Computer Engineering"],
            ["Academic Year", "2025-26"],
            ["Survey Cycle", f"{info['month']}/{info['year']}"],
            ["Survey Window", f"{info.get('window_start').strftime('%Y-%m-%d') if hasattr(info.get('window_start'), 'strftime') else 'N/A'} - {info.get('window_end').strftime('%Y-%m-%d') if hasattr(info.get('window_end'), 'strftime') else 'N/A'}"],
            [""],
            ["Likert Scale Definition"],
            ["0", "Not at all stressed / 😌"],
            ["1", "Slightly stressed / 🙂"],
            ["2", "Moderately stressed / 😐"],
            ["3", "Quite stressed / 😟"],
            ["4", "Extremely stressed / 😫"],
            [""],
            ["Stress Level Thresholds"],
            ["Low", "Avg Score < 1.5"],
            ["Medium", "Avg Score 1.5 - 2.5"],
            ["High", "Avg Score > 2.5"],
            [""],
            ["Participation Stats"],
            ["Total Responses", info["total_responses"]],
            ["Response Rate (%)", "N/A"] # We don't have total students in this context yet
        ]
        
        for r in rows:
            ws.append(r)
            for cell in ws[ws.max_row]:
                StressExcelReportGenerator._apply_border(cell)
        
        StressExcelReportGenerator._auto_adjust_columns(ws)

    @staticmethod
    def _create_summary_sheet(wb, data):
        ws = wb.create_sheet("Overall_Stress_Summary")
        summary = data["overall_summary"]
        
        ws.append(["Metric", "Value"])
        for cell in ws[1]:
            StressExcelReportGenerator._apply_header_style(cell)

        ws.append(["Overall Average Stress Score", summary["avg_score"]])
        ws.append(["Overall Stress Level", summary["level"]])
        for row in ws.iter_rows(min_row=2, max_row=3, max_col=2):
            for cell in row:
                StressExcelReportGenerator._apply_border(cell)

        ws.append([""])
        ws.append(["Stress Level", "Count", "Percentage"])
        for cell in ws[5]:
            StressExcelReportGenerator._apply_header_style(cell)
        
        for level, count in summary["counts"].items():
            ws.append([level, count, f"{summary['distribution'][level]}%"])
            for cell in ws[ws.max_row]:
                StressExcelReportGenerator._apply_border(cell)
            
        if summary["counts"] and any(summary["counts"].values()):
            chart = BarChart()
            chart.type = "col"
            chart.style = 10
            chart.title = "Stress Level Distribution"
            chart.y_axis.title = "Number of Students"
            chart.x_axis.title = "Stress Level"
            chart.height = 10 
            chart.width = 15
            
            # Label position
            chart.x_axis.tickLblPos = 'low'
            chart.y_axis.tickLblPos = 'nextTo'
            
            # Row 5 is header, Rows 6-8 are LOW, MODERATE, HIGH
            data_ref = Reference(ws, min_col=2, min_row=5, max_row=8)
            cats_ref = Reference(ws, min_col=1, min_row=6, max_row=8)
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats_ref)
            ws.add_chart(chart, "E2")
        
        StressExcelReportGenerator._auto_adjust_columns(ws)

    @staticmethod
    def _create_domain_sheet(wb, data):
        ws = wb.create_sheet("Domain_Wise_Stress")
        ws.append(["Domain Name", "Average Stress Score", "Stress Level", "High Stress %", "Rank"])
        for cell in ws[1]:
            StressExcelReportGenerator._apply_header_style(cell)

        row_idx = 2
        for d in data["domain_analysis"]:
            ws.append([d["name"], d["avg_score"], d["level"], f"{d['high_pct']}%", d["rank"]])
            for cell in ws[ws.max_row]:
                StressExcelReportGenerator._apply_border(cell)
            row_idx += 1
            
        if row_idx > 2:
            chart = BarChart()
            chart.title = "Domain Wise Stress Analysis"
            chart.x_axis.title = "Domains"
            chart.y_axis.title = "Average Score (0-4)"
            chart.height = 10
            chart.width = 15

            chart.x_axis.tickLblPos = 'low'
            chart.y_axis.tickLblPos = 'nextTo'

            data_ref = Reference(ws, min_col=2, min_row=1, max_row=row_idx-1)
            cats_ref = Reference(ws, min_col=1, min_row=2, max_row=row_idx-1)
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats_ref)
            ws.add_chart(chart, "G2")
        
        StressExcelReportGenerator._auto_adjust_columns(ws)

    @staticmethod
    def _create_question_sheet(wb, data):
        ws = wb.create_sheet("Question_Wise_Analysis")
        ws.append(["Question ID", "Question Text", "Domain", "Average Score", "Level", "High Count", "High %"])
        for cell in ws[1]:
            StressExcelReportGenerator._apply_header_style(cell)

        row_idx = 2
        for q in data["question_analysis"]:
            ws.append([q["qid"], q["text"], q["domain"], q["avg_score"], q["level"], q["high_count"], f"{q['high_pct']}%"])
            for cell in ws[ws.max_row]:
                StressExcelReportGenerator._apply_border(cell)
            row_idx += 1
            
        if row_idx > 2:
            chart = BarChart()
            chart.type = "bar"
            chart.title = "Question Wise Stress (Average Score)"
            chart.x_axis.title = "Avg Score"
            chart.y_axis.title = "Questions"
            chart.height = 15
            chart.width = 20

            # For bar charts, x_axis is vertical, y_axis is horizontal
            chart.x_axis.tickLblPos = 'low'
            chart.y_axis.tickLblPos = 'nextTo'

            max_r = min(15, row_idx - 1)
            data_ref = Reference(ws, min_col=4, min_row=1, max_row=max_r)
            cats_ref = Reference(ws, min_col=2, min_row=2, max_row=max_r)
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats_ref)
            ws.add_chart(chart, "I2")
        
        StressExcelReportGenerator._auto_adjust_columns(ws)

    @staticmethod
    def _create_distribution_sheet(wb, data):
        ws = wb.create_sheet("Stress_Level_Distribution")
        summary = data["overall_summary"]
        ws.append(["Stress Level", "Student Count", "Percentage"])
        for cell in ws[1]:
            StressExcelReportGenerator._apply_header_style(cell)

        for level, count in summary["counts"].items():
            ws.append([level, count, f"{summary['distribution'][level]}%"])
            for cell in ws[ws.max_row]:
                StressExcelReportGenerator._apply_border(cell)
            
        # Stacked Bar Chart (Mocked as BarChart with overlap)
        chart = BarChart()
        chart.type = "bar"
        chart.grouping = "stacked"
        chart.overlap = 100
        chart.title = "Stress Level Distribution"
        chart.height = 10
        chart.width = 15
        
        data_ref = Reference(ws, min_col=2, min_row=1, max_row=4)
        cats_ref = Reference(ws, min_col=1, min_row=2, max_row=4)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        ws.add_chart(chart, "E2")
        
        StressExcelReportGenerator._auto_adjust_columns(ws)

    @staticmethod
    def _create_raw_data_sheet(wb, data):
        ws = wb.create_sheet("Raw_Aggregated_Data")
        ws.append(["Domain", "Question ID", "Response Count", "Average Score"])
        for cell in ws[1]:
            StressExcelReportGenerator._apply_header_style(cell)

        for q in data["question_analysis"]:
            ws.append([q["domain"], q["qid"], data["survey_info"]["total_responses"], q["avg_score"]])
            for cell in ws[ws.max_row]:
                StressExcelReportGenerator._apply_border(cell)
        
        StressExcelReportGenerator._auto_adjust_columns(ws)

    @staticmethod
    def _create_action_plan_sheet(wb, data):
        ws = wb.create_sheet("HOD_Action_Plan")
        ws.append(["Target Batch", "Analysis Remarks", "Action Taken", "Last Updated"])
        
        # Apply styles to headers
        for cell in ws[1]:
            StressExcelReportGenerator._apply_header_style(cell)

        if not data.get("action_plans"):
            ws.append(["No action plan recorded yet", "-", "-", "-"])
            for cell in ws[ws.max_row]:
                StressExcelReportGenerator._apply_border(cell)
        else:
            for p in data["action_plans"]:
                ws.append([p["batch"], p["analysis"], p["action"], p["date"]])
                for cell in ws[ws.max_row]:
                    StressExcelReportGenerator._apply_border(cell)
        
        StressExcelReportGenerator._auto_adjust_columns(ws)

        # Enable text wrapping for long remarks
        for row in ws.iter_rows(min_row=2, max_col=3):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
