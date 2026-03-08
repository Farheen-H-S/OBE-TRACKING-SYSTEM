import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

class SurveyExcelReportGenerator:
    @staticmethod
    def generate(data):
        wb = Workbook()
        # Sheet 1: Matrix Data
        ws1 = wb.active
        ws1.title = "Analysis Report"

        survey = data.get('survey', {})
        statements = data.get('statements', [])
        teachers = data.get('teachers', [])

        # 1. Header Information
        ws1.append(["Teacher Feedback Analysis Report"])
        ws1.append(["Survey Name", survey.get('survey_name')])
        ws1.append(["Academic Year", survey.get('academic_year')])
        ws1.append(["Total Responses", data.get('total_responses', 0)])
        ws1.append([]) # Empty row

        # 2. Table Headers
        headers = ["#", "Teacher Name"]
        for stmt in statements:
            headers.append(stmt)
        headers.append("Achieved Score")
        
        ws1.append(headers)
        
        # Style Headers
        header_row = ws1.max_row
        header_fill = PatternFill(start_color="445D99", end_color="445D99", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws1[header_row]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            SurveyExcelReportGenerator._apply_border(cell)

        # 3. Data Rows
        for i, t in enumerate(teachers, 1):
            row = [i, t['teacher']]
            for stmt in statements:
                row.append(t['scores'].get(stmt, "-"))
            row.append(t['achieved_score'])
            
            ws1.append(row)
            for cell in ws1[ws1.max_row]:
                SurveyExcelReportGenerator._apply_border(cell)
                cell.alignment = Alignment(horizontal="center")

        SurveyExcelReportGenerator._auto_adjust_columns(ws1)

        # Sheet 2: Chart Data & Visualization
        from openpyxl.chart import BarChart, Reference
        ws2 = wb.create_sheet(title="Graphical Analysis")
        
        ws2.append(["Teacher Name", "Achieved Score"])
        for t in teachers:
            ws2.append([t['teacher'], t['achieved_score']])
        
        # Create Bar Chart
        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = "Teacher Feedback Performance Analysis"
        chart.y_axis.title = 'Score (1-5)'
        chart.x_axis.title = 'Teachers'
        
        data_ref = Reference(ws2, min_col=2, min_row=1, max_row=len(teachers) + 1)
        cats_ref = Reference(ws2, min_col=1, min_row=2, max_row=len(teachers) + 1)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        chart.shape = 4
        
        ws2.add_chart(chart, "D2")

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    @staticmethod
    def _apply_border(cell):
        thin = Side(border_style="thin", color="000000")
        cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

    @staticmethod
    def _auto_adjust_columns(ws):
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50) # Cap width
            ws.column_dimensions[column].width = adjusted_width
