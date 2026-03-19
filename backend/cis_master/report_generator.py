import openpyxl
import os
import re
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from django.db import models
from django.db.models import Avg, Count
from academics.models import Course, CO, COPOMapping, COPSOMapping, PO, PSO, AcademicSetup, COTarget
from attainment.models import COAttainment, POAttainment
from attainment.attainment_service import AttainmentService
from indirect_attainment.models import CourseIndirectAttainment, ActivityIndirectAttainment
from assessments.models import Assessment, MarksEntry
from users.models import Student, FacultyCourseAssignment
from surveys.models import SurveyMaster, SurveyQuestion, SurveyResponse, SurveyAnswer
from .models import CISType, CISTerm

def natural_sort_key(s):
    return [int(text) if text.isdigit() else text.lower()
            for text in re.split('([0-9]+)', str(s))]

# Color Constants (Matching User Images)
HEADER_DARK_BLUE = "2F5597"
HEADER_LIGHT_BLUE = "D9E1F2"
STAT_GREEN_LIGHT = "E2EFDA"
STAT_GREEN_MEDIUM = "C6E0B4"
STAT_ORANGE = "F8CBAD"
CT_YELLOW = "FFFF00"
CT_PINK = "FCE4D6"
LIGHT_YELLOW = "FFFF99"
BORDER_COLOR = "000000"
WHITE_TEXT = "FFFFFF"
BLACK_TEXT = "000000"
CO_BLUE = "ADD8E6"
MAPPING_ORANGE = "FACE8D"
CT_LIGHT_BLUE = "E7E6FE"
HEADER_LIGHT_PINK = "FCE4D6"

# Styling Objects
header_font = Font(bold=True, color=BLACK_TEXT)
yellow_fill = PatternFill(start_color=CT_YELLOW, end_color=CT_YELLOW, fill_type="solid")
red_font = Font(color="FF0000", bold=True)
center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

def get_border():
    return Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )

def apply_header_style(cell, fill_color=HEADER_DARK_BLUE, font_color=WHITE_TEXT, bold=True, size=11):
    cell.font = Font(bold=bold, color=font_color, size=size)
    cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = get_border()

def add_common_header(ws, title_type, faculty_name="Not Assigned", program_name="Computer Engineering Program"):
    # Logo Area (A1:B3)
    ws.merge_cells('A1:B3')
    # Try to load the real logo
    # Assuming the path: D:\Farheen\OBE-TRACKING-SYSTEM\frontend\webapp\src\assets\images\sflogo.jpg
    # We can use an absolute path or relative to project root
    base_dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    logo_path = os.path.join(base_dir, 'frontend', 'webapp', 'src', 'assets', 'images', 'sflogo.jpg')
    
    if os.path.exists(logo_path):
        try:
            img = Image(logo_path)
            # Adjust image size to fit roughly A1:B3
            img.width = 120
            img.height = 55
            ws.add_image(img, 'A1')
        except Exception:
            ws['A1'] = "SFL LOGO"
            ws['A1'].font = Font(bold=True, size=14)
    else:
        ws['A1'] = "SFL LOGO"
        ws['A1'].font = Font(bold=True, size=14)

    ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
    ws['A1'].border = get_border()

    max_col = ws.max_column if ws.max_column > 14 else 14
    target_col_letter = get_column_letter(max_col)

    ws.merge_cells(f'C1:{get_column_letter(max_col-1)}1')
    title_cell = ws['C1']
    title_cell.value = "Sandip Polytechnic, Nashik"
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center")

    ws.merge_cells(f'C2:{get_column_letter(max_col-1)}2')
    dept_cell = ws['C2']
    dept_cell.value = program_name
    dept_cell.font = Font(bold=True, size=12)
    dept_cell.alignment = Alignment(horizontal="center")

    ws.merge_cells(f'C3:{get_column_letter(max_col-1)}3')
    sub_title_cell = ws['C3']
    sub_title_cell.value = "CO Attainment" if "Analysis" in title_type or "Combine" in title_type else f"Evaluation Sheet for {title_type}"
    sub_title_cell.font = Font(bold=True, size=12)
    sub_title_cell.alignment = Alignment(horizontal="center")
    
    # Right top type label
    type_cell = ws.cell(row=1, column=max_col, value=title_type)
    type_cell.fill = PatternFill(start_color=LIGHT_YELLOW, end_color=LIGHT_YELLOW, fill_type="solid")
    type_cell.alignment = Alignment(horizontal="center")
    type_cell.border = get_border()
    type_cell.font = Font(bold=True)

def add_info_block(ws, course, academic_year, faculty_name, scheme="K", start_row=4):
    labels = [
        ("Academic Year", academic_year),
        ("Semester", f"{course.semester} (IV)" if course.semester == 4 else course.semester),
        ("Scheme", scheme),
        ("Name of Faculty", faculty_name),
        ("Name of Course & Code", f"{course.course_name} {course.course_code}"),
        ("Class & Division", "TYCO")
    ]
    
    for i, (label, value) in enumerate(labels):
        row = start_row + i
        ws.merge_cells(f'A{row}:B{row}')
        ws.merge_cells(f'C{row}:K{row}') # Reduced merge range for alignment
        
        cell_label = ws[f'A{row}']
        cell_label.value = label
        cell_label.font = Font(bold=True)
        cell_label.border = get_border()
        
        cell_val = ws[f'C{row}']
        cell_val.value = value
        cell_val.font = Font(bold=True)
        cell_val.alignment = Alignment(horizontal="left", indent=1) # Set alignment Left
        cell_val.border = get_border()
    
    return start_row + len(labels)

def create_cis_analysis_sheet(wb, course, academic_year, faculty_name, index=0):
    ws = wb.create_sheet("CIS_Analysis", index)
    add_common_header(ws, "CIS_Analysis", faculty_name)
    next_row = add_info_block(ws, course, academic_year, faculty_name)
    
    # Headers Row
    headers = ["CO", "CO Attainment", "Target", "% target achieved", "CO Atmnt.Gap", "Analysis", "Action Proposed Report"]
    current_row = next_row + 2
    for col, header in enumerate(headers, start=2):
        cell = ws.cell(row=current_row, column=col, value=header)
        apply_header_style(cell, fill_color=STAT_GREEN_MEDIUM, font_color=BLACK_TEXT)
        
    # Data Collection
    cos = CO.objects.filter(course_id=course, is_active=True)
    # Robust AY Matching
    ay_clean = academic_year.replace(' ', '') if academic_year else ""
    ay_spaced = ay_clean.replace('-', ' - ')
    ay_query = models.Q(academic_year__icontains=academic_year) | models.Q(academic_year__icontains=ay_clean) | models.Q(academic_year__icontains=ay_spaced)
    
    total_attainment = 0
    total_target = 0
    co_count = 0

    for co in cos:
        att = COAttainment.objects.filter(ay_query, co_id=co).first()
        target = COTarget.objects.filter(ay_query, course_id=course, co_id=co).first()
        
        if att:
            total_attainment += att.overall_attainment
        
        target_val = target.target_value if target else 2.86
        total_target += target_val
        co_count += 1

    # Calculation
    avg_attainment = total_attainment / co_count if co_count > 0 else 0
    avg_target = total_target / co_count if co_count > 0 else 2.86
    
    # Single Row Display
    row = current_row + 1
    
    # Course Code in CO column
    ws.cell(row=row, column=2, value=course.course_code).border = get_border()
    ws.cell(row=row, column=2).alignment = Alignment(horizontal="center")
    
    ws.cell(row=row, column=3, value=round(avg_attainment, 2) if avg_attainment else "-").border = get_border()
    ws.cell(row=row, column=3).alignment = Alignment(horizontal="center")
    
    ws.cell(row=row, column=4, value=round(avg_target, 2)).border = get_border()
    ws.cell(row=row, column=4).alignment = Alignment(horizontal="center")
    ws.cell(row=row, column=4).fill = PatternFill(start_color=STAT_GREEN_MEDIUM, end_color=STAT_GREEN_MEDIUM, fill_type="solid")
    
    perc = (avg_attainment / avg_target * 100) if avg_target and avg_attainment else 0
    ws.cell(row=row, column=5, value=round(perc, 2) if perc else "-").border = get_border()
    ws.cell(row=row, column=5).alignment = Alignment(horizontal="center")
    
    gap = avg_target - avg_attainment if avg_target and avg_attainment else 0
    ws.cell(row=row, column=6, value=round(gap, 2) if gap else "-").border = get_border()
    ws.cell(row=row, column=6).alignment = Alignment(horizontal="center")
    
    analysis = "Course Target is achieved" if gap <= 0 else "Course Target is not achieved"
    ws.cell(row=row, column=7, value=analysis if avg_attainment else "-").border = get_border()
    ws.cell(row=row, column=7).alignment = Alignment(horizontal="center", wrap_text=True)
    
    # ATR Text
    atr_text = course.course_atr or "We will try to achieve the target with extra efforts"
    ws.cell(row=row, column=8, value=atr_text).border = get_border()
    ws.cell(row=row, column=8).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # Footer
    footer_row = row + 2
    ws.cell(row=footer_row, column=2, value="Faculty").font = Font(bold=True)
    ws.cell(row=footer_row, column=5, value="NBA Coordinator").font = Font(bold=True)
    ws.cell(row=footer_row, column=8, value="HOD").font = Font(bold=True)

    # Column Widths
    ws.column_dimensions['B'].width = 8
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 20
    ws.column_dimensions['H'].width = 30

def create_all_combine_sheet(wb, course, academic_year, faculty_name, index=1):
    ws = wb.create_sheet("All Combine ", index)
    add_common_header(ws, "All Combine", faculty_name)
    next_row = add_info_block(ws, course, academic_year, faculty_name)
    
    # Complex multi-level headers (Matching Image 2)
    current_row = next_row + 1
    
    # Internal Assessment Header
    ws.merge_cells(start_row=current_row, start_column=3, end_row=current_row, end_column=8)
    cell_int = ws.cell(row=current_row, column=3, value="Internal Assessment 40%")
    apply_header_style(cell_int, fill_color=LIGHT_YELLOW, font_color=BLACK_TEXT)
    
    # External Assessment Header
    ws.merge_cells(start_row=current_row, start_column=10, end_row=current_row, end_column=12)
    cell_ext = ws.cell(row=current_row, column=10, value="External Assessment 60%")
    apply_header_style(cell_ext, fill_color=LIGHT_YELLOW, font_color=BLACK_TEXT)
    
    # Direct Attainment Header
    ws.merge_cells(start_row=current_row, start_column=14, end_row=current_row, end_column=14)
    cell_dir = ws.cell(row=current_row, column=14, value="Direct Attmnt.")
    apply_header_style(cell_dir, fill_color=LIGHT_YELLOW, font_color=BLACK_TEXT)

    # Course Exit Survey Header
    ws.merge_cells(start_row=current_row, start_column=15, end_row=current_row, end_column=15)
    cell_ces = ws.cell(row=current_row, column=15, value="Course Exit")
    apply_header_style(cell_ces, fill_color=LIGHT_YELLOW, font_color=BLACK_TEXT)
    
    current_row += 1
    # Individual Tool Headers
    headers = [
        "CO", "CT1", "CT2", "Assignment (SLA)", "FA-PR", "SA-PR (Internal)", "Avg(I)",
        " ", "SA-TH", "SA-PR (External)", "Avg(B)",
        " ", ".4A(I)+.6A(B) CO Attainment", "CES (Course Exit Survey)",
        "Total CO Attainment 0.8(L)+0.2 (M)", "Target", "CO Atmnt.Gap", "Target for next Batch"
    ]
    
    col_idx = 2
    for h in headers:
        if h == " ":
            col_idx += 1
            continue
        cell = ws.cell(row=current_row, column=col_idx, value=h)
        apply_header_style(cell, fill_color=STAT_GREEN_MEDIUM if col_idx > 14 else LIGHT_YELLOW, font_color=BLACK_TEXT)
        ws.column_dimensions[get_column_letter(col_idx)].width = 12 if len(h) > 8 else 8
        col_idx += 1
        
    # Data Rows
    # Robust AY Matching
    ay_clean = academic_year.replace(' ', '') if academic_year else ""
    ay_spaced = ay_clean.replace('-', ' - ')
    ay_query = models.Q(academic_year__icontains=academic_year) | models.Q(academic_year__icontains=ay_clean) | models.Q(academic_year__icontains=ay_spaced)
    
    cos = CO.objects.filter(course_id=course, is_active=True).order_by('co_number')
    row = current_row + 1
    
    tool_data = AttainmentService._calculate_detailed_tool_attainment(course.course_id, academic_year)
    co_indirect = AttainmentService._calculate_indirect_co_attainment(course.course_id, academic_year)
    
    for co in cos:
        tools = tool_data.get(co.co_id, {})
        ws.cell(row=row, column=2, value=co.co_number).border = get_border()
        ws.cell(row=row, column=2).alignment = Alignment(horizontal="center")
        
        # Internal
        def get_t_val(base_key, preferred_prefix=None):
            if preferred_prefix:
                k = f"{preferred_prefix}_{base_key}"
                if k in tools:
                    val = tools[k]
                    return val.get('level', "-") if isinstance(val, dict) else val
                return "-"
            
            for p in ['INTERNAL', 'EXTERNAL']:
                k = f"{p}_{base_key}"
                if k in tools:
                    val = tools[k]
                    return val.get('level', "-") if isinstance(val, dict) else val
            
            raw = tools.get(base_key, "-")
            if raw == "-":
                return "-"
            return raw.get('level', "-") if isinstance(raw, dict) else raw

        ct1 = get_t_val('FA_TH_1', 'INTERNAL')
        ct2 = get_t_val('FA_TH_2', 'INTERNAL')
        sla = get_t_val('SLA')
        fa_pr = get_t_val('FA_PR', 'INTERNAL')
        sa_pr_int = get_t_val('SA_PR', 'INTERNAL')
        
        ws.cell(row=row, column=3, value=ct1).border = get_border()
        ws.cell(row=row, column=3).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=4, value=ct2).border = get_border()
        ws.cell(row=row, column=4).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=5, value=sla).border = get_border()
        ws.cell(row=row, column=5).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=6, value=fa_pr).border = get_border()
        ws.cell(row=row, column=6).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=7, value=sa_pr_int).border = get_border()
        ws.cell(row=row, column=7).alignment = Alignment(horizontal="center")
        
        int_vals = [v for v in [ct1, ct2, sla, fa_pr, sa_pr_int] if isinstance(v, (int, float))]
        avg_i = sum(int_vals)/len(int_vals) if int_vals else 0
        ws.cell(row=row, column=8, value=round(avg_i, 2) if avg_i else "-").border = get_border()
        
        # External
        sa_th = get_t_val('SA_TH', 'EXTERNAL')
        sa_pr_ext = get_t_val('SA_PR', 'EXTERNAL')
        
        ws.cell(row=row, column=10, value=sa_th).border = get_border()
        ws.cell(row=row, column=10).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=11, value=sa_pr_ext).border = get_border()
        ws.cell(row=row, column=11).alignment = Alignment(horizontal="center")
        
        ext_vals = [v for v in [sa_th, sa_pr_ext] if isinstance(v, (int, float))]
        avg_b = sum(ext_vals)/len(ext_vals) if ext_vals else 0
        ws.cell(row=row, column=12, value=round(avg_b, 2) if avg_b else "-").border = get_border()
        
        # CO Attainment (Direct)
        direct_att = 0.4 * avg_i + 0.6 * avg_b
        ws.cell(row=row, column=14, value=round(direct_att, 2) if direct_att else "-").border = get_border()
        
        # CES
        ces_val = co_indirect.get(co.co_id, "-")
        ws.cell(row=row, column=15, value=round(ces_val, 2) if isinstance(ces_val, (int, float)) else "-").border = get_border()
        
        # Total
        total_att = 0.8 * direct_att + 0.2 * (ces_val if isinstance(ces_val, (int, float)) else 0)
        ws.cell(row=row, column=16, value=round(total_att, 2) if total_att else "-").border = get_border()
        
        target = COTarget.objects.filter(ay_query, course_id=course, co_id=co).first()
        t_val = target.target_value if target else 2.86
        ws.cell(row=row, column=17, value=t_val).border = get_border()
        ws.cell(row=row, column=17).fill = PatternFill(start_color=STAT_GREEN_MEDIUM, end_color=STAT_GREEN_MEDIUM, fill_type="solid")
        
        gap = t_val - total_att
        ws.cell(row=row, column=18, value=round(gap, 2)).border = get_border()
        ws.cell(row=row, column=19, value=t_val).border = get_border()
        ws.cell(row=row, column=19).fill = PatternFill(start_color=STAT_GREEN_MEDIUM, end_color=STAT_GREEN_MEDIUM, fill_type="solid")
        
        row += 1
    
    # Summary Average Row (Attainment of CO)
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=15)
    sum_lbl = ws.cell(row=row, column=2, value="Attainment of CO")
    sum_lbl.alignment = Alignment(horizontal="right")
    sum_lbl.font = Font(bold=True)
    sum_lbl.fill = PatternFill(start_color=STAT_ORANGE, end_color=STAT_ORANGE, fill_type="solid")
    sum_lbl.border = get_border()

    start_data_row = current_row + 1
    end_data_row = row - 1
    
    if end_data_row >= start_data_row:
        # Formulas for Total Attainment Average and Gap Average
        total_att_range = f"{get_column_letter(16)}{start_data_row}:{get_column_letter(16)}{end_data_row}"
        gap_range = f"{get_column_letter(18)}{start_data_row}:{get_column_letter(18)}{end_data_row}"
        
        
        all_att = []
        all_gaps = []
        for co_item in cos:
            t_data = tool_data.get(co_item.co_id, {})
            def local_get_t_val(base_key, preferred_prefix=None):
                if preferred_prefix:
                    k = f"{preferred_prefix}_{base_key}"
                    if k in t_data:
                        val = t_data[k]
                        return val.get('level', 0) if isinstance(val, dict) else val
                    return 0
                for p in ['INTERNAL', 'EXTERNAL']:
                    k = f"{p}_{base_key}"
                    if k in t_data:
                        val = t_data[k]
                        return val.get('level', 0) if isinstance(val, dict) else val
                raw = t_data.get(base_key, 0)
                if raw == 0:
                    return 0
                return raw.get('level', 0) if isinstance(raw, dict) else (raw if isinstance(raw, (int, float)) else 0)

            # Re-calculate exactly as in the loop
            l_ct1 = local_get_t_val('FA_TH_1', 'INTERNAL')
            l_ct2 = local_get_t_val('FA_TH_2', 'INTERNAL')
            l_sla = local_get_t_val('SLA')
            l_fa_pr = local_get_t_val('FA_PR', 'INTERNAL')
            l_sa_pr_int = local_get_t_val('SA_PR', 'INTERNAL')
            l_int_vals = [v for v in [l_ct1, l_ct2, l_sla, l_fa_pr, l_sa_pr_int] if isinstance(v, (int, float))]
            l_avg_i = sum(l_int_vals)/len(l_int_vals) if l_int_vals else 0
            
            l_sa_th = local_get_t_val('SA_TH', 'EXTERNAL')
            l_sa_pr_ext = local_get_t_val('SA_PR', 'EXTERNAL')
            l_ext_vals = [v for v in [l_sa_th, l_sa_pr_ext] if isinstance(v, (int, float))]
            l_avg_b = sum(l_ext_vals)/len(l_ext_vals) if l_ext_vals else 0
            
            l_direct_att = 0.4 * l_avg_i + 0.6 * l_avg_b
            l_ces_val = co_indirect.get(co_item.co_id, 0)
            if not isinstance(l_ces_val, (int, float)): l_ces_val = 0
            l_total_att = 0.8 * l_direct_att + 0.2 * l_ces_val
            all_att.append(l_total_att)
            
            l_target = COTarget.objects.filter(ay_query, course_id=course, co_id=co_item).first()
            l_t_val = l_target.target_value if l_target else 2.86
            all_gaps.append(l_t_val - l_total_att)

        avg_total_att = sum(all_att) / len(all_att) if all_att else 0
        avg_gap = sum(all_gaps) / len(all_gaps) if all_gaps else 0
        
        att_avg_cell = ws.cell(row=row, column=16, value=round(avg_total_att, 2))
        gap_avg_cell = ws.cell(row=row, column=18, value=round(avg_gap, 2))
        
        for c_idx in [16, 17, 18, 19]:
            cell = ws.cell(row=row, column=c_idx)
            cell.border = get_border()
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color=CT_YELLOW, end_color=CT_YELLOW, fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
            if c_idx == 17: cell.value = "" # Gap spacing
            if c_idx == 19: cell.value = "" # Target spacing
    
    row += 1
        
    # Legend Row
    legend_row = row + 1
    legend = "CT1 & CT2 : Class Test 1 & 2    SLA: Self Learning Assessments    FA-PR: Formative Assessment Practical    SA-PR:Summative Assessment Practical    SA-TH: Summative Assessment Theory"
    ws.merge_cells(start_row=legend_row, start_column=2, end_row=legend_row, end_column=19)
    legend_cell = ws.cell(row=legend_row, column=2, value=legend)
    legend_cell.font = Font(size=9, bold=True)
    legend_cell.fill = PatternFill(start_color=CO_BLUE, end_color=CO_BLUE, fill_type="solid")
    legend_cell.alignment = Alignment(horizontal="center")
    legend_cell.border = get_border()

def create_marks_sheet(wb, title, assessment_type, course, academic_year, students, faculty_name, index):
    ws = wb.create_sheet(title, index)
    add_common_header(ws, title, faculty_name)
    next_row = add_info_block(ws, course, academic_year, faculty_name)
    
    headers = ["Enrollment No.", "Roll no.", "Name of Student", f"Marks (out of ...)"]
    if assessment_type == "SA_TH":
        headers[3] = "Theory Marks out of 70 (SA-TH)"
    elif assessment_type == "SA_PR":
        headers[3] = f"Marks out of 25 ({title})"
    
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=next_row, column=col, value=header)
        apply_header_style(cell, fill_color=STAT_GREEN_MEDIUM if col == 4 else HEADER_LIGHT_BLUE, font_color=BLACK_TEXT)
    
    current_row = next_row + 1
    marks_list = []
    absent_count = 0
    
    # Flexible assessment lookup
    ay_clean = academic_year.replace(' ', '')
    ay_spaced = ay_clean.replace('-', ' - ')
    ay_query = models.Q(academic_year__icontains=academic_year) | models.Q(academic_year__icontains=ay_clean) | models.Q(academic_year__icontains=ay_spaced)
    
    # Try multiple naming patterns
    assessment = Assessment.objects.filter(
        ay_query,
        course_id=course, 
        assessment_type=assessment_type
    ).filter(
        models.Q(assessment_name__iexact=title) | 
        models.Q(assessment_name__icontains=title) |
        models.Q(assessment_name__icontains=assessment_type.replace('_', '-'))
    ).first()
    
    for student in students:
        ws.cell(row=current_row, column=1, value=student.enrollment_no).border = get_border()
        ws.cell(row=current_row, column=2, value=student.roll_no).border = get_border()
        ws.cell(row=current_row, column=3, value=student.name).border = get_border()
        ws.cell(row=current_row, column=3).alignment = Alignment(horizontal="left", indent=1)
        
        marks_entry = MarksEntry.objects.filter(assessment_id=assessment, student_id=student).first() if assessment else None
        marks = marks_entry.marks_obtained if marks_entry else None
        
        if marks is not None:
            marks_list.append(marks)
            val = marks
        else:
            absent_count += 1
            val = "-"
            
        cell_marks = ws.cell(row=current_row, column=4, value=val)
        cell_marks.border = get_border()
        cell_marks.alignment = Alignment(horizontal="center")
        current_row += 1
    
    # Stats Footer (Threshold-based attainment)
    avg = sum(marks_list)/len(marks_list) if marks_list else 0
    appeared = len(marks_list)
    max_m = assessment.max_marks or 100
    
    tp_norm = assessment_type.upper().replace('-', '').replace('_', '')
    is_practical = any(x in tp_norm for x in ['PR', 'SLA', 'PRACTICAL']) and 'SATH' not in tp_norm
    
    if is_practical:
        success_count = len([m for m in marks_list if m >= avg])
    else:
        # Theory Assessments use 40% binary threshold
        threshold = max_m * 0.4
        success_count = len([m for m in marks_list if m >= threshold])
        
    perc = (success_count / appeared * 100) if appeared > 0 else 0
    att_level = round((perc / 100) * 3, 2)
    
    stats = [
        ("Average", round(avg, 2)),
        ("Total Students Appeared", appeared),
        ("Absent", absent_count),
        ("Max Marks", max_m),
        ("CO Attainment", att_level)
    ]
    
    for label, value in stats:
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=3)
        lbl = ws.cell(row=current_row, column=1, value=label)
        lbl.fill = PatternFill(start_color=STAT_ORANGE, end_color=STAT_ORANGE, fill_type="solid")
        lbl.border = get_border()
        lbl.font = Font(bold=True)
        lbl.alignment = Alignment(horizontal="right", indent=1)
        
        val_cell = ws.cell(row=current_row, column=4, value=value)
        val_cell.border = get_border()
        val_cell.alignment = Alignment(horizontal="center")
        val_cell.font = Font(bold=True)
        current_row += 1

    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 25

def create_fa_pr_sheet(wb, course, academic_year, students, faculty_name, index=4):
    ws = wb.create_sheet("FA-PR (K3)", index)
    add_common_header(ws, "FA-PR (K3)", faculty_name)
    next_row = add_info_block(ws, course, academic_year, faculty_name)
    
    # Flexible assessment lookup
    ay_clean = academic_year.replace(' ', '')
    ay_spaced = ay_clean.replace('-', ' - ')
    ay_query = models.Q(academic_year__icontains=academic_year) | models.Q(academic_year__icontains=ay_clean) | models.Q(academic_year__icontains=ay_spaced)
    
    assessment = Assessment.objects.filter(
        ay_query,
        course_id=course, 
        assessment_type='FA_PR'
    ).first()

    config = assessment.configuration if assessment and assessment.configuration else {}
    
    # Dynamically determine practicals count
    marks_data = config.get('marksData', {})
    q_indices = set()
    for student_marks in marks_data.values():
        for q_idx in student_marks.keys():
            if q_idx.isdigit(): q_indices.add(int(q_idx))
    
    max_idx = max(q_indices) + 1 if q_indices else 0
    
    practicals = []
    raw_practicals = config.get('customQuestions', [])
    for i in range(max_idx if max_idx > 0 else 15):
        p_name = raw_practicals[i] if i < len(raw_practicals) and raw_practicals[i] else f"P{i+1}"
        practicals.append(p_name)

    # Sanitize Weights
    raw_weights = config.get('customWeights', [])
    weights = []
    for i in range(len(practicals)):
        w = raw_weights[i] if i < len(raw_weights) else None
        if w in [None, '', 0, '0']: w = 25 # Default for FA-PR
        weights.append(w)

    cos_map = config.get('userCos', [f"CO{(i//3)+1}" for i in range(len(practicals))])

    # 3-Row Header for FA-PR
    ws.cell(row=next_row+1, column=1, value="ENROLLMENT NO.").border = get_border()
    ws.cell(row=next_row+1, column=2, value="Roll no.").border = get_border()
    ws.cell(row=next_row+1, column=3, value="Name of Student").border = get_border()
    for c in [1,2,3]: ws.merge_cells(start_row=next_row+1, start_column=c, end_row=next_row+3, end_column=c)

    ws.cell(row=next_row+1, column=4, value="Practical No.").border = get_border()
    ws.cell(row=next_row+2, column=4, value="Max Marks").border = get_border()
    ws.cell(row=next_row+3, column=4, value="Course Outcome").border = get_border()
    for r in [1,2,3]: ws.cell(row=next_row+r, column=4).fill = PatternFill(start_color=HEADER_DARK_BLUE, end_color=HEADER_DARK_BLUE, fill_type="solid"); ws.cell(row=next_row+r, column=4).font = Font(color=WHITE_TEXT, bold=True)

    for i, p in enumerate(practicals):
        col = 5 + i
        ws.cell(row=next_row+1, column=col, value=p).border = get_border()
        ws.cell(row=next_row+2, column=col, value=weights[i]).border = get_border()
        ws.cell(row=next_row+3, column=col, value=cos_map[i] if i < len(cos_map) else "-").border = get_border()
        for r in [1,2,3]:
             c = ws.cell(row=next_row+r, column=col)
             c.alignment = Alignment(horizontal="center")
             c.fill = PatternFill(start_color=HEADER_LIGHT_BLUE if r==1 else HEADER_DARK_BLUE if r==2 else CO_BLUE, end_color=HEADER_LIGHT_BLUE if r==1 else HEADER_DARK_BLUE if r==2 else CO_BLUE, fill_type="solid")

    total_col = 5 + len(practicals)
    ws.cell(row=next_row+1, column=total_col, value="Total Marks").border = get_border()
    ws.merge_cells(start_row=next_row+1, start_column=total_col, end_row=next_row+3, end_column=total_col)
    apply_header_style(ws.cell(row=next_row+1, column=total_col))

    current_row = next_row + 4
    marks_list = []
    q_marks_collector = {i: [] for i in range(len(practicals))} # For question-wise stats
    
    for student in students:
        ws.cell(row=current_row, column=1, value=student.enrollment_no).border = get_border()
        ws.cell(row=current_row, column=2, value=student.roll_no).border = get_border()
        ws.cell(row=current_row, column=3, value=student.name).border = get_border()
        ws.cell(row=current_row, column=3).alignment = Alignment(horizontal="left", indent=1)
        
        marks_entry = MarksEntry.objects.filter(assessment_id=assessment, student_id=student).first() if assessment else None
        total_marks = marks_entry.marks_obtained if marks_entry else None
        
        marks_data = config.get('marksData', {})
        student_marks = marks_data.get(str(student.enrollment_no), {})
        
        for i in range(len(practicals)):
            q_val = student_marks.get(str(i), "-")
            
            # Type safety: Write as number if possible
            cell_val = q_val
            if q_val not in ["-", "", None]:
                try: cell_val = float(q_val)
                except: pass
                
            ws.cell(row=current_row, column=5+i, value=cell_val).border = get_border()
            ws.cell(row=current_row, column=5+i).alignment = Alignment(horizontal="center")
            
            if q_val not in ["-", "", None]:
                try: q_marks_collector[i].append(float(q_val))
                except: pass

        ws.cell(row=current_row, column=total_col, value=total_marks if total_marks is not None else "-").border = get_border()
        ws.cell(row=current_row, column=total_col).alignment = Alignment(horizontal="center")
        if total_marks is not None: marks_list.append(total_marks)
        current_row += 1

    # Calculate stats in Python instead of Excel formulas
    appeared_count = len(marks_list)
    avg_total = sum(marks_list) / appeared_count if appeared_count > 0 else 0
    
    # Average Row
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=3)
    lbl_cell = ws.cell(row=current_row, column=1, value="Average")
    lbl_cell.fill = PatternFill(start_color=STAT_ORANGE, end_color=STAT_ORANGE, fill_type="solid")
    lbl_cell.border = get_border(); lbl_cell.font = Font(bold=True)
    lbl_cell.alignment = Alignment(horizontal="right", indent=1)

    for i in range(len(practicals)):
        q_col = 5 + i
        q_avg = sum(q_marks_collector[i]) / len(q_marks_collector[i]) if q_marks_collector[i] else 0
        ws.cell(row=current_row, column=q_col, value=round(q_avg, 2)).border = get_border()
        ws.cell(row=current_row, column=q_col).alignment = Alignment(horizontal="center")
    
    ws.cell(row=current_row, column=total_col, value=round(avg_total, 2)).border = get_border()
    ws.cell(row=current_row, column=total_col).alignment = Alignment(horizontal="center")
    current_row += 1

    # Total Appeared Row
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=3)
    lbl_cell = ws.cell(row=current_row, column=1, value="Total Appeared")
    lbl_cell.fill = PatternFill(start_color=STAT_ORANGE, end_color=STAT_ORANGE, fill_type="solid")
    lbl_cell.border = get_border(); lbl_cell.font = Font(bold=True)
    lbl_cell.alignment = Alignment(horizontal="right", indent=1)

    for i in range(len(practicals)):
        q_col = 5 + i
        ws.cell(row=current_row, column=q_col, value=len(q_marks_collector[i])).border = get_border()
        ws.cell(row=current_row, column=q_col).alignment = Alignment(horizontal="center")
    
    ws.cell(row=current_row, column=total_col, value=appeared_count).border = get_border()
    ws.cell(row=current_row, column=total_col).alignment = Alignment(horizontal="center")
    current_row += 1

    # % above Avg Row
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=3)
    lbl_cell = ws.cell(row=current_row, column=1, value="% above Avg")
    lbl_cell.fill = PatternFill(start_color=STAT_ORANGE, end_color=STAT_ORANGE, fill_type="solid")
    lbl_cell.border = get_border(); lbl_cell.font = Font(bold=True)
    lbl_cell.alignment = Alignment(horizontal="right", indent=1)

    for i in range(len(practicals)):
        q_col = 5 + i
        q_list = q_marks_collector[i]
        q_avg = sum(q_list) / len(q_list) if q_list else 0
        success = len([m for m in q_list if m >= q_avg]) if q_list else 0
        perc = (success / len(q_list) * 100) if q_list else 0
        ws.cell(row=current_row, column=q_col, value=f"{round(perc, 2)}%").border = get_border()
        ws.cell(row=current_row, column=q_col).alignment = Alignment(horizontal="center")

    total_success = len([m for m in marks_list if m >= avg_total]) if marks_list else 0
    total_perc = (total_success / appeared_count * 100) if appeared_count > 0 else 0
    ws.cell(row=current_row, column=total_col, value=f"{round(total_perc, 2)}%").border = get_border()
    ws.cell(row=current_row, column=total_col).alignment = Alignment(horizontal="center")
    current_row += 1

    # CO Attainment Row
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=3)
    att_lbl = ws.cell(row=current_row, column=1, value="CO Attainment")
    att_lbl.fill = PatternFill(start_color=STAT_ORANGE, end_color=STAT_ORANGE, fill_type="solid")
    att_lbl.border = get_border(); att_lbl.font = Font(bold=True); att_lbl.alignment = Alignment(horizontal="right", indent=1)
    
    for i in range(len(practicals)):
        q_marks = q_marks_collector[i]
        q_avg = sum(q_marks)/len(q_marks) if q_marks else 0
        
        q_pass = len([m for m in q_marks if m >= q_avg])
        q_perc = (q_pass / len(q_marks) * 100) if q_marks else 0
        ws.cell(row=current_row, column=5+i, value=round(q_perc * 3 / 100, 2)).border = get_border()
        ws.cell(row=current_row, column=5+i).alignment = Alignment(horizontal="center")
    
    # Total column attainment (Average threshold)
    appeared = len(marks_list)
    avg_total = sum(marks_list)/len(marks_list) if marks_list else 0
    if appeared > 0:
        pass_above_avg = len([m for m in marks_list if m >= avg_total])
        perc_above_avg = (pass_above_avg / appeared * 100)
    else:
        perc_above_avg = 0
    att_level = round(perc_above_avg * 3 / 100, 2)
    total_att_cell = ws.cell(row=current_row, column=total_col, value=att_level)
    total_att_cell.border = get_border()
    total_att_cell.alignment = Alignment(horizontal="center")
    total_att_cell.font = Font(bold=True)
    current_row += 2

    # CO Attainment Table (Image 2)
    co_stats = {}
    for i, co_val in enumerate(cos_map):
        if i >= len(practicals): break
        co_key = f"CO{co_val}" if not str(co_val).upper().startswith("CO") else str(co_val).upper()
        if co_key not in co_stats: co_stats[co_key] = []
        co_stats[co_key].extend(q_marks_collector[i])

    ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=4)
    tbl_hdr = ws.cell(row=current_row, column=2, value="CO attainment through FA-PR")
    apply_header_style(tbl_hdr, fill_color=CT_YELLOW, font_color=BLACK_TEXT)
    current_row += 1
    
    ws.cell(row=current_row, column=2, value="% CO Attained").border = get_border()
    apply_header_style(ws.cell(row=current_row, column=2), fill_color=HEADER_LIGHT_PINK, font_color=BLACK_TEXT)
    ws.merge_cells(start_row=current_row, start_column=3, end_row=current_row, end_column=4)
    ws.cell(row=current_row, column=3, value="CO attainment").border = get_border()
    apply_header_style(ws.cell(row=current_row, column=3), fill_color=HEADER_LIGHT_PINK, font_color=BLACK_TEXT)
    current_row += 1
    
    for co_name in sorted(co_stats.keys()):
        c_marks = co_stats[co_name]
        c_avg = sum(c_marks)/len(c_marks) if c_marks else 0
        
        c_pass = len([m for m in c_marks if m >= c_avg])
        c_perc = (c_pass / len(c_marks) * 100) if c_marks else 0
        c_level = round((c_perc * 3 / 100), 2)
        
        ws.cell(row=current_row, column=2, value=co_name).border = get_border()
        ws.cell(row=current_row, column=2).alignment = Alignment(horizontal="center")
        ws.cell(row=current_row, column=3, value=round(c_perc, 2)).border = get_border()
        ws.cell(row=current_row, column=3).fill = PatternFill(start_color=STAT_GREEN_MEDIUM, end_color=STAT_GREEN_MEDIUM, fill_type="solid")
        ws.cell(row=current_row, column=3).alignment = Alignment(horizontal="center")
        ws.cell(row=current_row, column=4, value=c_level).border = get_border()
        ws.cell(row=current_row, column=4).alignment = Alignment(horizontal="center")
        current_row += 1

    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 35
    for i in range(len(practicals)): ws.column_dimensions[get_column_letter(5+i)].width = 6

def create_ct_sheet(wb, ct_num, course, academic_year, students, faculty_name, index):
    title = f"Class test 0{ct_num}" # Match sheet name in Image 2
    ws = wb.create_sheet(title, index)
    add_common_header(ws, title, faculty_name)
    next_row = add_info_block(ws, course, academic_year, faculty_name)
    
    ay_clean = academic_year.replace(' ', '')
    ay_spaced = ay_clean.replace('-', ' - ')
    ay_query = models.Q(academic_year__icontains=academic_year) | models.Q(academic_year__icontains=ay_clean) | models.Q(academic_year__icontains=ay_spaced)
    
    assessment = Assessment.objects.filter(
        ay_query,
        course_id=course, 
        assessment_type='FA_TH'
    ).filter(
        models.Q(assessment_name__icontains=f"CT{ct_num}") | 
        models.Q(assessment_name__icontains=f"Test {ct_num}") |
        models.Q(assessment_name__icontains=f"Class Test {ct_num}")
    ).first()

    config = assessment.configuration if assessment and assessment.configuration else {}
    
    # Dynamically determine questions count
    marks_data = config.get('marksData', {})
    q_indices = set()
    for student_marks in marks_data.values():
        for q_idx in student_marks.keys():
            if q_idx.isdigit(): q_indices.add(int(q_idx))
    
    max_idx = max(q_indices) + 1 if q_indices else 0
    
    # Sanitize Questions
    raw_questions = config.get('customQuestions', [])
    questions = []
    for i in range(max_idx if max_idx > 0 else 14):
        q_name = raw_questions[i] if i < len(raw_questions) and raw_questions[i] else f"Q{i+1}"
        questions.append(q_name)

    # Sanitize Weights
    raw_weights = config.get('customWeights', [])
    weights = []
    for i in range(len(questions)):
        w = raw_weights[i] if i < len(raw_weights) else None
        if w in [None, '', 0, '0']: w = 2 # Default for CT
        weights.append(w)

    cos_map = config.get('userCos', [1]*len(questions))
    
    h_row = next_row + 1
    # 3-Row Header for CT (A-C)
    ws.cell(row=h_row, column=1, value="ENROLLMENT NO.").border = get_border()
    ws.cell(row=h_row, column=2, value="Roll no.").border = get_border()
    ws.cell(row=h_row, column=3, value="Name of Student").border = get_border()
    for c in [1,2,3]: 
        ws.merge_cells(start_row=h_row, start_column=c, end_row=h_row+2, end_column=c)
        apply_header_style(ws.cell(row=h_row, column=c), fill_color=HEADER_LIGHT_BLUE, font_color=BLACK_TEXT)

    # Labels Col (D)
    ws.cell(row=h_row, column=4, value="Q->").fill = PatternFill(start_color=CT_YELLOW, end_color=CT_YELLOW, fill_type="solid")
    ws.cell(row=h_row+1, column=4, value="Wt->").fill = PatternFill(start_color=CT_YELLOW, end_color=CT_YELLOW, fill_type="solid")
    ws.cell(row=h_row+2, column=4, value="CO ->").fill = PatternFill(start_color=CT_YELLOW, end_color=CT_YELLOW, fill_type="solid")
    for r in range(h_row, h_row+3): 
        c_lbl = ws.cell(row=r, column=4)
        c_lbl.border = get_border()
        c_lbl.font = Font(bold=True)
        c_lbl.alignment = Alignment(horizontal="center")

    for i, q in enumerate(questions):
        col = 5 + i
        ws.cell(row=h_row, column=col, value=q).border = get_border()
        ws.cell(row=h_row+1, column=col, value=weights[i]).border = get_border()
        ws.cell(row=h_row+2, column=col, value=cos_map[i] if i < len(cos_map) else "-").border = get_border()
        for r in range(h_row, h_row+3): 
            ws.cell(row=r, column=col).fill = PatternFill(start_color=CT_LIGHT_BLUE if r == h_row else HEADER_DARK_BLUE if r == h_row+1 else HEADER_LIGHT_BLUE, end_color=CT_LIGHT_BLUE if r == h_row else HEADER_DARK_BLUE if r == h_row+1 else HEADER_LIGHT_BLUE, fill_type="solid")
            ws.cell(row=r, column=col).alignment = Alignment(horizontal="center")

    # Column for Total
    total_col = 5 + len(questions)
    ws.cell(row=h_row, column=total_col, value="Total").border = get_border()
    apply_header_style(ws.cell(row=h_row, column=total_col), fill_color=HEADER_DARK_BLUE)
    ws.merge_cells(start_row=h_row, start_column=total_col, end_row=h_row+2, end_column=total_col)
    
    current_row = h_row + 3
    marks_list = []
    q_marks_collector = {i: [] for i in range(len(questions))} # For question-wise stats
    
    for student in students:
        ws.cell(row=current_row, column=1, value=student.enrollment_no).border = get_border()
        ws.cell(row=current_row, column=2, value=student.roll_no).border = get_border()
        ws.cell(row=current_row, column=3, value=student.name).border = get_border()
        ws.cell(row=current_row, column=3).alignment = Alignment(horizontal="left", indent=1)
        
        marks_entry = MarksEntry.objects.filter(assessment_id=assessment, student_id=student).first() if assessment else None
        total_val = marks_entry.marks_obtained if marks_entry else None
        
        marks_data = config.get('marksData', {})
        student_marks = marks_data.get(str(student.enrollment_no), {})

        # Fill granular marks from config
        for i in range(len(questions)):
            q_val = student_marks.get(str(i), "-")
            
            # Type safety: Write as number if possible
            cell_val = q_val
            if q_val not in ["-", "", None]:
                try: cell_val = float(q_val)
                except: pass
                
            ws.cell(row=current_row, column=5+i, value=cell_val).border = get_border()
            ws.cell(row=current_row, column=5+i).alignment = Alignment(horizontal="center")
            
            if q_val not in ["-", "", None]:
                try: q_marks_collector[i].append(float(q_val))
                except: pass
            
        ws.cell(row=current_row, column=total_col, value=total_val if total_val is not None else "-").border = get_border()
        ws.cell(row=current_row, column=total_col).alignment = Alignment(horizontal="center")
        if total_val is not None: marks_list.append(total_val)
        current_row += 1

    # Calculate stats in Python instead of Excel formulas
    appeared_count = len(marks_list)
    avg_total = sum(marks_list) / appeared_count if appeared_count > 0 else 0
    
    # Average Row
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=3)
    lbl_cell = ws.cell(row=current_row, column=1, value="Average")
    lbl_cell.fill = PatternFill(start_color=STAT_ORANGE, end_color=STAT_ORANGE, fill_type="solid")
    lbl_cell.border = get_border(); lbl_cell.font = Font(bold=True)
    lbl_cell.alignment = Alignment(horizontal="right", indent=1)

    for i in range(len(questions)):
        q_col = 5 + i
        q_avg = sum(q_marks_collector[i]) / len(q_marks_collector[i]) if q_marks_collector[i] else 0
        ws.cell(row=current_row, column=q_col, value=round(q_avg, 2)).border = get_border()
        ws.cell(row=current_row, column=q_col).alignment = Alignment(horizontal="center")
    
    ws.cell(row=current_row, column=total_col, value=round(avg_total, 2)).border = get_border()
    ws.cell(row=current_row, column=total_col).alignment = Alignment(horizontal="center")
    current_row += 1

    # Total Appeared Row
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=3)
    lbl_cell = ws.cell(row=current_row, column=1, value="Total Appeared")
    lbl_cell.fill = PatternFill(start_color=STAT_ORANGE, end_color=STAT_ORANGE, fill_type="solid")
    lbl_cell.border = get_border(); lbl_cell.font = Font(bold=True)
    lbl_cell.alignment = Alignment(horizontal="right", indent=1)

    for i in range(len(questions)):
        q_col = 5 + i
        ws.cell(row=current_row, column=q_col, value=len(q_marks_collector[i])).border = get_border()
        ws.cell(row=current_row, column=q_col).alignment = Alignment(horizontal="center")
    
    ws.cell(row=current_row, column=total_col, value=appeared_count).border = get_border()
    ws.cell(row=current_row, column=total_col).alignment = Alignment(horizontal="center")
    current_row += 1

    # % above Avg Row
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=3)
    lbl_cell = ws.cell(row=current_row, column=1, value="% above Avg")
    lbl_cell.fill = PatternFill(start_color=STAT_ORANGE, end_color=STAT_ORANGE, fill_type="solid")
    lbl_cell.border = get_border(); lbl_cell.font = Font(bold=True)
    lbl_cell.alignment = Alignment(horizontal="right", indent=1)

    for i in range(len(questions)):
        q_col = 5 + i
        q_list = q_marks_collector[i]
        q_avg = sum(q_list) / len(q_list) if q_list else 0
        success = len([m for m in q_list if m >= q_avg]) if q_list else 0
        perc = (success / len(q_list) * 100) if q_list else 0
        ws.cell(row=current_row, column=q_col, value=f"{round(perc, 2)}%").border = get_border()
        ws.cell(row=current_row, column=q_col).alignment = Alignment(horizontal="center")

    total_success = len([m for m in marks_list if m >= avg_total]) if marks_list else 0
    total_perc = (total_success / appeared_count * 100) if appeared_count > 0 else 0
    ws.cell(row=current_row, column=total_col, value=f"{round(total_perc, 2)}%").border = get_border()
    ws.cell(row=current_row, column=total_col).alignment = Alignment(horizontal="center")
    current_row += 1

    # Attainment Level Row (Requires Python logic for the threshold or a complex IF formula)
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=3)
    att_lbl = ws.cell(row=current_row, column=1, value="CO Attainment")
    att_lbl.fill = PatternFill(start_color=STAT_ORANGE, end_color=STAT_ORANGE, fill_type="solid")
    att_lbl.border = get_border(); att_lbl.font = Font(bold=True); att_lbl.alignment = Alignment(horizontal="right", indent=1)
    
    # Python-calculated attainment level based on 40% threshold for CT
    appeared = len(marks_list)
    avg = sum(marks_list)/len(marks_list) if marks_list else 0
    max_total = assessment.max_marks or 20
    
    threshold_total = max_total * 0.4
    success_total = len([m for m in marks_list if m >= threshold_total])
    perc_total = (success_total / appeared * 100) if appeared > 0 else 0
    att_level = round(perc_total * 3 / 100, 2)
    
    for i in range(len(questions)):
        q_stats_list = q_marks_collector[i]
        q_avg = sum(q_stats_list)/len(q_stats_list) if q_stats_list else 0
        q_max = 2.0
        if i < len(config.get('customWeights', [])):
            try: q_max = float(config.get('customWeights')[i])
            except: pass
        elif assessment.max_marks:
            q_max = assessment.max_marks / len(questions)
            
        q_threshold = q_max * 0.4
        q_success = len([m for m in q_stats_list if m >= q_threshold])
        q_perc = (q_success / len(q_stats_list) * 100) if q_stats_list else 0
        q_level = round(q_perc * 3 / 100, 2)
        
        ws.cell(row=current_row, column=5+i, value=q_level).border = get_border()
        ws.cell(row=current_row, column=5+i).alignment = Alignment(horizontal="center")
        
    ws.cell(row=current_row, column=total_col, value=att_level).border = get_border()
    ws.cell(row=current_row, column=total_col).alignment = Alignment(horizontal="center")
    ws.cell(row=current_row, column=total_col).font = Font(bold=True)
    current_row += 2

    # Image 2: CO Attainment Summary Table (Aggregated by CO)
    co_stats = {} # {co_num: [marks]}
    for i, co_val in enumerate(cos_map):
        if i >= len(questions): break
        co_key = f"CO{co_val}" if not str(co_val).upper().startswith("CO") else str(co_val).upper()
        if co_key not in co_stats: co_stats[co_key] = []
        co_stats[co_key].extend(q_marks_collector[i])
        
    ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=4)
    tbl_hdr = ws.cell(row=current_row, column=2, value="CO attainment through FA-TH")
    apply_header_style(tbl_hdr, fill_color=CT_YELLOW, font_color=BLACK_TEXT)
    current_row += 1
    
    ws.cell(row=current_row, column=2, value="% CO Attained").border = get_border()
    apply_header_style(ws.cell(row=current_row, column=2), fill_color=HEADER_LIGHT_PINK, font_color=BLACK_TEXT)
    ws.merge_cells(start_row=current_row, start_column=3, end_row=current_row, end_column=4)
    ws.cell(row=current_row, column=3, value="CO attainment").border = get_border()
    apply_header_style(ws.cell(row=current_row, column=3), fill_color=HEADER_LIGHT_PINK, font_color=BLACK_TEXT)
    current_row += 1
    
    for co_name in sorted(co_stats.keys()):
        c_marks = co_stats[co_name]
        
        # Determine average max marks for this CO group
        c_q_indices = [idx for idx, cv in enumerate(cos_map) if (f"CO{cv}" if not str(cv).upper().startswith("CO") else str(cv).upper()) == co_name]
        c_max_sum = 0
        for idx in c_q_indices:
            qm = 2.0
            if idx < len(config.get('customWeights', [])):
                try: qm = float(config.get('customWeights')[idx])
                except: pass
            elif assessment.max_marks:
                qm = assessment.max_marks / len(questions)
            c_max_sum += qm
        
        c_max_avg = c_max_sum / len(c_q_indices) if c_q_indices else 2.0
        c_threshold = c_max_avg * 0.4
        c_success = len([m for m in c_marks if m >= c_threshold])
        c_perc = (c_success / len(c_marks) * 100) if c_marks else 0
        c_level = round(c_perc * 3 / 100, 2)
        
        ws.cell(row=current_row, column=2, value=co_name).border = get_border()
        ws.cell(row=current_row, column=2).alignment = Alignment(horizontal="center")
        
        ws.cell(row=current_row, column=3, value=round(c_perc, 2)).border = get_border() # Percentage success
        ws.cell(row=current_row, column=3).fill = PatternFill(start_color=STAT_GREEN_MEDIUM, end_color=STAT_GREEN_MEDIUM, fill_type="solid")
        ws.cell(row=current_row, column=3).alignment = Alignment(horizontal="center")
        
        ws.cell(row=current_row, column=4, value=c_level).border = get_border()
        ws.cell(row=current_row, column=4).alignment = Alignment(horizontal="center")
        current_row += 1

    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 35
    for i in range(len(questions)): ws.column_dimensions[get_column_letter(5+i)].width = 6

def create_sla_sheet(wb, course, academic_year, students, faculty_name, index):
    ws = wb.create_sheet("SLA", index)
    add_common_header(ws, "SLA", faculty_name)
    next_row = add_info_block(ws, course, academic_year, faculty_name)
    
    headers = ["ENROLLMENT NO.", "Roll no.", "Name of Student", "Assignment 01", "Assignment 02", "Assignment 03", "Assignment 04", "Assignment 05", "Assignment 06", "Total"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=next_row, column=col, value=h)
        apply_header_style(cell, fill_color=STAT_GREEN_MEDIUM if col > 3 else HEADER_LIGHT_BLUE, font_color=BLACK_TEXT)
    
    current_row = next_row + 1
    marks_list = []
    absent_count = 0
    q_marks_collector = {i: [] for i in range(6)} # For assignment-wise stats
    
    # Flexible assessment lookup
    ay_clean = academic_year.replace(' ', '')
    ay_spaced = ay_clean.replace('-', ' - ')
    ay_query = models.Q(academic_year__icontains=academic_year) | models.Q(academic_year__icontains=ay_clean) | models.Q(academic_year__icontains=ay_spaced)
    
    assessment = Assessment.objects.filter(
        ay_query,
        course_id=course, 
        assessment_type='SLA'
    ).first()
    
    config = assessment.configuration if assessment and assessment.configuration else {}
    marks_data = config.get('marksData', {})

    for student in students:
        ws.cell(row=current_row, column=1, value=student.enrollment_no).border = get_border()
        ws.cell(row=current_row, column=2, value=student.roll_no).border = get_border()
        ws.cell(row=current_row, column=3, value=student.name).border = get_border()
        ws.cell(row=current_row, column=3).alignment = Alignment(horizontal="left", indent=1)
        
        marks_entry = MarksEntry.objects.filter(assessment_id=assessment, student_id=student).first() if assessment else None
        total_marks = marks_entry.marks_obtained if marks_entry else None
        
        student_marks = marks_data.get(str(student.enrollment_no), {})

        # Fill student marks (SLA)
        # The 'questions' variable is not defined in this scope for SLA.
        # Assuming a fixed number of assignments (6) as per the original code's q_marks_collector initialization.
        # The instruction implies using 'questions', but it's not available.
        # To maintain syntactic correctness and adhere to the spirit of the change (type safety),
        # I will apply the type safety logic to the existing loop structure.
        # If 'questions' was intended to be passed or derived, that would be a larger change.
        # Attempt to find practicals if not explicitly defined
        practicals_count = len(config.get('customQuestions', [])) or 6
        for i in range(practicals_count): # Use dynamic or fallback count
            q_val = student_marks.get(str(i), "-")
            
            cell_val = q_val
            if q_val not in ["-", "", None]:
                try: cell_val = float(q_val)
                except: pass
                
            ws.cell(row=current_row, column=4+i, value=cell_val).border = get_border()
            ws.cell(row=current_row, column=4+i).alignment = Alignment(horizontal="center")
            
            if q_val not in ["-", "", None]:
                try: q_marks_collector[i].append(float(q_val))
                except: pass
            
        if total_marks is not None:
            marks_list.append(total_marks)
            val = total_marks
        else:
            absent_count += 1
            val = "-"
            
        ws.cell(row=current_row, column=10, value=val).border = get_border()
        ws.cell(row=current_row, column=10).alignment = Alignment(horizontal="center")
        current_row += 1

    # Calculate stats in Python instead of Excel formulas
    total_col = 10
    appeared_count = len(marks_list)
    avg_total_val = sum(marks_list) / appeared_count if appeared_count > 0 else 0
    practicals_count = 6 # As per range(6) in original code
    
    # Average Row
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=3)
    lbl_cell = ws.cell(row=current_row, column=1, value="Average")
    lbl_cell.fill = PatternFill(start_color=STAT_ORANGE, end_color=STAT_ORANGE, fill_type="solid")
    lbl_cell.border = get_border(); lbl_cell.font = Font(bold=True)
    lbl_cell.alignment = Alignment(horizontal="right", indent=1)

    for i in range(practicals_count):
        q_col = 4 + i
        q_avg = sum(q_marks_collector[i]) / len(q_marks_collector[i]) if q_marks_collector[i] else 0
        ws.cell(row=current_row, column=q_col, value=round(q_avg, 2)).border = get_border()
        ws.cell(row=current_row, column=q_col).alignment = Alignment(horizontal="center")
    
    ws.cell(row=current_row, column=total_col, value=round(avg_total_val, 2)).border = get_border()
    ws.cell(row=current_row, column=total_col).alignment = Alignment(horizontal="center")
    current_row += 1

    # Total Appeared Row
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=3)
    lbl_cell = ws.cell(row=current_row, column=1, value="Total Appeared")
    lbl_cell.fill = PatternFill(start_color=STAT_ORANGE, end_color=STAT_ORANGE, fill_type="solid")
    lbl_cell.border = get_border(); lbl_cell.font = Font(bold=True)
    lbl_cell.alignment = Alignment(horizontal="right", indent=1)

    for i in range(practicals_count):
        q_col = 4 + i
        ws.cell(row=current_row, column=q_col, value=len(q_marks_collector[i])).border = get_border()
        ws.cell(row=current_row, column=q_col).alignment = Alignment(horizontal="center")
    
    ws.cell(row=current_row, column=total_col, value=appeared_count).border = get_border()
    ws.cell(row=current_row, column=total_col).alignment = Alignment(horizontal="center")
    current_row += 1

    # % above Avg Row
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=3)
    lbl_cell = ws.cell(row=current_row, column=1, value="% above Avg")
    lbl_cell.fill = PatternFill(start_color=STAT_ORANGE, end_color=STAT_ORANGE, fill_type="solid")
    lbl_cell.border = get_border(); lbl_cell.font = Font(bold=True)
    lbl_cell.alignment = Alignment(horizontal="right", indent=1)

    for i in range(practicals_count):
        q_col = 4 + i
        q_list = q_marks_collector[i]
        q_avg = sum(q_list) / len(q_list) if q_list else 0
        success = len([m for m in q_list if m >= q_avg]) if q_list else 0
        perc = (success / len(q_list) * 100) if q_list else 0
        ws.cell(row=current_row, column=q_col, value=f"{round(perc, 2)}%").border = get_border()
        ws.cell(row=current_row, column=q_col).alignment = Alignment(horizontal="center")

    total_success = len([m for m in marks_list if m >= avg_total_val]) if marks_list else 0
    total_perc = (total_success / appeared_count * 100) if appeared_count > 0 else 0
    ws.cell(row=current_row, column=total_col, value=f"{round(total_perc, 2)}%").border = get_border()
    ws.cell(row=current_row, column=total_col).alignment = Alignment(horizontal="center")
    current_row += 1

    # Attainment Level Row
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=3)
    att_lbl = ws.cell(row=current_row, column=1, value="CO Attainment")
    att_lbl.fill = PatternFill(start_color=STAT_ORANGE, end_color=STAT_ORANGE, fill_type="solid")
    att_lbl.border = get_border(); att_lbl.font = Font(bold=True); att_lbl.alignment = Alignment(horizontal="right", indent=1)
    
    # Python-calculated attainment level using average-threshold logic (SLA)
    appeared = len(marks_list)
    avg_total = sum(marks_list)/len(marks_list) if marks_list else 0
    if appeared > 0:
        pass_above_avg = len([m for m in marks_list if m >= avg_total])
        perc_above_avg = (pass_above_avg / appeared * 100)
    else:
        perc_above_avg = 0
    att_level = round(perc_above_avg * 3 / 100, 2)
    
    for i in range(6):
        q_marks = q_marks_collector[i]
        q_avg = sum(q_marks)/len(q_marks) if q_marks else 0
        
        q_pass = len([m for m in q_marks if m >= q_avg])
        q_perc = (q_pass / len(q_marks) * 100) if q_marks else 0
        q_level = round(q_perc * 3 / 100, 2)
        
        ws.cell(row=current_row, column=4+i, value=q_level).border = get_border()
        ws.cell(row=current_row, column=4+i).alignment = Alignment(horizontal="center")
        
    ws.cell(row=current_row, column=total_col, value=att_level).border = get_border()
    ws.cell(row=current_row, column=total_col).alignment = Alignment(horizontal="center")
    current_row += 2

    # CO Attainment Table
    co_stats = {} # Map assignment numbers to COs if possible, or just use 1 CO per tool as fallback
    # In SLA, we often have 1 CO per course or assignment. Assuming direct marks collection.
    for i in range(6):
        co_key = f"CO{i+1}" # Fallback
        if co_key not in co_stats: co_stats[co_key] = []
        co_stats[co_key].extend(q_marks_collector[i])

    ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=4)
    ws.cell(row=current_row, column=2, value="CO attainment through SLA").fill = PatternFill(start_color=CT_YELLOW, end_color=CT_YELLOW, fill_type="solid")
    current_row += 1
    
    ws.cell(row=current_row, column=2, value="% CO Attained").border = get_border()
    ws.merge_cells(start_row=current_row, start_column=3, end_row=current_row, end_column=4)
    ws.cell(row=current_row, column=3, value="CO attainment").border = get_border()
    current_row += 1
    
    for co_name in sorted(co_stats.keys()):
        if not co_stats[co_name]: continue
        c_marks = co_stats[co_name]
        c_avg = sum(c_marks)/len(c_marks) if c_marks else 0
        
        c_pass = len([m for m in c_marks if m >= c_avg])
        c_perc = (c_pass / len(c_marks) * 100) if c_marks else 0
        c_level = round(c_perc * 3 / 100, 2)
        
        ws.cell(row=current_row, column=2, value=co_name).border = get_border()
        ws.cell(row=current_row, column=3, value=round(c_perc, 2)).border = get_border()
        ws.cell(row=current_row, column=3).fill = PatternFill(start_color=STAT_GREEN_MEDIUM, end_color=STAT_GREEN_MEDIUM, fill_type="solid")
        ws.cell(row=current_row, column=4, value=c_level).border = get_border()
        current_row += 1

    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 35
    for i in range(4, 11):
        ws.column_dimensions[get_column_letter(i)].width = 12

def create_ces_sheet(wb, course, academic_year, students, faculty_name, index=8):
    ws = wb.create_sheet("CES. ", index)
    add_common_header(ws, "CES", faculty_name)
    next_row = add_info_block(ws, course, academic_year, faculty_name)
    
    # Flexible survey lookup
    ay_clean = academic_year.replace(' ', '') if academic_year else ""
    ay_short = ay_clean
    if len(ay_clean) == 9 and ay_clean[4] == '-': # 2025-2026
        ay_short = ay_clean[:5] + ay_clean[7:] # 2025-26
        
    ay_spaced = ay_clean.replace('-', ' - ')
    ay_query = models.Q(academic_year__icontains=academic_year) | \
               models.Q(academic_year__icontains=ay_clean) | \
               models.Q(academic_year__icontains=ay_short) | \
               models.Q(academic_year__icontains=ay_spaced)
    
    # Find approved CES survey for this course, prioritized by number of responses
    # (This reliably finds the 'actual' survey among test surveys)
    survey = SurveyMaster.objects.filter(
        course_id=course, 
        survey_category='course_exit',
        status='APPROVED'
    ).annotate(resp_count=Count('responses')).order_by('-resp_count', '-survey_id').first()
    
    # Fallback to latest if none approved
    if not survey:
        survey = SurveyMaster.objects.filter(
            course_id=course, 
            survey_category='course_exit'
        ).annotate(resp_count=Count('responses')).order_by('-resp_count', '-survey_id').first()
    questions = SurveyQuestion.objects.filter(survey_id=survey).order_by('question_id') if survey else []
    
    headers_row1 = ["ENROLLMENT NO.", "Roll no.", "Name of Student"]
    headers_row2 = ["", "", ""]
    
    for q in questions:
        co_num = q.co_id.co_number if q.co_id else "N/A"
        headers_row1.append(str(co_num))
        headers_row2.append(q.question_text)
    
    headers_row1.append("Total")
    headers_row2.append("")

    # Apply Header Styles (2 rows)
    for col, (h1, h2) in enumerate(zip(headers_row1, headers_row2), start=1):
        if col <= 3:
            ws.merge_cells(start_row=next_row, start_column=col, end_row=next_row+1, end_column=col)
            cell = ws.cell(row=next_row, column=col, value=h1)
            apply_header_style(cell, fill_color=HEADER_LIGHT_BLUE, font_color=BLACK_TEXT)
        else:
            cell1 = ws.cell(row=next_row, column=col, value=h1)
            cell2 = ws.cell(row=next_row+1, column=col, value=h2)
            apply_header_style(cell1, fill_color=STAT_GREEN_MEDIUM if h1 != "Total" else HEADER_DARK_BLUE, font_color=BLACK_TEXT if h1 != "Total" else WHITE_TEXT)
            apply_header_style(cell2, fill_color=STAT_GREEN_MEDIUM if h1 != "Total" else HEADER_DARK_BLUE, font_color=BLACK_TEXT if h1 != "Total" else WHITE_TEXT)
            if h1 == "Total":
                ws.merge_cells(start_row=next_row, start_column=col, end_row=next_row+1, end_column=col)

    current_row = next_row + 2
    marks_list = []
    q_marks_collector = {i: [] for i in range(len(questions))} # For CO-wise stats
    
    for student in students:
        ws.cell(row=current_row, column=1, value=student.enrollment_no).border = get_border()
        ws.cell(row=current_row, column=2, value=student.roll_no).border = get_border()
        ws.cell(row=current_row, column=3, value=student.name).border = get_border()
        ws.cell(row=current_row, column=3).alignment = Alignment(horizontal="left", indent=1)
        
        row_total = 0
        row_count = 0
        for i, q in enumerate(questions):
            ans = SurveyAnswer.objects.filter(question_id=q, response_id__student_id=student).first()
            val = ans.answer_value if ans else "-"
            ws.cell(row=current_row, column=4+i, value=val).border = get_border()
            ws.cell(row=current_row, column=4+i).alignment = Alignment(horizontal="center")
            if isinstance(val, (int, float)):
                q_marks_collector[i].append(val)
                row_total += val
                row_count += 1
        
        row_avg = row_total / row_count if row_count > 0 else "-"
        ws.cell(row=current_row, column=4+len(questions), value=round(row_avg, 2) if isinstance(row_avg, float) else row_avg).border = get_border()
        ws.cell(row=current_row, column=4+len(questions)).alignment = Alignment(horizontal="center")
        
        if isinstance(row_avg, (int, float)): marks_list.append(row_avg)
        current_row += 1

    # Statistical Footer
    start_answers_row = next_row + 1
    end_answers_row = current_row - 1
    
    # Calculate stats in Python
    appeared_count = len(marks_list)
    
    # Average Rating Row
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=3)
    lbl_cell = ws.cell(row=current_row, column=1, value="Average Rating")
    lbl_cell.fill = PatternFill(start_color=STAT_ORANGE, end_color=STAT_ORANGE, fill_type="solid")
    lbl_cell.border = get_border(); lbl_cell.font = Font(bold=True)
    lbl_cell.alignment = Alignment(horizontal="right", indent=1)

    for i in range(len(questions)):
        q_col = 4 + i
        q_list = q_marks_collector[i]
        q_avg = sum(q_list) / len(q_list) if q_list else 0
        ws.cell(row=current_row, column=q_col, value=round(q_avg, 2)).border = get_border()
        ws.cell(row=current_row, column=q_col).alignment = Alignment(horizontal="center")
    current_row += 1

    # Total Responses Row
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=3)
    lbl_cell = ws.cell(row=current_row, column=1, value="Total Responses")
    lbl_cell.fill = PatternFill(start_color=STAT_ORANGE, end_color=STAT_ORANGE, fill_type="solid")
    lbl_cell.border = get_border(); lbl_cell.font = Font(bold=True)
    lbl_cell.alignment = Alignment(horizontal="right", indent=1)

    for i in range(len(questions)):
        q_col = 4 + i
        ws.cell(row=current_row, column=q_col, value=len(q_marks_collector[i])).border = get_border()
        ws.cell(row=current_row, column=q_col).alignment = Alignment(horizontal="center")
    current_row += 1
    
    current_row += 1
    # CO Summary Table for CES
    ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=4)
    ws.cell(row=current_row, column=2, value="CO attainment through CES").fill = PatternFill(start_color=CT_YELLOW, end_color=CT_YELLOW, fill_type="solid")
    current_row += 1
    
    ws.cell(row=current_row, column=2, value="Course Outcome").border = get_border()
    ws.merge_cells(start_row=current_row, start_column=3, end_row=current_row, end_column=4)
    ws.cell(row=current_row, column=3, value="Avg Rating").border = get_border()
    current_row += 1

    for i, q in enumerate(questions):
        q_marks = q_marks_collector[i]
        q_avg = sum(q_marks)/len(q_marks) if q_marks else 0
        # Survey rating is typically 1-5, scaling to 3
        q_level = round((q_avg / 5) * 3, 2)
        ws.cell(row=current_row, column=2, value=f"CO{i+1}").border = get_border()
        ws.cell(row=current_row, column=3, value=round(q_avg, 2)).border = get_border()
        ws.cell(row=current_row, column=3).fill = PatternFill(start_color=STAT_GREEN_MEDIUM, end_color=STAT_GREEN_MEDIUM, fill_type="solid")
        ws.cell(row=current_row, column=4, value=q_level).border = get_border()
        current_row += 1
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 35
    for i in range(4, 4+len(questions)+1): ws.column_dimensions[get_column_letter(i)].width = 10

def create_mapping_sheet(wb, course, academic_year, faculty_name, index=9):
    ws = wb.create_sheet("CO-PO-PSO Mapping & PO-PSO Att.", index)
    add_common_header(ws, "Mapping & PO-PSO Att.", faculty_name)

    next_row = add_info_block(ws, course, academic_year, faculty_name)

    # Fetch Data
    pos = list(PO.objects.filter(program_id=course.program_id, is_active=True).order_by('po_number'))
    psos = list(PSO.objects.filter(program_id=course.program_id, is_active=True).order_by('pso_number'))
    cos = list(CO.objects.filter(course_id=course, is_active=True).order_by('co_number'))
    
    po_count = len(pos)
    pso_count = len(psos)
    
    # Robust AY Matching
    ay_clean = academic_year.replace(' ', '') if academic_year else ""
    ay_spaced = ay_clean.replace('-', ' - ')
    ay_query = models.Q(academic_year__icontains=academic_year) | models.Q(academic_year__icontains=ay_clean) | models.Q(academic_year__icontains=ay_spaced)

    # 2. Matrix Headers
    matrix_start_row = next_row + 2
    
    ws.cell(row=matrix_start_row, column=1, value="CO No.").border = get_border()
    ws.cell(row=matrix_start_row, column=1).font = Font(bold=True)
    ws.cell(row=matrix_start_row, column=1).fill = PatternFill(start_color=CO_BLUE, end_color=CO_BLUE, fill_type="solid")

    # POs Group Header
    if po_count > 0:
        ws.merge_cells(start_row=matrix_start_row, start_column=2, end_row=matrix_start_row, end_column=2+po_count-1)
        ws.cell(row=matrix_start_row, column=2, value="POs").border = get_border()
        ws.cell(row=matrix_start_row, column=2).fill = PatternFill(start_color=MAPPING_ORANGE, end_color=MAPPING_ORANGE, fill_type="solid")
        ws.cell(row=matrix_start_row, column=2).alignment = Alignment(horizontal="center")
        ws.cell(row=matrix_start_row, column=2).font = Font(bold=True)

    # PSOs Group Header
    if pso_count > 0:
        ws.merge_cells(start_row=matrix_start_row, start_column=2+po_count, end_row=matrix_start_row, end_column=2+po_count+pso_count-1)
        ws.cell(row=matrix_start_row, column=2+po_count, value="PSOs").border = get_border()
        ws.cell(row=matrix_start_row, column=2+po_count).fill = PatternFill(start_color=MAPPING_ORANGE, end_color=MAPPING_ORANGE, fill_type="solid")
        ws.cell(row=matrix_start_row, column=2+po_count).alignment = Alignment(horizontal="center")
        ws.cell(row=matrix_start_row, column=2+po_count).font = Font(bold=True)

    # Individual Headers
    col_idx = 2
    for po in pos:
        c = ws.cell(row=matrix_start_row+1, column=col_idx, value=po.po_number)
        c.border = get_border()
        c.font = Font(bold=True)
        c.fill = PatternFill(start_color=CO_BLUE, end_color=CO_BLUE, fill_type="solid")
        col_idx += 1
        
    for pso in psos:
        c = ws.cell(row=matrix_start_row+1, column=col_idx, value=pso.pso_number)
        c.border = get_border()
        c.font = Font(bold=True)
        c.fill = PatternFill(start_color=CO_BLUE, end_color=CO_BLUE, fill_type="solid")
        col_idx += 1

    # 3. Matrix Data & Calculation Prep
    current_row = matrix_start_row + 2
    mapping_data = [] 
    
    for co in cos:
        row_data = [] # [weight1, weight2...]
        ws.cell(row=current_row, column=1, value=co.co_number).border = get_border()
        ws.cell(row=current_row, column=1).font = Font(bold=True)
        ws.cell(row=current_row, column=1).fill = PatternFill(start_color=CO_BLUE, end_color=CO_BLUE, fill_type="solid")
        
        col_idx = 2
        # PO Mappings
        for po in pos:
            m = COPOMapping.objects.filter(co_id=co, po_id=po).first()
            val = m.weightage if m else 0
            row_data.append(val)
            ws.cell(row=current_row, column=col_idx, value=val if val else "").border = get_border()
            ws.cell(row=current_row, column=col_idx).font = Font(bold=True)
            col_idx += 1
            
        # PSO Mappings
        for pso in psos:
            m = COPSOMapping.objects.filter(co_id=co, pso_id=pso).first()
            val = m.weightage if m else 0
            row_data.append(val)
            ws.cell(row=current_row, column=col_idx, value=val if val else "").border = get_border()
            ws.cell(row=current_row, column=col_idx).font = Font(bold=True)
            col_idx += 1
            
        mapping_data.append(row_data)
        current_row += 1
        
    # Average Row
    ws.cell(row=current_row, column=1, value="Average").border = get_border()
    ws.cell(row=current_row, column=1).font = Font(bold=True)
    
    col_sums = [0] * (po_count + pso_count)
    col_counts = [0] * (po_count + pso_count)
    
    for r_data in mapping_data:
        for i, val in enumerate(r_data):
            if val:
                col_sums[i] += val
                col_counts[i] += 1
                
    for i in range(po_count + pso_count):
        avg = col_sums[i] / col_counts[i] if col_counts[i] > 0 else 0
        ws.cell(row=current_row, column=2+i, value=round(avg, 2) if avg else "-").border = get_border()
        ws.cell(row=current_row, column=2+i).font = Font(bold=True)

    matrix_end_row = current_row
    
    # 4. CO Descriptions (Right Side) - Start Column 13 (M)
    desc_start_col = 13
    ws.merge_cells(start_row=matrix_start_row, start_column=desc_start_col, end_row=matrix_start_row, end_column=desc_start_col+1)
    ws.cell(row=matrix_start_row, column=desc_start_col, value="Course outcomes").border = get_border()
    ws.cell(row=matrix_start_row, column=desc_start_col).font = header_font
    
    # Sub headers
    ws.cell(row=matrix_start_row+1, column=desc_start_col, value="CO No.").border = get_border()
    ws.cell(row=matrix_start_row+1, column=desc_start_col+1, value="Description").border = get_border()
    
    desc_row = matrix_start_row + 2
    for co in cos:
        ws.cell(row=desc_row, column=desc_start_col, value=co.co_number).border = get_border()
        ws.cell(row=desc_row, column=desc_start_col+1, value=co.description).border = get_border()
        desc_row += 1
        
    # 5. Direct CO Attainment Table (Bottom Left)
    att_start_row = matrix_end_row + 4
    
    # Header
    ws.merge_cells(start_row=att_start_row, start_column=1, end_row=att_start_row, end_column=2)
    ws.cell(row=att_start_row, column=1, value="Direct CO Attainment").fill = yellow_fill
    ws.cell(row=att_start_row, column=1).font = red_font
    ws.cell(row=att_start_row, column=1).alignment = center_align
    ws.cell(row=att_start_row, column=1).border = get_border()
    
    ws.cell(row=att_start_row+1, column=1, value="CO").border = get_border()
    ws.cell(row=att_start_row+1, column=1).fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    ws.cell(row=att_start_row+1, column=2, value="Attainment").border = get_border()
    ws.cell(row=att_start_row+1, column=2).fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    
    curr_att_row = att_start_row + 2
    co_attainments = []
    
    for co in cos:
        att = COAttainment.objects.filter(ay_query, co_id=co).first()
        val = att.overall_attainment if att else 0
        co_attainments.append(val)
        
        ws.cell(row=curr_att_row, column=1, value=co.co_number).border = get_border()
        ws.cell(row=curr_att_row, column=2, value=val).border = get_border()
        curr_att_row += 1
        
    # Average
    avg_att = sum(co_attainments) / len(co_attainments) if co_attainments else 0
    ws.cell(row=curr_att_row, column=1, value="Average").border = get_border()
    ws.cell(row=curr_att_row, column=1).font = Font(bold=True)
    ws.cell(row=curr_att_row, column=1).fill = PatternFill(start_color=MAPPING_ORANGE, end_color=MAPPING_ORANGE, fill_type="solid")
    
    ws.cell(row=curr_att_row, column=2, value=round(avg_att, 2)).border = get_border()
    ws.cell(row=curr_att_row, column=2).font = Font(bold=True)
    ws.cell(row=curr_att_row, column=2).fill = PatternFill(start_color=MAPPING_ORANGE, end_color=MAPPING_ORANGE, fill_type="solid")

    # 6. Direct PO/PSO Attainment Table (Bottom Right aligned)
    # Align header with PO columns
    po_att_header_row = curr_att_row + 2
    
    ws.merge_cells(start_row=po_att_header_row, start_column=2, end_row=po_att_header_row, end_column=max(2+po_count+pso_count-1, 2))
    ws.cell(row=po_att_header_row, column=2, value="Direct POs & PSOs Attainment").fill = yellow_fill
    ws.cell(row=po_att_header_row, column=2).font = red_font
    ws.cell(row=po_att_header_row, column=2).alignment = center_align
    ws.cell(row=po_att_header_row, column=2).border = get_border()
    
    # Subheaders (PO1...PSO3)
    col_idx = 2
    for po in pos:
        ws.cell(row=po_att_header_row+1, column=col_idx, value=po.po_number).border = get_border()
        ws.cell(row=po_att_header_row+1, column=col_idx).font = Font(bold=True)
        ws.cell(row=po_att_header_row+1, column=col_idx).alignment = center_align
        col_idx += 1
    for pso in psos:
        ws.cell(row=po_att_header_row+1, column=col_idx, value=pso.pso_number).border = get_border()
        ws.cell(row=po_att_header_row+1, column=col_idx).font = Font(bold=True)
        ws.cell(row=po_att_header_row+1, column=col_idx).alignment = center_align
        col_idx += 1
        
    # Data Row
    data_row = po_att_header_row + 2
    col_idx = 2
    
    # Calculate PO Attainment: Sum(CO_Att * Weight) / Sum(Weights)
    for i in range(po_count + pso_count):
        numerator = 0
        denominator = 0
        for co_idx, co_att in enumerate(co_attainments):
            weight = mapping_data[co_idx][i]
            if weight:
                numerator += co_att * weight
                denominator += weight
        
        final_att = numerator / denominator if denominator > 0 else 0
        ws.cell(row=data_row, column=col_idx, value=round(final_att, 2)).border = get_border()
        ws.cell(row=data_row, column=col_idx).font = Font(bold=True)
        ws.cell(row=data_row, column=col_idx).alignment = center_align
        col_idx += 1

    # Formatting adjustments
    # Set Column Widths
    ws.column_dimensions['A'].width = 12 # CO Col (Matrix)
    for i in range(2, 2 + po_count + pso_count):
        col_letter = get_column_letter(i)
        ws.column_dimensions[col_letter].width = 6 # PO/PSO Cols
        
    desc_start_letter = get_column_letter(desc_start_col)
    desc_text_letter = get_column_letter(desc_start_col+1)
    ws.column_dimensions[desc_start_letter].width = 12 # CO Col (Desc)
    ws.column_dimensions[desc_text_letter].width = 60 # Description Text
    
    # Set Row Heights for better spacing
    ws.row_dimensions[1].height = 30 # Main Title
    ws.row_dimensions[2].height = 25 # College Name
    ws.row_dimensions[3].height = 20 # Program Name

def generate_cis_report(course_id, academic_year=None, batch_id=None):
    try:
        course = Course.objects.get(pk=course_id)
    except (Course.DoesNotExist, ValueError):
        return None

    # Robust AY Matching: Check for both spaced and unspaced versions
    ay_val = academic_year or (AcademicSetup.objects.first().academic_year if AcademicSetup.objects.exists() else "2024-25")
    ay_clean = ay_val.replace(' ', '')
    ay_spaced = ay_clean.replace('-', ' - ')
    ay_query = models.Q(academic_year__icontains=ay_val) | models.Q(academic_year__icontains=ay_clean) | models.Q(academic_year__icontains=ay_spaced)

    assignment = FacultyCourseAssignment.objects.filter(
        ay_query,
        course_id=course
    ).first()
    
    # Robust faculty fetching: try normalized AY, then fall back to course's last active assignment
    if not assignment:
        assignment = FacultyCourseAssignment.objects.filter(course_id=course, is_active=True).first()
        
    faculty_name = assignment.faculty_id.name if assignment and assignment.faculty_id else "Not Assigned"

    # Robust student fetching: prioritize students who have marks in this course
    # This prevents blank past reports after students are promoted to a new semester
    from assessments.models import MarksEntry
    
    # Use the robust ay_query for assessment matching to find students
    # We prefix the query with assessment_id__ to filter MarksEntry based on Assessment's AY
    ay_query_marks = models.Q(assessment_id__academic_year=ay_val) | models.Q(assessment_id__academic_year=ay_clean) | models.Q(assessment_id__academic_year=ay_spaced)
    
    marks_students = MarksEntry.objects.filter(
        assessment_id__course_id=course
    ).filter(ay_query_marks).values_list('student_id', flat=True).distinct()
    
    if marks_students.exists():
        students = list(Student.objects.filter(student_id__in=marks_students))
    else:
        # Fallback to current enrollment logic if no marks found yet
        # We ignore semester when batch_id is present to account for promoted students
        student_filters = {
            'program_id': course.program_id,
            'is_active': True
        }
        
        # Only use semester if we don't have a batch to narrow it down
        if not batch_id or batch_id == 'All':
            student_filters['semester'] = course.semester
        
        if batch_id and batch_id != 'All':
            if isinstance(batch_id, str) and '-' in batch_id:
                from academics.models import Batch
                year = batch_id.split('-')[0].strip()
                batch = Batch.objects.filter(batch_year=year).first()
                if batch:
                    student_filters['batch_id'] = batch.batch_id
            else:
                student_filters['batch_id'] = batch_id

        students = list(Student.objects.filter(**student_filters))
        if not students:
            student_filters.pop('semester', None)
            if course.class_year:
                student_filters['class_year'] = course.class_year
            students = list(Student.objects.filter(**student_filters))

    students.sort(key=lambda x: natural_sort_key(x.roll_no or ""))

    wb = openpyxl.Workbook()
    wb.remove(wb.active) 
    
    # Final Sheet Order
    create_cis_analysis_sheet(wb, course, academic_year, faculty_name, 0)
    create_all_combine_sheet(wb, course, academic_year, faculty_name, 1)
    create_ct_sheet(wb, 1, course, academic_year, students, faculty_name, 2)
    create_ct_sheet(wb, 2, course, academic_year, students, faculty_name, 3)
    create_fa_pr_sheet(wb, course, academic_year, students, faculty_name, 4)
    create_sla_sheet(wb, course, academic_year, students, faculty_name, 5)
    create_marks_sheet(wb, "SA-TH", "SA_TH", course, academic_year, students, faculty_name, 6)
    create_marks_sheet(wb, "SA-PR", "SA_PR", course, academic_year, students, faculty_name, 7)
    create_ces_sheet(wb, course, academic_year, students, faculty_name, 8)
    create_mapping_sheet(wb, course, academic_year, faculty_name, 9)

    return wb
